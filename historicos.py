"""Lectura de informes históricos para la comparativa del portal.

Un "histórico" es un informe de una inspección anterior del mismo tramo (de
PCC o de otro contratista). Como todos vienen en la plantilla del contrato, se
leen por encabezado y no por posición fija:

  - cabecera (col A = etiqueta, col C = valor): Fecha, Gasoducto, Tramo, Inspector
  - tabla de potenciales: ABSCISADO, ON [mV], OFF [mV], latitud/longitud

En **DCVG** el informe no guarda un perfil de potenciales sino defectos de
recubrimiento: la hoja 'Inspección DCVG' se lee por etiqueta de encabezado
(la plantilla del contratista pone la severidad %IR en UNA columna y la de PCC
en tres — AA/CA/CC) y la metadata sale de la hoja 'Informe'.

El resultado se guarda como **CSV liviano** (unas decenas de KB por tramo, en
vez de los ~2 MB del .xlsx) y se publica en Supabase, que es de donde el
dashboard y su PDF leen la comparación. Los CSV son la copia auditable local:
NO se suben al repo (es público).

Uso desde el cargador: `cargar_historico.py <informe.xlsx>`.
"""
import csv
import datetime
import os
import re
import unicodedata

import openpyxl

MESES = ['', 'Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep',
         'Oct', 'Nov', 'Dic']

# Hoja de datos según el tipo de inspección
HOJAS = (('Potenciales CIPS', 'CIPS'), ('Potenciales PAP', 'PAP'))
HOJA_DCVG = 'Inspección DCVG'

CAMPOS_CSV = ('abscisa', 'on', 'off', 'lat', 'lon', 'fecha')
CAMPOS_CSV_DCVG = ('clase', 'abscisa', 'referencia', 'on', 'off', 'caracter',
                   'ol_re', 'p_re', 'severidad_pct', 'clasificacion',
                   'profundidad', 'resistividad', 'lat', 'lon')

# Umbrales de clasificación de severidad (%IR), iguales a generator/db.
CLASES_SEVERIDAD = ((15, 'Muy Pequeño'), (35, 'Pequeño'), (60, 'Mediano'))


def _txt(v):
    t = ''.join(c for c in unicodedata.normalize('NFD', str(v or ''))
                if unicodedata.category(c) != 'Mn')
    return re.sub(r'\s+', ' ', t).strip().lower()


def _num(v):
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(',', '.'))
    except (TypeError, ValueError):
        return None


def periodo_legible(fecha):
    """'30/11/2023 AL 01/12/2023' -> 'Nov 2023'. '' si no se entiende."""
    if isinstance(fecha, (datetime.datetime, datetime.date)):
        return f"{MESES[fecha.month]} {fecha.year}"
    txt = str(fecha or '').strip()
    if not txt:
        return ''
    # OJO: primero el formato ISO; si no, '2024-03-12' entra por la regla
    # dd/mm/aaaa leyendo '24-03-12' y devuelve el año equivocado.
    m = re.search(r'\b(\d{4})[/\-.](\d{1,2})[/\-.](\d{1,2})\b', txt)
    if m:                                   # aaaa-mm-dd
        a, mes = int(m.group(1)), int(m.group(2))
        if 1 <= mes <= 12:
            return f"{MESES[mes]} {a}"
    m = re.search(r'\b(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{2,4})\b', txt)
    if m:                                   # dd/mm/aaaa
        mes, a = int(m.group(2)), int(m.group(3))
        a = a + 2000 if a < 100 else a
        if 1 <= mes <= 12:
            return f"{MESES[mes]} {a}"
    return ''


def datos_del_nombre(nombre):
    """Lo que se puede deducir del nombre codificado del archivo:
    CIPS_REP_R_DOR_11_23_1300006811_551003090_TEL_Rev0.xlsx"""
    partes = os.path.splitext(os.path.basename(nombre))[0].split('_')
    out = {}
    if partes:
        out['tipo'] = partes[0].upper()
    m = re.match(r'^[A-Z]+_(?:REP|PPM)_([RTLA])_([A-Z0-9]+)_(\d{2})_(\d{2})_'
                 r'(\d+)_(\d+)_', os.path.basename(nombre).upper())
    if m:
        out['letra'], out['sigla'] = m.group(1), m.group(2)
        mes, anio = int(m.group(3)), int(m.group(4))
        if 1 <= mes <= 12:
            out['periodo'] = f"{MESES[mes]} {2000 + anio}"
        out['ot'], out['contrato'] = m.group(5), m.group(6)
    return out


def _cabecera(ws):
    """{etiqueta_normalizada: valor} de las filas 5-9 (col A = etiqueta,
    col C = valor; col K/S = etiquetas de las otras dos columnas)."""
    out = {}
    for r in range(4, 10):
        for c_et, c_val in ((1, 3), (11, 13), (19, 21)):
            et = _txt(ws.cell(row=r, column=c_et).value)
            val = ws.cell(row=r, column=c_val).value
            if et and val not in (None, ''):
                out[et] = val
    return out


def _fila_encabezado_tabla(ws, hasta=20):
    """Fila donde está 'ABSCISADO' (los datos empiezan 2 filas después)."""
    for r in range(1, hasta + 1):
        for c in range(1, 6):
            if _txt(ws.cell(row=r, column=c).value).startswith('abscisado'):
                return r
    return 10


def _cols_potencial(ws, fila_enc):
    """Columnas de ON y OFF del bloque 'POTENCIAL NEGATIVO 1' (el medido),
    leídas de la fila de subencabezados."""
    ini = None
    for c in range(1, 40):
        if 'potencial negativo 1' in _txt(ws.cell(row=fila_enc, column=c).value):
            ini = c
            break
    if ini is None:
        return (5, 6)                      # respaldo: E/F
    on = off = None
    for c in range(ini, min(ini + 6, 40)):
        sub = _txt(ws.cell(row=fila_enc + 1, column=c).value)
        if sub.startswith('on') and on is None:
            on = c
        elif sub.startswith('off') and off is None:
            off = c
    return (on or ini, off or (ini + 1))


# ── DCVG ─────────────────────────────────────────────────────────────────────
# Etiquetas del encabezado de 'Inspección DCVG'. Se leen por texto y no por
# posición porque cada contratista mueve las columnas: TELMACOM pone la
# severidad %IR en una sola columna y PCC la parte en tres (AA/CA/CC).
_ETIQUETAS_DCVG = (
    ('item', 'item'),
    ('referencia', 'referencias'),
    ('distancia', 'distancia tramo'),
    ('abscisa', 'abscisa'),
    ('lat', 'latitud'),
    ('lon', 'longitud'),
    ('altitud', 'altitud'),
    ('forma', 'forma'),
    ('caracter', 'caracter'),
    ('ol_re', 'ol/re'),
    ('potencial', 'potencial estructura-suelo'),
    ('pulso', 'pulso'),
    ('p_re', 'p/re'),
    ('profundidad', 'profundidad'),
    ('clasificacion', 'severidad [clasificacion'),
    ('severidad', 'severidad [%ir'),
    ('resistividad', 'resistividad'),
    ('observaciones', 'observaciones'),
)


def clasificar_severidad(pct):
    """%IR (0-100) -> 'Muy Pequeño' | 'Pequeño' | 'Mediano' | 'Grande'."""
    if pct is None:
        return None
    for tope, nombre in CLASES_SEVERIDAD:
        if pct <= tope:
            return nombre
    return 'Grande'


def _normalizar_clasificacion(v, pct):
    """La columna del informe a veces trae '-', '#DIV/0!' o el texto en otra
    caja. Si no se entiende, se recalcula del %IR."""
    t = _txt(v)
    for _, nombre in CLASES_SEVERIDAD:
        if t == _txt(nombre):
            return nombre
    if t == 'grande':
        return 'Grande'
    return clasificar_severidad(pct)


def _normalizar_caracter(v):
    """'C-A' / 'Catódico-Anódico' -> 'CA'. Vacío si no se entiende."""
    t = _txt(v).replace(' ', '')
    if not t or t in ('-', 'on-off'):
        return ''
    if t.upper() in ('AA', 'CA', 'CC', 'AC'):       # ya viene compacto
        return t.upper()
    letras = [p[0].upper() for p in re.split(r'[-/]', t) if p]
    par = ''.join(letras[:2])
    return par if par in ('AA', 'CA', 'CC', 'AC') else ''


def _meta_informe(wb):
    """{etiqueta_normalizada: valor} de la hoja 'Informe' (o de la primera hoja
    que tenga la cabecera). La etiqueta va en una celda y el valor en la primera
    celda con contenido a su derecha, porque cada plantilla usa otra columna
    (G/U/AC en DCVG, G/V/AF en PAP-CIPS)."""
    if 'Informe' not in wb.sheetnames:
        return {}
    ws = wb['Informe']
    out = {}
    for r in range(3, 15):
        for c in range(1, 34):
            et = _txt(ws.cell(row=r, column=c).value)
            if not et or len(et) > 40:
                continue
            for c2 in range(c + 1, min(c + 13, 40)):
                v = ws.cell(row=r, column=c2).value
                if v in (None, ''):
                    continue
                if not _txt(v):                     # celda numérica/fecha: sirve
                    out.setdefault(et, v)
                    break
                out.setdefault(et, v)
                break
    return out


def _mapa_columnas_dcvg(ws, fila_enc):
    """{campo: (col_ini, col_fin)} — cada etiqueta manda hasta donde empieza la
    siguiente, para que 'SEVERIDAD [%IR]' cubra las tres columnas AA/CA/CC."""
    encontrados = []
    for c in range(1, 40):
        t = _txt(ws.cell(row=fila_enc, column=c).value)
        if not t:
            continue
        for campo, etiqueta in _ETIQUETAS_DCVG:
            if t.startswith(etiqueta):
                encontrados.append((c, campo))
                break
    encontrados.sort()
    mapa = {}
    for i, (c, campo) in enumerate(encontrados):
        fin = encontrados[i + 1][0] - 1 if i + 1 < len(encontrados) else c
        mapa.setdefault(campo, (c, max(c, fin)))
    # ON / OFF viven en el subencabezado, dentro del bloque de potencial
    ini, fin = mapa.get('potencial', (0, -1))
    for c in range(ini, fin + 1):
        sub = _txt(ws.cell(row=fila_enc + 1, column=c).value)
        if sub.startswith('on') and 'on' not in mapa:
            mapa['on'] = (c, c)
        elif sub.startswith('off') and 'off' not in mapa:
            mapa['off'] = (c, c)
    return mapa


def _fila_encabezado_dcvg(ws, hasta=20):
    for r in range(1, hasta + 1):
        for c in range(1, 8):
            if _txt(ws.cell(row=r, column=c).value).startswith('referencias'):
                return r
    return 6


def _leer_dcvg(wb, ruta):
    """Postes y defectos de la hoja 'Inspección DCVG'."""
    ws = wb[HOJA_DCVG]
    fila_enc = _fila_encabezado_dcvg(ws)
    mapa = _mapa_columnas_dcvg(ws, fila_enc)

    def celda(fila, campo, numerico=False):
        rango = mapa.get(campo)
        if not rango:
            return None
        for c in range(rango[0], rango[1] + 1):
            v = fila[c - 1] if c - 1 < len(fila) else None
            if v in (None, ''):
                continue
            return _num(v) if numerico else v
        return None

    puntos = []
    for fila in ws.iter_rows(min_row=fila_enc + 2, values_only=True):
        if any(_txt(x).startswith(('elaboro', 'reviso', 'aprobo')) for x in fila
               if isinstance(x, str)):
            break
        absc = celda(fila, 'abscisa', True)
        if absc is None:
            continue
        referencia = str(celda(fila, 'referencia') or '').strip()
        on, off = celda(fila, 'on', True), celda(fila, 'off', True)
        es_defecto = 'defecto' in _txt(referencia)
        if not es_defecto and on is None and off is None:
            continue                       # referencia sin lectura: no aporta
        p = {'clase': 'defecto' if es_defecto else 'poste',
             'abscisa': absc, 'referencia': referencia,
             'lat': celda(fila, 'lat', True), 'lon': celda(fila, 'lon', True)}
        if es_defecto:
            sev = celda(fila, 'severidad', True)
            # el informe guarda la fracción (0,111) y el portal usa % (11,1)
            pct = round(sev * 100, 2) if sev is not None and sev <= 1.5 else sev
            p.update({
                'caracter': _normalizar_caracter(celda(fila, 'caracter')),
                'ol_re': celda(fila, 'ol_re', True),
                'p_re': celda(fila, 'p_re', True),
                'severidad_pct': pct,
                'clasificacion': _normalizar_clasificacion(
                    celda(fila, 'clasificacion'), pct),
                'profundidad': celda(fila, 'profundidad', True),
                'resistividad': celda(fila, 'resistividad', True)})
        else:
            p.update({'on': on, 'off': off})
        puntos.append(p)
    return puntos


def resumen_dcvg(puntos):
    """Indicadores del histórico DCVG: cuántos defectos, de qué severidad y
    cada cuánto (densidad por km), que es lo comparable entre inspecciones."""
    defectos = [p for p in puntos if p.get('clase') == 'defecto']
    postes = [p for p in puntos if p.get('clase') == 'poste']
    conteo = {'Muy Pequeño': 0, 'Pequeño': 0, 'Mediano': 0, 'Grande': 0}
    for d in defectos:
        if d.get('clasificacion') in conteo:
            conteo[d['clasificacion']] += 1
    abscisas = [p['abscisa'] for p in puntos
                if isinstance(p.get('abscisa'), (int, float))]
    long_m = (max(abscisas) - min(abscisas)) if abscisas else None
    sev = [d['severidad_pct'] for d in defectos
           if isinstance(d.get('severidad_pct'), (int, float))]
    offs = [p['off'] for p in postes if isinstance(p.get('off'), (int, float))]
    return {
        'n': len(puntos), 'n_defectos': len(defectos), 'n_postes': len(postes),
        'por_clasificacion': conteo,
        'n_criticos': conteo['Mediano'] + conteo['Grande'],
        'long_m': long_m,
        'densidad_km': (round(len(defectos) / (long_m / 1000), 2)
                        if long_m else None),
        'prom_severidad': round(sum(sev) / len(sev), 2) if sev else None,
        'max_severidad': round(max(sev), 2) if sev else None,
        'prom_off': round(sum(offs) / len(offs), 1) if offs else None,
    }


def leer_historico(ruta):
    """Lee un informe de inspección anterior. Devuelve
    {tramo, tipo, periodo, fecha, gasoducto, contratista, inspector, fuente,
     puntos: [{abscisa, on, off, lat, lon, fecha}], resumen}"""
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    hoja = tipo = None
    for nombre, t in HOJAS:
        if nombre in wb.sheetnames:
            hoja, tipo = nombre, t
            break
    if hoja is None and HOJA_DCVG in wb.sheetnames:
        wb.close()
        # la hoja DCVG se recorre dos veces (encabezado y datos): sin read_only
        return _historico_dcvg(openpyxl.load_workbook(ruta, data_only=True), ruta)
    if hoja is None:
        wb.close()
        raise ValueError(f"{os.path.basename(ruta)}: no tiene hoja de datos "
                         f"(se esperaba 'Potenciales CIPS', 'Potenciales PAP' "
                         f"o '{HOJA_DCVG}')")
    ws = wb[hoja]
    cab = _cabecera(ws)

    def g(*etiquetas):
        for e in etiquetas:
            for k, v in cab.items():
                if k.startswith(_txt(e)):
                    return v
        return ''

    fila_enc = _fila_encabezado_tabla(ws)
    c_on, c_off = _cols_potencial(ws, fila_enc)
    puntos = []
    for fila in ws.iter_rows(min_row=fila_enc + 2, values_only=True):
        def v(col):
            return fila[col - 1] if col - 1 < len(fila) else None
        # el bloque de firmas cierra la tabla
        if any(_txt(x).startswith(('elaboro', 'reviso', 'aprobo')) for x in fila
               if isinstance(x, str)):
            break
        absc, on, off = _num(v(2)), _num(v(c_on)), _num(v(c_off))
        if absc is None or off is None:
            continue
        fecha_pt = v(3)
        puntos.append({
            'abscisa': absc, 'on': on, 'off': off,
            'lat': _num(v(19)), 'lon': _num(v(20)),
            'fecha': fecha_pt.strftime('%Y-%m-%d')
                     if isinstance(fecha_pt, (datetime.datetime, datetime.date))
                     else (str(fecha_pt).split()[0] if fecha_pt else '')})
    wb.close()

    fecha = g('fecha')
    del_nombre = datos_del_nombre(ruta)
    hist = {
        'tramo': str(g('tramo') or '').strip(),
        'tipo': tipo,
        'periodo': periodo_legible(fecha) or del_nombre.get('periodo', ''),
        'fecha': str(fecha or '').strip(),
        'gasoducto': str(g('gasoducto') or '').strip(),
        'contratista': str(g('contratista') or '').strip(),
        'inspector': str(g('inspector') or '').strip(),
        'fuente': os.path.basename(ruta),
        'puntos': puntos,
    }
    hist['resumen'] = resumen(puntos)
    return hist


def _historico_dcvg(wb, ruta):
    """Histórico de una inspección DCVG: la metadata sale de la hoja 'Informe'
    (la hoja de datos no tiene cabecera) y los puntos, de 'Inspección DCVG'."""
    meta = _meta_informe(wb)

    def g(*etiquetas):
        for e in etiquetas:                        # primero coincidencia exacta
            if _txt(e) in meta:
                return meta[_txt(e)]
        for e in etiquetas:
            for k, v in meta.items():
                if k.startswith(_txt(e)):
                    return v
        return ''

    puntos = _leer_dcvg(wb, ruta)
    wb.close()
    fecha = g('fecha')
    del_nombre = datos_del_nombre(ruta)
    hist = {
        'tramo': str(g('tramo') or '').strip(),
        'tipo': 'DCVG',
        'periodo': periodo_legible(fecha) or del_nombre.get('periodo', ''),
        'fecha': str(fecha or '').strip(),
        'gasoducto': str(g('gasoducto') or '').strip(),
        'contratista': str(g('contratista') or '').strip(),
        'inspector': str(g('inspector') or '').strip(),
        'fuente': os.path.basename(ruta),
        'puntos': puntos,
    }
    hist['resumen'] = resumen_dcvg(puntos)
    return hist


def resumen(puntos):
    """Indicadores del histórico (mismo criterio que el portal: OFF ≤ -850 mV)."""
    offs = [p['off'] for p in puntos if isinstance(p.get('off'), (int, float))]
    if not offs:
        return {'n': 0, 'fuera': None, 'pct_prot': None, 'prom_off': None,
                'min_off': None, 'max_off': None}
    fuera = sum(1 for o in offs if o > -850)
    return {'n': len(offs), 'fuera': fuera,
            'pct_prot': round(100 * (len(offs) - fuera) / len(offs), 2),
            'prom_off': round(sum(offs) / len(offs), 1),
            'min_off': round(min(offs), 1), 'max_off': round(max(offs), 1)}


# ── CSV liviano ──────────────────────────────────────────────────────────────

def a_csv(hist, ruta):
    """Guarda el histórico como CSV: una cabecera comentada con los metadatos
    y una fila por punto."""
    with open(ruta, 'w', newline='', encoding='utf-8') as f:
        for k in ('tramo', 'tipo', 'periodo', 'fecha', 'gasoducto',
                  'contratista', 'fuente'):
            f.write(f"# {k}: {hist.get(k, '')}\n")
        campos = (CAMPOS_CSV_DCVG if hist.get('tipo') == 'DCVG'
                  else CAMPOS_CSV)
        w = csv.DictWriter(f, fieldnames=campos, extrasaction='ignore')
        w.writeheader()
        w.writerows(hist['puntos'])
    return ruta


def desde_csv(ruta):
    """Reconstruye el histórico desde el CSV generado por `a_csv`."""
    meta, filas = {}, []
    with open(ruta, encoding='utf-8') as f:
        lineas = f.readlines()
    datos = []
    for ln in lineas:
        if ln.startswith('#'):
            k, _, v = ln[1:].partition(':')
            meta[k.strip()] = v.strip()
        else:
            datos.append(ln)
    numericos = ('abscisa', 'on', 'off', 'lat', 'lon', 'ol_re', 'p_re',
                 'severidad_pct', 'profundidad', 'resistividad')
    for fila in csv.DictReader(datos):
        filas.append({k: (_num(v) if k in numericos else (v or ''))
                      for k, v in fila.items()})
    meta['puntos'] = filas
    meta['resumen'] = (resumen_dcvg(filas) if meta.get('tipo') == 'DCVG'
                       else resumen(filas))
    return meta
