"""Codificación de los nombres de archivo de los entregables (TGI).

Formato acordado con TGI:

    tipo inspección _ REP|PPM _ R|T _ sigla de la línea _ mes _ año _ OT _
    # de contrato _ PCC _ Rev.A

Ejemplo real: ``DCVG_REP_R_ARM_03_25_1300013506_551007370_PCC_Rev.A``

- **R/T** y **sigla** salen de `Infraestrutura TGI.xlsx` (columnas `Tipo` y
  `SIGLAS`), buscando el tramo por nombre. Es la misma fuente del autollenado.
- **mes/año** salen de la fecha de la inspección (`info['fecha']`), en
  cualquiera de los formatos que llegan de campo.
- **OT** y **contrato** son los campos de Datos Generales.

Si algún dato falta, la parte se omite (no se dejan '__' en el nombre) y
`faltantes()` dice qué quedó por fuera, para avisarle al ingeniero.
"""
import datetime
import os
import re
import unicodedata

import openpyxl

from generator import resource_path

ARCHIVO_INFRA = 'Infraestrutura TGI.xlsx'

# Primera letra del tipo de línea (columna `Tipo` de Infraestrutura TGI).
_LETRA_TIPO = {'troncal': 'T', 'ramal': 'R', 'aislado': 'A', 'loop': 'L'}
# Prioridad ante tramos duplicados con siglas distintas: manda la línea
# principal sobre el loop (p.ej. 'La Belleza - Vasconia' → VRMB, no BEVV).
_PRIORIDAD_TIPO = {'T': 0, 'R': 0, 'A': 0, 'L': 1, '': 2}

_cache_infra = None


def _sin_tildes(s):
    return ''.join(c for c in unicodedata.normalize('NFD', str(s or ''))
                   if unicodedata.category(c) != 'Mn')


def _norm(s):
    """Nombre de tramo comparable: sin tildes, sin el 'PK 46+265' del final,
    sin espacios de más y en minúsculas. OJO: hay tramos que se LLAMAN
    'PK 7+200 - PK 17+500'; ahí no se recorta nada."""
    t = _sin_tildes(s).lower()
    sin_pk = re.sub(r'\s*\(?\bpk\b\s*\d.*', '', t)   # 'Armenia PK 46+265' -> 'armenia'
    if sin_pk.strip():
        t = sin_pk
    return re.sub(r'\s+', ' ', t).strip()


def _limpiar(parte):
    """Deja una parte del nombre apta para un archivo: sin tildes, sin espacios
    ni separadores, en mayúsculas."""
    t = _sin_tildes(parte).upper()
    return re.sub(r'[^A-Z0-9.+-]', '', t)


def _cargar_infra():
    """[(nombre_normalizado, sigla, letra)] de `Infraestrutura TGI.xlsx`."""
    global _cache_infra
    if _cache_infra is not None:
        return _cache_infra
    filas = []
    try:
        ruta = resource_path(ARCHIVO_INFRA)
        if os.path.exists(ruta):
            wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
            ws = wb['Infra'] if 'Infra' in wb.sheetnames else wb[wb.sheetnames[0]]
            encabezado, idx = None, {}
            for fila in ws.iter_rows(values_only=True):
                if encabezado is None:
                    if fila and any(str(c or '').strip().upper() == 'TRAMOS'
                                    for c in fila):
                        encabezado = fila
                        idx = {str(c or '').strip().upper(): i
                               for i, c in enumerate(fila)}
                    continue
                tramo = fila[idx['TRAMOS']] if 'TRAMOS' in idx else None
                if not tramo or str(tramo).strip().upper() == 'TRAMOS':
                    continue          # la hoja repite el encabezado a media tabla
                sigla = fila[idx['SIGLAS']] if 'SIGLAS' in idx else None
                tipo = fila[idx['TIPO']] if 'TIPO' in idx else None
                filas.append((_norm(tramo), str(sigla or '').strip(),
                              _LETRA_TIPO.get(_norm(tipo), '')))
            wb.close()
    except Exception:
        filas = []
    _cache_infra = filas
    return filas


def sigla_tramo(tramo):
    """(sigla, letra R/T/A, encontrado) del tramo según Infraestrutura TGI.

    Si el tramo no está en el archivo devuelve una sigla de respaldo hecha con
    el propio nombre, para que el informe igual salga con nombre usable.
    """
    objetivo = _norm(tramo)
    if objetivo:
        filas = _cargar_infra()
        exactas = [f for f in filas if f[0] == objetivo and f[1]]
        if exactas:                                   # coincidencia exacta
            _n, sigla, letra = min(
                exactas, key=lambda f: _PRIORIDAD_TIPO.get(f[2], 2))
            return (sigla.upper(), letra, True)
        # coincidencia parcial: solo con nombres de largo razonable, para no
        # pegarle a cualquier cosa
        candidatos = [f for f in filas
                      if f[1] and len(f[0]) >= 4
                      and (objetivo in f[0]
                           or (len(objetivo) >= 4 and f[0] in objetivo))]
        if candidatos:                # el más parecido en largo, línea antes que loop
            nombre, sigla, letra = min(
                candidatos, key=lambda f: (abs(len(f[0]) - len(objetivo)),
                                           _PRIORIDAD_TIPO.get(f[2], 2)))
            return (sigla.upper(), letra, True)
    respaldo = _limpiar(re.sub(r'\s+', '', _norm(tramo)))[:12]
    return (respaldo, '', False)


def mes_anio(fecha):
    """('03', '25') desde la fecha de la inspección; ('', '') si no se entiende."""
    if isinstance(fecha, (datetime.datetime, datetime.date)):
        return (f"{fecha.month:02d}", f"{fecha.year % 100:02d}")
    txt = str(fecha or '').strip()
    if not txt:
        return ('', '')
    txt = txt.split()[0].replace('.', '-').replace('/', '-')
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d-%m-%y", "%Y-%m-%d", "%m-%d-%Y"):
        try:
            d = datetime.datetime.strptime(txt, fmt)
            return (f"{d.month:02d}", f"{d.year % 100:02d}")
        except ValueError:
            continue
    return ('', '')


def nombre_archivo(info, doc="REP", ext=".xlsx"):
    """Nombre codificado del entregable. `doc` = 'REP' (informe) o 'PPM'."""
    sigla, letra, _ = sigla_tramo(info.get('tramo'))
    mm, aa = mes_anio(info.get('fecha'))
    if not letra:
        # respaldo: el tipo de ducto de Datos Generales (Troncal/Ramal)
        letra = _LETRA_TIPO.get(_norm(info.get('tipo_ducto')), '')
    partes = [
        _limpiar(info.get('tipo_inspeccion') or ''),
        _limpiar(doc),
        letra,
        _limpiar(sigla),
        mm,
        aa,
        _limpiar(info.get('ot') or ''),
        _limpiar(info.get('contrato') or ''),
        "PCC",
        "Rev.A",
    ]
    return "_".join(p for p in partes if p) + (ext or "")


def faltantes(info):
    """Etiquetas de lo que no se pudo poner en el nombre (para avisar)."""
    fuera = []
    if not _limpiar(info.get('tipo_inspeccion') or ''):
        fuera.append('Tipo de inspección')
    sigla, letra, hallado = sigla_tramo(info.get('tramo'))
    if not hallado:
        fuera.append('sigla del tramo (no está en Infraestrutura TGI.xlsx)')
    if not (letra or _LETRA_TIPO.get(_norm(info.get('tipo_ducto')))):
        fuera.append('Tipo Ducto (Troncal/Ramal/Loop/Aislado)')
    if not all(mes_anio(info.get('fecha'))):
        fuera.append('Fecha')
    if not _limpiar(info.get('ot') or ''):
        fuera.append('OT')
    if not _limpiar(info.get('contrato') or ''):
        fuera.append('Contrato')
    return fuera
