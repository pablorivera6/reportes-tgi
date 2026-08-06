"""Genera el entregable de 'Inspección Visual de Interfases' en Excel a partir
de los datos de FastField (info + inspecciones con PK, GPS, observación, fotos).

Diseño limpio (paleta PCC): una fila por FOTO, con las celdas de Ítem/PK/
Coordenadas/Observación combinadas verticalmente por interfase. Las imágenes se
incrustan escaladas a un tamaño uniforme y legible.

Va al ZIP de entrega en 04_Anexos/Inspeccion_Visual_Interfases/.
"""
from __future__ import annotations

import io

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Paleta PCC
_ROJO = "C8102E"
_GRIS = "F2F2F2"
_GRIS_L = "FAFAFA"
_TXT = "1A1A1A"

_THIN = Side(style="thin", color="D0D0D0")
_BORDE = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_FOTO_W = 210          # px
_FOTO_H = 150          # px
_PX2PT = 0.75          # alto de fila (pt) ~ px * 0.75


def _absc(pk):
    """'136300' o '136+300' -> 'K 136+300'. Deja strings ya formateados."""
    s = str(pk or "").strip()
    if not s:
        return ""
    if "+" in s:
        return s if s.upper().startswith("K") else f"K {s}"
    try:
        m = int(float(s))
        return f"K {m // 1000:03d}+{m % 1000:03d}"
    except (TypeError, ValueError):
        return s


def _scale(img_bytes):
    """(w, h) escalados a la caja _FOTO_W x _FOTO_H conservando proporción."""
    try:
        from PIL import Image as PILImage
        with PILImage.open(io.BytesIO(img_bytes)) as im:
            w, h = im.size
    except Exception:
        return _FOTO_W, _FOTO_H
    if not w or not h:
        return _FOTO_W, _FOTO_H
    r = min(_FOTO_W / w, _FOTO_H / h)
    return max(1, int(w * r)), max(1, int(h * r))


def construir_excel(info: dict, inspecciones: list, fotos_map: dict) -> bytes:
    """info: {tramo, fecha, tecnico, contratista, ...}.
    inspecciones: [{pk, lat, lon, observacion, fotos:[filename,...]}].
    fotos_map: {filename: bytes}.
    Devuelve los bytes del .xlsx.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Interfases"
    ws.sheet_view.showGridLines = False

    cols = ["N°", "Abscisa", "Latitud", "Longitud", "Observación",
            "Registro Fotográfico"]
    anchos = [6, 14, 13, 13, 46, 32]
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ncol = len(cols)
    last = get_column_letter(ncol)

    blanco = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
    negro = Font(color=_TXT, size=10, name="Calibri")
    bold = Font(color=_TXT, bold=True, size=10, name="Calibri")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ── Título ────────────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{last}1")
    t = ws["A1"]
    t.value = "INSPECCIÓN VISUAL DE INTERFASES"
    t.fill = PatternFill("solid", fgColor=_ROJO)
    t.font = Font(color="FFFFFF", bold=True, size=15, name="Calibri")
    t.alignment = center
    ws.row_dimensions[1].height = 28
    ws.merge_cells(f"A2:{last}2")
    s = ws["A2"]
    s.value = "PCC Integrity — Protección Catódica de Colombia"
    s.font = Font(color=_ROJO, bold=True, size=9, name="Calibri")
    s.alignment = Alignment(horizontal="center")

    # ── Bloque de información (2 pares label/valor por fila) ──────────────────
    pares = [
        ("Tramo", info.get("tramo") or ""),
        ("Fecha", info.get("fecha") or ""),
        ("Gasoducto", info.get("gasoducto") or ""),
        ("Contrato", info.get("contrato") or ""),
        ("Inspector", info.get("tecnico") or info.get("inspector") or ""),
        ("OT", info.get("ot") or ""),
        ("Contratista", info.get("contratista") or ""),
        ("Tipo", "Inspección Visual de Interfases"),
    ]
    r = 4
    for i in range(0, len(pares), 2):
        for j, (lab, val) in enumerate(pares[i:i + 2]):
            c0 = 1 + j * 3
            lc = ws.cell(row=r, column=c0, value=lab)
            lc.font = bold
            lc.fill = PatternFill("solid", fgColor=_GRIS)
            lc.alignment = left
            lc.border = _BORDE
            ws.merge_cells(start_row=r, start_column=c0 + 1,
                           end_row=r, end_column=c0 + 2 if j == 0 else ncol)
            vc = ws.cell(row=r, column=c0 + 1, value=str(val))
            vc.font = negro
            vc.alignment = left
            vc.border = _BORDE
        r += 1

    # ── Encabezado de la tabla ────────────────────────────────────────────────
    r += 1
    head_row = r
    for i, cname in enumerate(cols, 1):
        cell = ws.cell(row=r, column=i, value=cname)
        cell.fill = PatternFill("solid", fgColor=_ROJO)
        cell.font = blanco
        cell.alignment = center
        cell.border = _BORDE
    ws.row_dimensions[r].height = 22
    r += 1

    # ── Filas: una por foto, celdas de datos combinadas por interfase ────────
    foto_col = ncol
    for idx, insp in enumerate(inspecciones, 1):
        fotos = [f for f in (insp.get("fotos") or []) if fotos_map.get(f)]
        nfilas = max(1, len(fotos))
        r0, r1 = r, r + nfilas - 1

        def _celda(col, val, align=center, zebra=idx % 2 == 0):
            if r1 > r0:
                ws.merge_cells(start_row=r0, start_column=col,
                               end_row=r1, end_column=col)
            cc = ws.cell(row=r0, column=col, value=val)
            cc.font = negro
            cc.alignment = align
            cc.border = _BORDE
            if zebra:
                cc.fill = PatternFill("solid", fgColor=_GRIS_L)
            # bordes en todas las filas combinadas
            for rr in range(r0, r1 + 1):
                ws.cell(row=rr, column=col).border = _BORDE

        _celda(1, idx)
        _celda(2, _absc(insp.get("pk")))
        _celda(3, insp.get("lat"))
        _celda(4, insp.get("lon"))
        _celda(5, insp.get("observacion") or "", align=left)

        if not fotos:
            cc = ws.cell(row=r0, column=foto_col, value="(sin foto)")
            cc.font = Font(color="999999", italic=True, size=9)
            cc.alignment = center
            cc.border = _BORDE
            ws.row_dimensions[r0].height = 20
        else:
            for k, fn in enumerate(fotos):
                rr = r0 + k
                ws.cell(row=rr, column=foto_col).border = _BORDE
                try:
                    b = fotos_map[fn]
                    w, h = _scale(b)
                    xi = XLImage(io.BytesIO(b))
                    xi.width, xi.height = w, h
                    ws.add_image(xi, f"{get_column_letter(foto_col)}{rr}")
                    ws.row_dimensions[rr].height = max(h * _PX2PT + 6, 24)
                except Exception:
                    ws.cell(row=rr, column=foto_col, value=fn).font = negro
        r = r1 + 1

    # ancho de la columna de fotos acorde al tamaño de imagen
    ws.column_dimensions[get_column_letter(foto_col)].width = _FOTO_W / 7.0

    ws.freeze_panes = ws.cell(row=head_row + 1, column=1)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
