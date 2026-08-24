"""Lectura de informes históricos (de PCC o de otro contratista) para la
comparativa del portal.

Los informes vienen en la misma plantilla del contrato, así que se leen por
encabezado: tramo, fecha y contratista de la cabecera, y abscisa/ON/OFF de la
hoja de potenciales. El resultado se guarda como CSV liviano (auditable) y se
publica en Supabase para alimentar el dashboard y su PDF.
"""
import os

import openpyxl
import pytest

from historicos import (leer_historico, a_csv, desde_csv, periodo_legible,
                        datos_del_nombre)

REAL = "/private/tmp/hist_dorada.xlsx"


def _informe_sintetico(tmp_path, filas=5, hoja="Potenciales CIPS"):
    """Réplica mínima de la plantilla CIPS: cabecera + tabla desde la fila 12."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = hoja
    ws["A5"], ws["C5"] = "Fecha", "30/11/2023 AL 01/12/2023"
    ws["A6"], ws["C6"] = "Gasoducto/Estación", "Centro_Oriente"
    ws["A7"], ws["C7"] = "Tramo", "La Dorada"
    ws["A8"], ws["C8"] = "Inspector", "Luis Humberto Ortiz"
    ws["S5"], ws["U5"] = "No de contrato", "551003090"
    ws["S7"], ws["U7"] = "Contratista", "TELMACOM SAS"
    ws["A10"], ws["B10"], ws["C10"] = "ÍTEM", "ABSCISADO", "FECHA"
    ws["E10"] = "POTENCIAL NEGATIVO 1 TGI"
    ws["E11"], ws["F11"] = "ON [mV]", "OFF [mV]"
    for i in range(filas):
        r = 12 + i
        ws.cell(row=r, column=2, value=i * 10)              # B abscisa
        ws.cell(row=r, column=5, value=-2100.0 - i)         # E ON
        ws.cell(row=r, column=6, value=-1125.0 - i)         # F OFF
        ws.cell(row=r, column=19, value=5.44 + i / 1000)    # S lat
        ws.cell(row=r, column=20, value=-74.68)             # T lon
    fin = 12 + filas
    ws.cell(row=fin + 1, column=3, value="ELABORÓ")         # bloque de firmas
    ruta = str(tmp_path / "hist.xlsx")
    wb.save(ruta)
    return ruta


def test_lee_cabecera_y_puntos(tmp_path):
    h = leer_historico(_informe_sintetico(tmp_path))
    assert h["tramo"] == "La Dorada"
    assert h["tipo"] == "CIPS"
    assert h["contratista"] == "TELMACOM SAS"
    assert len(h["puntos"]) == 5
    p = h["puntos"][0]
    assert p["abscisa"] == 0 and p["on"] == -2100.0 and p["off"] == -1125.0
    assert round(p["lat"], 2) == 5.44


def test_no_incluye_el_bloque_de_firmas(tmp_path):
    h = leer_historico(_informe_sintetico(tmp_path, filas=3))
    assert len(h["puntos"]) == 3
    assert all(p["off"] is not None for p in h["puntos"])


def test_periodo_desde_la_fecha(tmp_path):
    h = leer_historico(_informe_sintetico(tmp_path))
    assert h["periodo"] == "Nov 2023"


def test_periodo_legible():
    assert periodo_legible("30/11/2023 AL 01/12/2023") == "Nov 2023"
    assert periodo_legible("2024-03-12") == "Mar 2024"
    assert periodo_legible("") == ""


def test_datos_del_nombre_del_archivo():
    d = datos_del_nombre("CIPS_REP_R_DOR_11_23_1300006811_551003090_TEL_Rev0.xlsx")
    assert d["tipo"] == "CIPS"
    assert d["periodo"] == "Nov 2023"
    assert d["ot"] == "1300006811"
    assert d["contrato"] == "551003090"
    assert d["sigla"] == "DOR"


def test_ida_y_vuelta_por_csv(tmp_path):
    h = leer_historico(_informe_sintetico(tmp_path))
    csv = str(tmp_path / "h.csv")
    a_csv(h, csv)
    assert os.path.getsize(csv) > 0
    h2 = desde_csv(csv)
    assert h2["tramo"] == h["tramo"] and h2["tipo"] == h["tipo"]
    assert h2["periodo"] == h["periodo"]
    assert len(h2["puntos"]) == len(h["puntos"])
    assert h2["puntos"][0]["off"] == h["puntos"][0]["off"]


def test_resumen_de_proteccion(tmp_path):
    h = leer_historico(_informe_sintetico(tmp_path))
    # todos los OFF están por debajo de -850 mV: 100% protegido
    assert h["resumen"]["n"] == 5
    assert h["resumen"]["pct_prot"] == 100.0


@pytest.mark.skipif(not os.path.exists(REAL), reason="informe real no disponible")
def test_informe_real_la_dorada():
    h = leer_historico(REAL)
    assert h["tramo"] == "La Dorada"
    assert h["contratista"] == "TELMACOM SAS"
    assert h["periodo"] == "Nov 2023"
    assert 700 <= len(h["puntos"]) <= 800
    offs = [p["off"] for p in h["puntos"] if p["off"] is not None]
    assert len(offs) == len(h["puntos"])
    assert all(-3000 < o < 0 for o in offs)
