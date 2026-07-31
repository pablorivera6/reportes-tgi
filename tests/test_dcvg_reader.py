"""Reader del FastField DCVG y de Resistividades."""
import os

import openpyxl
import pytest

from dcvg_reader import (parse_pk, parse_coords, caracter_corto,
                         clasificar_severidad, leer_dcvg_fastfield,
                         leer_resistividades_fastfield)

_REF = "/private/tmp/dcvg_ref"


def test_parse_pk():
    assert parse_pk("5+760") == 5760
    assert parse_pk("136+300") == 136300
    assert parse_pk("PK 2+000") == 2000
    assert parse_pk("00+000") == 0
    assert parse_pk("") is None
    assert parse_pk(None) is None


def test_parse_coords():
    assert parse_coords("4.5290577,-75.7461438") == (4.5290577, -75.7461438)
    assert parse_coords("  4.53 , -75.74 ") == (4.53, -75.74)
    assert parse_coords("") == (None, None)
    assert parse_coords(None) == (None, None)


def test_caracter_corto():
    assert caracter_corto("Catódico-Catódico") == "CC"
    assert caracter_corto("Catódico-Anódico") == "CA"
    assert caracter_corto("Anódico-Anódico") == "AA"
    assert caracter_corto("") == ""


def test_clasificar_severidad():
    assert clasificar_severidad(1) == "Muy Pequeño"
    assert clasificar_severidad(15) == "Muy Pequeño"
    assert clasificar_severidad(16) == "Pequeño"
    assert clasificar_severidad(35) == "Pequeño"
    assert clasificar_severidad(36) == "Mediano"
    assert clasificar_severidad(60) == "Mediano"
    assert clasificar_severidad(61) == "Grande"
    assert clasificar_severidad(100) == "Grande"
    assert clasificar_severidad(None) == ""


def _fastfield_sintetico(tmp_path):
    wb = openpyxl.Workbook()
    root = wb.active; root.title = "Root"
    root.append(["Contratista", "Fecha", "Cliente",
                 "Troncal o ramal inspeccionado", "Técnico a cargo"])
    root.append(["PCC", "07-28-2026", "TGI", "ramal Montenegro", "Juan Perez"])
    s5 = wb.create_sheet("subform_5")
    s5.append(["Submission Id", "Tipo de poste", "PK", "ON", "OFF",
               "Voltaje AC", "Resistencia ", "Coordenadas"])
    s5.append(["x", "Poste de potencial", "6+000", -1682, -1169, 0.442, 0.3,
               "4.5309806,-75.7461"])
    s9 = wb.create_sheet("subform_9")
    s9.append(["Submission Id", "Sector", "Ubicación", "PK del defecto (Abscisa)",
               "Forma N", "Forma S", "Forma E", "Forma O", "OL/RE",
               "Profundidad (M)", "Clasificación ", "Forma del defecto",
               "Comentarios", "Caracter de la indicación"])
    s9.append(["x", "Montenegro", "4.5290577,-75.7461", "5+760",
               54.8, 23.8, 26.1, 52.9, 95.7, 190, "Grande", "12",
               "cruce", "Catódico-Catódico"])
    ruta = os.path.join(tmp_path, "dcvg_ff.xlsx")
    wb.save(ruta)
    return ruta


def test_leer_dcvg_fastfield(tmp_path):
    d = leer_dcvg_fastfield(_fastfield_sintetico(tmp_path))
    assert d["meta"]["tecnico"] == "Juan Perez"
    assert d["meta"]["tramo"] == "ramal Montenegro"
    assert len(d["postes"]) == 1
    p = d["postes"][0]
    assert p["pk_m"] == 6000 and p["on"] == -1682 and p["off"] == -1169
    assert round(p["lat"], 4) == 4.5310
    assert len(d["defectos"]) == 1
    df = d["defectos"][0]
    assert df["pk_m"] == 5760
    assert df["forma_n"] == 54.8 and df["forma_e"] == 26.1
    assert df["ol_re"] == 95.7 and df["profundidad"] == 190
    assert df["caracter"] == "CC"
    assert round(df["lat"], 4) == 4.5291


@pytest.mark.skipif(not os.path.isdir(_REF), reason="archivos reales no disponibles")
def test_leer_montenegro_real():
    d = leer_dcvg_fastfield(os.path.join(_REF, "Dcvg Fastfield.xlsx"))
    assert d["postes"], "debe haber postes"
    assert d["defectos"], "debe haber defectos"
    # la mayoría de defectos con abscisa; los sin PK (gaps del técnico) se
    # conservan con pk_m=None para que el generador los omita avisando.
    con_pk = [x for x in d["defectos"] if x["pk_m"] is not None]
    assert len(con_pk) >= len(d["defectos"]) - 1
    assert con_pk[0]["caracter"] in ("AA", "CA", "CC", "")
    r = leer_resistividades_fastfield(os.path.join(_REF, "Resistividades Fastfield.xlsx"))
    assert r and all(x["pk_m"] is not None for x in r)
    assert r[0]["r1"] is not None
