"""Bloque 'PARÁMETROS OPERATIVOS EN RECTIFICADORES' (URPC).

Cada plantilla lo pone en filas y columnas distintas:

              PAP/CIPS        DCVG
  título      B77             B43
  datos       fila 80+        fila 46+
  Disp. V     Q               O
  Disp. I     T               Q
  TAPS        U               S
  NEG 1 [mV]  W               T
  NEG 1 [A]   AC              Z

Como el generador tenía quemadas las de PAP —y el flujo DCVG ni siquiera
llamaba a fill_rectificadores—, los datos de las URPC salían vacíos en el
informe DCVG.
"""
import os

import openpyxl
import pytest

from generator import ReportGenerator, resource_path

RECT = [{'nombre': 'URPC Balboa', 'voltaje_nominal': 30, 'corriente_nominal': 20,
         'ultima_inspeccion': {'vdc_salida': 12, 'idc_salida': 8,
                               'disponibilidad_v': 40, 'disponibilidad_i': 40,
                               'taps': '3-2'},
         'conexion_estructura': {'pot_on': '-1.5', 'pot_off': '-1.1',
                                 'corriente': '5'}}]


def _informe(tmp_path, plantilla, rect=RECT, nombre="i.xlsx"):
    gen = ReportGenerator(resource_path(plantilla))
    gen.fill_rectificadores(rect)
    out = os.path.join(tmp_path, nombre)
    gen.save(out)
    return openpyxl.load_workbook(out)["Informe"], gen


def test_urpc_dcvg_en_sus_columnas(tmp_path):
    ws, _ = _informe(tmp_path, "DCVG_REP.xlsx")
    assert ws["B46"].value == 'URPC Balboa'
    assert ws["E46"].value == 30       # V nominal
    assert ws["H46"].value == 20       # I nominal
    assert ws["K46"].value == 12       # V operacional
    assert ws["N46"].value == 8        # I operacional
    assert ws["O46"].value == 40       # disponibilidad V  (en PAP es Q)
    assert ws["Q46"].value == 40       # disponibilidad I  (en PAP es T)
    assert ws["S46"].value == '3-2'    # TAPS              (en PAP es U)
    assert 'ON: -1.5' in str(ws["T46"].value)   # NEG 1 [mV] (en PAP es W)
    assert ws["Z46"].value == '5'      # NEG 1 [A]         (en PAP es AC)


def test_no_pisa_el_resumen_de_indicaciones(tmp_path):
    """El bloque DCVG solo tiene 7 filas (46-52); no debe invadir la tabla de
    RESUMEN DE INDICACIONES que va justo debajo."""
    muchos = [dict(RECT[0], nombre=f"URPC {i}") for i in range(12)]
    ws, gen = _informe(tmp_path, "DCVG_REP.xlsx", muchos, "muchos.xlsx")
    assert ws["B53"].value == 'RESUMEN DE INDICACIONES'
    assert ws["B54"].value is not None
    assert gen.rect_omitidos > 0       # avisa cuántos no cupieron


def test_pap_mantiene_sus_columnas(tmp_path):
    ws, _ = _informe(tmp_path, "EN BLANCO.xlsx", nombre="pap.xlsx")
    assert ws["B80"].value == 'URPC Balboa'
    assert ws["E80"].value == 30
    assert ws["K80"].value == 12
    assert ws["Q80"].value == 40       # disponibilidad V
    assert ws["T80"].value == 40       # disponibilidad I
    assert ws["U80"].value == '3-2'    # TAPS
    assert 'ON: -1.5' in str(ws["W80"].value)
    assert ws["AC80"].value == '5'
    assert ws["A83"].value == 'CONCLUSIONES'   # no se pisa la sección
