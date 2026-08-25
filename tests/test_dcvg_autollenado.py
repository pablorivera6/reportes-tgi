"""El FastField de DCVG trae el tramo inspeccionado ('Troncal o ramal'), la
fecha y el contratista. Si no se llevan a Datos Generales, el informe sale sin
tramo: encabezado vacío, columna TRAMO de la hoja Hallazgos vacía y el nombre
del archivo sin la sigla de la línea.
"""
import openpyxl
import pytest

from dcvg_reader import info_desde_meta, leer_dcvg_fastfield


def _fastfield(tmp_path):
    wb = openpyxl.Workbook()
    r = wb.active
    r.title = "Root"
    r.append(["Contratista", "Fecha", "Cliente", "Troncal o ramal",
              "Técnico a cargo"])
    r.append(["PCC", "12/03/2025", "TGI", "Ramal Armenia", "Juan Perez"])
    s5 = wb.create_sheet("subform_5")
    s5.append(["Tipo de poste", "PK", "ON", "OFF", "Voltaje AC", "Resistencia",
               "Coordenadas"])
    s5.append(["Poste de potencial", "5+000", -1600, -1100, 0.4, 0.3, "4.52,-75.74"])
    ruta = str(tmp_path / "ff.xlsx")
    wb.save(ruta)
    return ruta


def test_info_desde_meta_trae_el_tramo(tmp_path):
    meta = leer_dcvg_fastfield(_fastfield(tmp_path))["meta"]
    info = info_desde_meta(meta)
    assert info["tramo"] == "Ramal Armenia"
    assert info["fecha"] == "12/03/2025"
    assert info["contratista"] == "PCC"
    assert info["inspector"] == "Juan Perez"


def test_no_inventa_campos_vacios():
    assert info_desde_meta({}) == {}
    assert info_desde_meta({"tramo": "", "tecnico": "Ana"}) == {"inspector": "Ana"}
    assert info_desde_meta(None) == {}


def test_el_tramo_llega_a_la_columna_tramo_de_hallazgos(tmp_path):
    """Con el tramo en Datos Generales, la columna F (TRAMO) de la hoja
    Hallazgos queda con el tramo inspeccionado."""
    from cips_adapter import cips_a_hallazgos
    from generator import ReportGenerator, resource_path
    meta = leer_dcvg_fastfield(_fastfield(tmp_path))["meta"]
    info = dict(info_desde_meta(meta), gasoducto="Mariquita-Cali",
                tipo_inspeccion="DCVG")
    hall = cips_a_hallazgos([
        {"abscisa_val": 5500, "observaciones": "cruce de caño",
         "referencia": "cruce de caño", "lat": 4.5, "lon": -75.7}])
    gen = ReportGenerator(resource_path("DCVG_REP.xlsx"))
    gen.fill_general_info(info)
    gen.fill_hallazgos(hall, info)
    out = str(tmp_path / "h.xlsx")
    gen.save(out)
    ws = openpyxl.load_workbook(out)["Hallazgos"]
    fila = next(r for r in range(12, 40)
                if ws.cell(row=r, column=1).value == 1)
    assert ws.cell(row=fila, column=6).value == "Ramal Armenia"   # F TRAMO
    assert ws.cell(row=fila, column=5).value == "Mariquita-Cali"  # E GASODUCTO


def _logger(tmp_path, tecnico="Evelio Alvarez", nombre="logger.xlsx"):
    """Data cruda del equipo: trae el técnico en 'Survey Info'."""
    wb = openpyxl.Workbook()
    si = wb.active
    si.title = "Survey Info"
    si.append(["Survey Name", "DCVG ARMENIA"])
    si.append(["Technician Name", tecnico])
    sd = wb.create_sheet("Survey Data")
    sd.append(["Data No", "Station No", "Latitude", "Longitude"])
    sd.append([1, 100, 4.5, -75.7])
    ruta = str(tmp_path / nombre)
    wb.save(ruta)
    return ruta


def test_inspector_sale_de_la_data_cruda(tmp_path):
    """El inspector debe salir del logger (es el nombre que coincide con el
    listado de equipos), no del 'Técnico a cargo' del FastField."""
    meta = leer_dcvg_fastfield(_fastfield(tmp_path))["meta"]   # técnico: Juan Perez
    info = info_desde_meta(meta, [_logger(tmp_path)])
    assert info["inspector"] == "Evelio Alvarez"
    # lo demás sigue viniendo del FastField
    assert info["tramo"] == "Ramal Armenia"
    assert info["contratista"] == "PCC"


def test_si_el_logger_no_trae_tecnico_usa_el_del_fastfield(tmp_path):
    meta = leer_dcvg_fastfield(_fastfield(tmp_path))["meta"]
    vacio = _logger(tmp_path, tecnico="", nombre="sin_tecnico.xlsx")
    assert info_desde_meta(meta, [vacio])["inspector"] == "Juan Perez"
    assert info_desde_meta(meta, [])["inspector"] == "Juan Perez"
    assert info_desde_meta(meta)["inspector"] == "Juan Perez"


def test_sin_fastfield_igual_toma_el_del_logger(tmp_path):
    info = info_desde_meta({}, [_logger(tmp_path)])
    assert info["inspector"] == "Evelio Alvarez"
