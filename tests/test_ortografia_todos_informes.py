"""La corrección ortográfica aplica a TODOS los informes (no solo CIPS):
potenciales PAP, hallazgos y PPM."""
import os

import openpyxl

from generator import ReportGenerator
from ppm_generator import PPMGenerator


def test_pap_corrige_observaciones(tmp_path):
    gen = ReportGenerator()
    pots = [{'abscisa': 100, 'ref_geografica': 'cruse de via',
             'observaciones': 'valvula sin proteccion', 'on_mv': -1100,
             'off_mv': -900}]
    gen.fill_potenciales_pap(pots)
    out = os.path.join(tmp_path, "pap.xlsx")
    gen.save(out)
    ws = openpyxl.load_workbook(out)['Potenciales PAP']
    assert ws.cell(row=12, column=4).value == 'Cruce de vía'
    assert ws.cell(row=12, column=27).value == 'Válvula sin protección'


def test_hallazgos_corrige_descripcion(tmp_path):
    gen = ReportGenerator()
    gen.fill_hallazgos([{'abscisa_val': 10, 'tipo': 'cruce',
                         'descripcion': 'cruse linea de alta tencion',
                         'lat': 4.0, 'lon': -73.0}], {})
    out = os.path.join(tmp_path, "h.xlsx")
    gen.save(out)
    ws = openpyxl.load_workbook(out)['Hallazgos']
    assert ws.cell(row=18, column=13).value == 'Cruce línea de alta tensión'


def test_ppm_corrige_comentario(tmp_path):
    out = os.path.join(tmp_path, "ppm.xlsx")
    PPMGenerator().generate({}, [{'abscisa': 5, 'observaciones': 'tramo aerio',
                                  'on_mv': -1100, 'off_mv': -900,
                                  'lat': 4.0, 'lon': -73.0}], [], out)
    ws = openpyxl.load_workbook(out)['CIPS - PAP']
    assert ws.cell(row=2, column=17).value == 'Tramo aéreo'
