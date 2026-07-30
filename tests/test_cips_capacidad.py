"""Un survey CIPS con MÁS puntos que la capacidad de la hoja (bloque de firmas
al final) no debe tronar con 'MergedCell is read-only': fill_cips escribe hasta
la capacidad, avisa cuántos recortó (cips_truncados) y deja el bloque de firmas
intacto. Caso real: OCENSA El Porvenir–Miraflores (survey muy largo)."""
import openpyxl

from generator import ReportGenerator, resource_path


def _capacidad(ws):
    firmas = next((r for r in range(12, ws.max_row + 1)
                   if ws.cell(row=r, column=3).value == 'ELABORÓ'), None)
    resumen = min([m.min_row for m in ws.merged_cells.ranges
                   if m.min_row >= 12] + [firmas])
    return resumen - 12, firmas


def test_capacidad_grande_para_ocensa():
    # Tras extender la hoja, la capacidad debe superar con creces los ~29k
    # puntos que antes hacían tronar el informe (caso OCENSA).
    gen = ReportGenerator(resource_path("CIPS EN BLANCO.xlsx"))
    cap, _ = _capacidad(gen.wb['Potenciales CIPS'])
    assert cap >= 60000, f"capacidad insuficiente: {cap}"


def test_survey_largo_extiende_formato(tmp_path):
    # 35 000 puntos (más que las filas pre-formateadas ~29 336): no debe tronar,
    # no debe recortar, y las filas nuevas quedan formateadas + contadas.
    import os
    gen = ReportGenerator(resource_path("CIPS EN BLANCO.xlsx"))
    n = 35000
    datos = [{'abscisa_val': i, 'on_mv': -1100.0, 'off_mv': -900.0,
              'on_limpio': -1100.0, 'off_limpio': -900.0} for i in range(n)]
    gen.fill_cips(datos)   # no debe lanzar
    assert gen.cips_truncados == 0
    ws = gen.wb['Potenciales CIPS']
    # una fila más allá de la zona pre-formateada tiene potencial y formato
    r = 12 + 34000
    assert ws.cell(row=r, column=5).value == -1100.0
    assert 'K' in ws.cell(row=r, column=2).number_format
    assert str(ws.cell(row=r, column=22).value).startswith('=IF(F')


def test_survey_normal_no_recorta():
    gen = ReportGenerator(resource_path("CIPS EN BLANCO.xlsx"))
    datos = [{'abscisa_val': i * 5, 'on_mv': -1100.0, 'off_mv': -900.0,
              'on_limpio': -1100.0, 'off_limpio': -900.0} for i in range(50)]
    gen.fill_cips(datos)
    assert gen.cips_truncados == 0
