"""La fecha de inspección de cada punto (día del timestamp 'On Time' del
archivo crudo) debe arrastrarse a la hoja CIPS (col FECHA), al PPM y a los
hallazgos. Con varios archivos de fechas distintas, cada punto lleva la suya."""
import datetime
import os

import openpyxl
import pandas as pd
import shapefile

from cips_lrs import procesar_cips_lrs
from cips_adapter import lrs_df_a_cips_dicts
from generator import ReportGenerator, resource_path
from ppm_generator import PPMGenerator


def _archivo(tmp_path, shp_real, dia, nombre, base_dist=0):
    sf = shapefile.Reader(shp_real)
    pts = sf.shapes()[0].points
    muestra = pts[base_dist:base_dist + 8]
    lons = [p[0] for p in muestra]
    lats = [p[1] for p in muestra]
    n = len(muestra)
    t0 = datetime.datetime(dia.year, dia.month, dia.day, 8, 0, 0)
    survey = pd.DataFrame({
        "Data No": range(base_dist + 1, base_dist + n + 1),
        "Dist From Start": [i * 10.0 for i in range(n)],
        "On Voltage": [-1.5] * n, "Off Voltage": [-1.0] * n,
        "Latitude": lats, "Longitude": lons,
        "Comment": [None] * n, "DCP/Feature/DCVG Anomaly": [None] * n,
        "On Time": [t0 + datetime.timedelta(minutes=i) for i in range(n)],
    })
    dcp = pd.DataFrame({"Data No": [base_dist + 1], "Device ID": ["X"],
                        "Comments": ["ini"]})
    ruta = os.path.join(tmp_path, nombre)
    with pd.ExcelWriter(ruta, engine="openpyxl") as w:
        survey.to_excel(w, sheet_name="Survey Data", index=False)
        dcp.to_excel(w, sheet_name="DCP Data", index=False)
    return ruta


def test_fecha_por_punto_en_cips_ppm(tmp_path, shp_real):
    a1 = _archivo(tmp_path, shp_real, datetime.date(2026, 6, 27), "d1.xlsx", 0)
    a2 = _archivo(tmp_path, shp_real, datetime.date(2026, 6, 28), "d2.xlsx", 20)
    df = procesar_cips_lrs([a1, a2], shp_real)
    dicts = lrs_df_a_cips_dicts(df)

    fechas = {d['fecha'] for d in dicts if d.get('fecha')}
    assert '27/06/2026' in fechas and '28/06/2026' in fechas, f"fechas: {fechas}"

    # Hoja CIPS: columna C = FECHA por punto
    gen = ReportGenerator(resource_path("CIPS EN BLANCO.xlsx"))
    gen.fill_cips(dicts)
    ws = gen.wb['Potenciales CIPS']
    en_hoja = {ws.cell(row=r, column=3).value
               for r in range(12, 12 + len(dicts))}
    assert '27/06/2026' in en_hoja and '28/06/2026' in en_hoja

    # PPM: columna 6 = fecha por punto
    out = os.path.join(tmp_path, "ppm.xlsx")
    PPMGenerator().generate({'route_id': 'R'}, [], [], out, cips=dicts)
    wp = openpyxl.load_workbook(out)['CIPS - PAP']
    ppm_fechas = {wp.cell(row=r, column=6).value
                  for r in range(2, 2 + len(dicts))}
    assert '27/06/2026' in ppm_fechas and '28/06/2026' in ppm_fechas
