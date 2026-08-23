"""DESCRIPCIÓN DE LA LÍNEA OBJETO DE ESTUDIO del informe DCVG.

En la plantilla DCVG esa sección viene VACÍA (a diferencia de la de PAP, que
trae el párrafo de un informe de ejemplo). Se redacta con los datos de la
inspección, con la misma estructura del formato.
"""
import os

import openpyxl

from generator import ReportGenerator, resource_path

INFO = {'tramo': 'Ramal Armenia', 'gasoducto': 'Mariquita-Cali',
        'longitud_km': 45.3, 'tipo_recubrimiento': 'Tricapa', 'diametro': '6"',
        'tipo_inspeccion': 'DCVG', 'distrito': 'D07'}


def _informe(tmp_path, plantilla, info, nombre="i.xlsx"):
    gen = ReportGenerator(resource_path(plantilla))
    gen.fill_general_info(dict(info))
    out = os.path.join(tmp_path, nombre)
    gen.save(out)
    return openpyxl.load_workbook(out)["Informe"]


def test_descripcion_de_la_linea_dcvg(tmp_path):
    ws = _informe(tmp_path, "DCVG_REP.xlsx", INFO)
    d = str(ws["A26"].value or '')
    assert "Ramal Armenia" in d
    assert "Mariquita-Cali" in d
    assert "45,3 Km" in d
    assert "Tricapa" in d
    assert "6 in" in d           # el diámetro va sin comillas
    assert "Pereira" not in d    # nada del formato de ejemplo


def test_con_rectificadores_menciona_el_mecanismo(tmp_path):
    info = dict(INFO, rectificadores_tgi="Balboa, Cerritos")
    ws = _informe(tmp_path, "DCVG_REP.xlsx", info)
    d = str(ws["A26"].value or '')
    assert "corriente impresa" in d and "Balboa, Cerritos" in d


def test_sin_rectificadores_no_inventa_el_mecanismo(tmp_path):
    info = dict(INFO, rectificadores_tgi="[ESCRIBIR RECTIFICADORES TGI]")
    ws = _informe(tmp_path, "DCVG_REP.xlsx", info)
    d = str(ws["A26"].value or '')
    assert "corriente impresa" not in d
    assert "ESCRIBIR" not in d


def test_sin_tramo_no_escribe_nada(tmp_path):
    ws = _informe(tmp_path, "DCVG_REP.xlsx", {'tipo_inspeccion': 'DCVG'})
    assert ws["A26"].value in (None, '')


def test_no_toca_la_descripcion_de_pap(tmp_path):
    """En PAP esa sección la escribe el ingeniero: no se sobrescribe."""
    ws = _informe(tmp_path, "EN BLANCO.xlsx", dict(INFO, tipo_inspeccion='PAP'),
                  nombre="pap.xlsx")
    assert "Ramal Pereira" in str(ws["A30"].value)
