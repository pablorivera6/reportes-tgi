"""Generación del informe DCVG: hojas Inspección DCVG y Resistividad."""
import os

import openpyxl

from generator import ReportGenerator, resource_path


def _postes():
    return [
        {"tipo": "Poste de potencial", "pk_m": 0, "on": -1682.0, "off": -1169.0,
         "vac": 0.44, "resistencia": 0.3, "lat": 4.53, "lon": -75.74},
        {"tipo": "Poste de potencial", "pk_m": 6000, "on": -1600.0, "off": -1100.0,
         "vac": 0.4, "resistencia": 0.3, "lat": 4.50, "lon": -75.75},
    ]


def _defectos():
    return [
        # defecto entre los dos postes -> pulso interpolable
        {"sector": "M", "lat": 4.52, "lon": -75.745, "pk_m": 3000,
         "forma_n": 54.8, "forma_s": 23.8, "forma_e": 26.1, "forma_o": 52.9,
         "ol_re": 95.7, "profundidad": 190, "clasificacion_campo": "Grande",
         "posicion_reloj": "12", "comentarios": "cruce", "caracter": "CC"},
        # defecto sin PK -> se omite (no rompe)
        {"sector": "M", "lat": None, "lon": None, "pk_m": None,
         "forma_n": 1, "forma_s": 1, "forma_e": 1, "forma_o": 1,
         "ol_re": 10, "profundidad": 100, "clasificacion_campo": "",
         "posicion_reloj": "", "comentarios": "", "caracter": "CA"},
    ]


def _gen():
    return ReportGenerator(resource_path("DCVG_REP.xlsx"))


def test_fill_dcvg_ordena_y_mapea(tmp_path):
    gen = _gen()
    # un hallazgo entre el defecto (3000) y el poste (6000)
    hall = [{'abscisa_val': 4500, 'descripcion': 'Cruce vía', 'lat': 4.5,
             'lon': -75.7, 'tipo': 'Cruce'}]
    gen.fill_dcvg(_postes(), _defectos(), hallazgos=hall)
    out = os.path.join(tmp_path, "dcvg.xlsx")
    gen.save(out)
    ws = openpyxl.load_workbook(out)["Inspección DCVG"]

    # 3 filas escritas (2 postes + 1 defecto con PK); el defecto sin PK se omite
    assert gen.dcvg_omitidos == 1
    # orden por abscisa intercalando el hallazgo: 0, 3000(defecto), 4500(hallazgo), 6000
    assert ws.cell(row=8, column=4).value == 0        # D abscisa poste0
    assert ws.cell(row=9, column=4).value == 3000     # defecto
    assert ws.cell(row=10, column=4).value == 4500    # hallazgo intercalado
    assert ws.cell(row=10, column=2).value == 'Cruce vía'   # B referencia
    assert ws.cell(row=11, column=4).value == 6000    # poste
    # poste: ON/OFF en N/O y pulso P=ABS(N-O)
    assert ws.cell(row=8, column=14).value == -1682.0   # N
    assert ws.cell(row=8, column=15).value == -1169.0   # O
    assert str(ws.cell(row=8, column=16).value) == "=ABS(N8-O8)"  # P
    # defecto: forma N->H, E->I, S->J, O->K
    assert ws.cell(row=9, column=8).value == 54.8    # H (N)
    assert ws.cell(row=9, column=9).value == 26.1    # I (E)
    assert ws.cell(row=9, column=10).value == 23.8   # J (S)
    assert ws.cell(row=9, column=11).value == 52.9   # K (O)
    assert ws.cell(row=9, column=12).value == "CC"   # L caracter
    assert ws.cell(row=9, column=13).value == 95.7   # M OL/RE
    # %IR en U (porque CC) = OL/RE como fracción para el formato '0%' (=M/100),
    # y P/RE (Q) interpolada
    assert str(ws.cell(row=9, column=21).value) == "=M9/100"              # U
    assert str(ws.cell(row=9, column=17).value).startswith("=((P")        # Q interpolación
    # clasificación V con umbrales en fracción (0.15/0.35/0.6)
    v = str(ws.cell(row=9, column=22).value)
    assert "Muy Pequeño" in v and "0.15" in v and "0.6" in v
    # distancia C = fórmula
    assert ws.cell(row=9, column=3).value == "=D9-$D$8"
    # firmas intactas debajo
    assert ws.cell(row=239, column=3).value and 'ELABOR' in str(ws.cell(row=239, column=3).value).upper()


def test_fill_resistividad(tmp_path):
    gen = _gen()
    # a propósito desordenadas -> deben quedar ascendentes por abscisa
    resist = [{"pk_m": 136750, "sector": "Potrero", "profundidad": 192,
               "lat": 4.96, "lon": -75.78, "r1": 3.8, "r2": 2.3, "r3": 2.0},
              {"pk_m": 136500, "sector": "Potrero", "profundidad": 198,
               "lat": 4.97, "lon": -75.77, "r1": 6.1, "r2": 3.2, "r3": 2.8}]
    gen.fill_resistividad(resist)
    out = os.path.join(tmp_path, "r.xlsx")
    gen.save(out)
    ws = openpyxl.load_workbook(out)["Resistividad"]
    # ordenadas ascendente: 136500 antes que 136750
    assert ws.cell(row=9, column=1).value == 136500    # A abscisa (menor primero)
    assert ws.cell(row=10, column=1).value == 136750
    assert ws.cell(row=9, column=6).value == 6.1       # F R1
    assert ws.cell(row=9, column=8).value == 3.2       # H R2
    assert ws.cell(row=9, column=10).value == 2.8      # J R3
    # la fórmula ρ del template sigue ahí
    assert str(ws.cell(row=9, column=12).value).startswith("=2*PI()")
