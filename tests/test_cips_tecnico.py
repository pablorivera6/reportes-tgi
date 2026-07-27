"""Al subir archivos CIPS se identifica el técnico (hoja 'Survey Info' ->
'Technician Name') para autollenar inspector y seriales de sus equipos."""
import os

import openpyxl
import pandas as pd

from cips_lrs import tecnico_de_archivos


def _archivo(tmp_path, nombre, tecnico):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Survey Data"
    ws.append(["Data No", "On Voltage"])
    ws.append([1, -1.5])
    info = wb.create_sheet("Survey Info")
    info.append(["Field", "Value"])
    info.append(["Name of P/L", "Ramal X"])
    info.append(["Technician Name", tecnico])
    info.append(["SurveyName", "s1"])
    ruta = os.path.join(tmp_path, nombre)
    wb.save(ruta)
    return ruta


def test_extrae_tecnico(tmp_path):
    a = _archivo(tmp_path, "a.xlsx", "EVELIO ALVAREZ")
    assert tecnico_de_archivos([a]) == "EVELIO ALVAREZ"


def test_tecnico_mas_frecuente(tmp_path):
    a = _archivo(tmp_path, "a.xlsx", "EVELIO ALVAREZ")
    b = _archivo(tmp_path, "b.xlsx", "EVELIO ALVAREZ")
    c = _archivo(tmp_path, "c.xlsx", "OTRO TECNICO")
    assert tecnico_de_archivos([a, b, c]) == "EVELIO ALVAREZ"


def test_sin_survey_info(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.title = "Survey Data"
    ruta = os.path.join(tmp_path, "x.xlsx")
    wb.save(ruta)
    assert tecnico_de_archivos([ruta]) == ""
