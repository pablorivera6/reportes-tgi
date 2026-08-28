"""Histórico DCVG: lectura del informe de una inspección DCVG anterior.

Los DCVG de otro contratista (TELMACOM 2024) vienen en una plantilla distinta
a la de PCC: la severidad %IR va en UNA columna (Q) y no en tres (AA/CA/CC), la
profundidad en metros y no en cm, y el carácter como 'C-A' y no 'CA'. Por eso
la hoja se lee por ETIQUETA de encabezado y no por posición, igual que el
histórico CIPS. La metadata (tramo, fecha, contratista) no está en la hoja de
datos sino en la hoja 'Informe'.
"""
import openpyxl
import pytest

from historicos import leer_historico, a_csv, desde_csv


def _hoja_informe(wb):
    ws = wb.create_sheet("Informe")
    ws["A6"], ws["G6"] = "Fecha", "Del 20/04/2024 al 22/04/2024"
    ws["A7"], ws["G7"] = "Gasoducto", "Mariquita - Cali"
    ws["A8"], ws["G8"] = "Tramo", "Montenegro"
    ws["A9"], ws["G9"] = "Inspector", "Miguel Rincón"
    ws["AA6"], ws["AC6"] = "No de contrato", "551003090"
    ws["AA7"], ws["AC7"] = "Contratista", "TELMACOM SAS"
    ws["AA8"], ws["AC8"] = "OT", "1300007688"
    return ws


def _encabezado_telmacom(ws):
    """Plantilla del contratista: %IR en una sola columna (Q)."""
    for col, txt in [("A6", "ÍTEM"), ("B6", "REFERENCIAS GEOGRÁFICAS"),
                     ("C6", "DISTANCIA TRAMO\n[m]"), ("D6", "ABSCISA"),
                     ("E6", "LATITUD"), ("F6", "LONGITUD"), ("G6", "ALTITUD\n[msnm]"),
                     ("H6", "FORMA [mV]"), ("L6", "CARÁCTER"), ("M6", "OL/RE\n[mV]"),
                     ("N6", "POTENCIAL ESTRUCTURA-SUELO [mV]"), ("P6", "P/RE\n[mV] "),
                     ("Q6", "SEVERIDAD [%IR]"), ("R6", "PROFUNDIDAD [m]"),
                     ("S6", "SEVERIDAD [CLASIFICACIÓN]"), ("T6", "RESISTIVIDAD [ohm-cm]"),
                     ("U6", "OBSERVACIONES")]:
        ws[col] = txt
    ws["N7"], ws["O7"] = "ON", "OFF"


def _informe_dcvg(tmp_path, nombre="hist_dcvg.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inspección DCVG"
    _encabezado_telmacom(ws)
    # 2 postes con ON/OFF
    for i, (r, absc, on, off) in enumerate(
            [(8, 0, -1369, -1051), (9, 1000, -1313, -1030)]):
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value="Poste de Potencial")
        ws.cell(row=r, column=4, value=absc)
        ws.cell(row=r, column=5, value=4.884)
        ws.cell(row=r, column=6, value=-75.914)
        ws.cell(row=r, column=14, value=on)
        ws.cell(row=r, column=15, value=off)
        ws.cell(row=r, column=19, value="-")
    # 2 defectos: %IR en FRACCIÓN (0.111 = 11,1 %)
    for i, (r, absc, olre, pre, frac, clas, car) in enumerate(
            [(10, 510, 22, 197.487, 0.11139973, "Muy Pequeño", "C-A"),
             (11, 1660, 260, 614.217, 0.42327, "Mediano", "A-A")]):
        ws.cell(row=r, column=1, value=i + 3)
        ws.cell(row=r, column=2, value="Defecto")
        ws.cell(row=r, column=4, value=absc)
        ws.cell(row=r, column=5, value=4.88)
        ws.cell(row=r, column=6, value=-75.92)
        ws.cell(row=r, column=12, value=car)
        ws.cell(row=r, column=13, value=olre)
        ws.cell(row=r, column=16, value=pre)
        ws.cell(row=r, column=17, value=frac)
        ws.cell(row=r, column=18, value=2.6)
        ws.cell(row=r, column=19, value=clas)
        ws.cell(row=r, column=20, value=132313.65)
    ws.cell(row=13, column=2, value="ELABORÓ")          # bloque de firmas
    ws.cell(row=14, column=4, value=99999)              # basura tras las firmas
    _hoja_informe(wb)
    ruta = str(tmp_path / nombre)
    wb.save(ruta)
    return ruta


def test_tipo_y_metadata_desde_la_hoja_informe(tmp_path):
    h = leer_historico(_informe_dcvg(tmp_path))
    assert h["tipo"] == "DCVG"
    assert h["tramo"] == "Montenegro"
    assert h["contratista"] == "TELMACOM SAS"
    assert h["periodo"] == "Abr 2024"


def test_separa_postes_y_defectos(tmp_path):
    h = leer_historico(_informe_dcvg(tmp_path))
    postes = [p for p in h["puntos"] if p["clase"] == "poste"]
    defectos = [p for p in h["puntos"] if p["clase"] == "defecto"]
    assert len(postes) == 2 and len(defectos) == 2
    assert postes[0]["on"] == -1369 and postes[0]["off"] == -1051
    assert defectos[0]["abscisa"] == 510 and defectos[0]["ol_re"] == 22


def test_severidad_en_porcentaje_y_caracter_normalizado(tmp_path):
    """El informe trae la fracción (0.111); el portal usa % (11.1) y 'CA'."""
    h = leer_historico(_informe_dcvg(tmp_path))
    d = [p for p in h["puntos"] if p["clase"] == "defecto"]
    assert d[0]["severidad_pct"] == pytest.approx(11.1, abs=0.05)
    assert d[0]["clasificacion"] == "Muy Pequeño"
    assert d[0]["caracter"] == "CA"
    assert d[1]["severidad_pct"] == pytest.approx(42.3, abs=0.05)
    assert d[1]["caracter"] == "AA"


def test_no_pasa_del_bloque_de_firmas(tmp_path):
    h = leer_historico(_informe_dcvg(tmp_path))
    assert len(h["puntos"]) == 4
    assert all(p["abscisa"] != 99999 for p in h["puntos"])


def test_resumen_dcvg(tmp_path):
    r = leer_historico(_informe_dcvg(tmp_path))["resumen"]
    assert r["n_defectos"] == 2 and r["n_postes"] == 2
    assert r["por_clasificacion"]["Muy Pequeño"] == 1
    assert r["por_clasificacion"]["Mediano"] == 1
    assert r["n_criticos"] == 1
    assert r["long_m"] == 1660
    assert r["densidad_km"] == pytest.approx(2 / 1.66, abs=0.01)


def test_clasificacion_se_recalcula_si_el_informe_la_trae_rota(tmp_path):
    """Varios informes traen '#DIV/0!' en la columna de clasificación."""
    ruta = _informe_dcvg(tmp_path, "roto.xlsx")
    wb = openpyxl.load_workbook(ruta)
    wb["Inspección DCVG"].cell(row=11, column=19, value="#DIV/0!")
    wb.save(ruta)
    d = [p for p in leer_historico(ruta)["puntos"] if p["clase"] == "defecto"]
    assert d[1]["clasificacion"] == "Mediano"       # 42,3 % -> Mediano


def test_plantilla_pcc_con_severidad_en_tres_columnas(tmp_path):
    """La plantilla de PCC parte %IR en AA/CA/CC (S/T/U): igual debe leerse."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inspección DCVG"
    for col, txt in [("A6", "ÍTEM"), ("B6", "REFERENCIAS GEOGRÁFICAS"),
                     ("D6", "ABSCISA"), ("E6", "LATITUD"), ("F6", "LONGITUD"),
                     ("L6", "CARÁCTER"), ("M6", "OL/RE\n[mV]"),
                     ("N6", "POTENCIAL ESTRUCTURA-SUELO [mV]"), ("P6", "PULSO\n[mV]"),
                     ("Q6", "P/RE\n[mV] "), ("R6", "PROFUNDIDAD [cm]"),
                     ("S6", "SEVERIDAD [%IR]"), ("V6", "SEVERIDAD [CLASIFICACIÓN]"),
                     ("W6", "RESISTIVIDAD [Ohm-cm]"), ("X6", "OBSERVACIONES")]:
        ws[col] = txt
    ws["N7"], ws["O7"], ws["S7"], ws["T7"], ws["U7"] = "ON", "OFF", "AA", "CA", "CC"
    ws.cell(row=8, column=2, value="Defecto")
    ws.cell(row=8, column=4, value=300)
    ws.cell(row=8, column=12, value="CC")
    ws.cell(row=8, column=13, value=40)
    ws.cell(row=8, column=17, value=200)            # Q = P/RE
    ws.cell(row=8, column=21, value=0.20)           # U = %IR (CC)
    ws.cell(row=8, column=22, value="Pequeño")      # V = clasificación
    _hoja_informe(wb)
    ruta = str(tmp_path / "pcc.xlsx")
    wb.save(ruta)
    d = [p for p in leer_historico(ruta)["puntos"] if p["clase"] == "defecto"]
    assert len(d) == 1
    assert d[0]["severidad_pct"] == pytest.approx(20.0, abs=0.05)
    assert d[0]["p_re"] == 200 and d[0]["caracter"] == "CC"


def test_csv_conserva_los_campos_del_defecto(tmp_path):
    h = leer_historico(_informe_dcvg(tmp_path))
    ruta = a_csv(h, str(tmp_path / "h.csv"))
    v = desde_csv(ruta)
    assert v["tipo"] == "DCVG" and v["tramo"] == "Montenegro"
    d = [p for p in v["puntos"] if p["clase"] == "defecto"]
    assert len(d) == 2
    assert d[0]["severidad_pct"] == pytest.approx(11.1, abs=0.05)
    assert d[0]["clasificacion"] == "Muy Pequeño"
    assert v["resumen"]["n_defectos"] == 2
