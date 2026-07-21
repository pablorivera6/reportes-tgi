"""Los comentarios de campo del survey CIPS deben ir también a la hoja
Hallazgos del formato, con todas las columnas (abscisa, coordenadas,
gasoducto/tramo, fecha, tipo, descripción). Los marcadores de solo
abscisado ('pk 1+000', '0+031', 'PK 002+000 No existe') NO son hallazgos."""
import openpyxl

from cips_adapter import cips_a_hallazgos
from generator import ReportGenerator, resource_path


def _dicts():
    return [
        {'abscisa_val': 19, 'observaciones': 'salida valvula pk 0+000',
         'lat': 4.627, 'lon': -75.686},
        {'abscisa_val': 45, 'observaciones': '0+031', 'lat': 4.627, 'lon': -75.686},
        {'abscisa_val': 491, 'observaciones': 'cruse caño', 'lat': 4.628, 'lon': -75.685},
        {'abscisa_val': 802, 'observaciones': 'pk 1+000', 'lat': 4.629, 'lon': -75.684},
        {'abscisa_val': 1877, 'observaciones': 'partidura de cable',
         'lat': 4.630, 'lon': -75.683},
        {'abscisa_val': 2000, 'observaciones': '', 'lat': 4.631, 'lon': -75.682},
        {'abscisa_val': 5070, 'observaciones': 'pk 5+000 abcisado',
         'lat': 4.632, 'lon': -75.681},
    ]


def test_filtra_marcadores_y_mapea_campos():
    h = cips_a_hallazgos(_dicts())
    descs = [x['descripcion'] for x in h]
    assert descs == ['salida valvula pk 0+000', 'cruse caño', 'partidura de cable']
    assert h[0]['abscisa_val'] == 19 and h[0]['lat'] == 4.627
    assert h[2]['tipo'] != ''    # clasificado (cable)
    assert 'cable' in h[2]['tipo'].lower()


def test_hallazgos_cips_en_hoja_del_formato(tmp_path):
    import os
    gen = ReportGenerator(resource_path("CIPS EN BLANCO.xlsx"))
    info = {'fecha': '27/06/2026', 'gasoducto': 'Mariquita-Cali',
            'tramo': 'Salento', 'inspector': 'X', 'contrato': 'C1', 'ot': 'OT1',
            'contratista': 'PCC', 'tipo_inspeccion': 'CIPS'}
    h = cips_a_hallazgos(_dicts())
    gen.fill_hallazgos(h, info)
    out = os.path.join(tmp_path, "h.xlsx")
    gen.save(out)

    ws = openpyxl.load_workbook(out)['Hallazgos']
    fila = 18
    assert ws.cell(row=fila, column=1).value == 1                 # ITEM
    assert ws.cell(row=fila, column=2).value == 19                # abscisa inicio
    assert ws.cell(row=fila, column=5).value == 'Mariquita-Cali'  # gasoducto
    assert ws.cell(row=fila, column=6).value == 'Salento'         # tramo
    assert ws.cell(row=fila, column=7).value == 4.627             # lat
    assert ws.cell(row=fila, column=11).value == '27/06/2026'     # fecha
    assert ws.cell(row=fila, column=13).value == 'salida valvula pk 0+000'
    assert ws.cell(row=fila + 2, column=13).value == 'partidura de cable'
