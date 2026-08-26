"""PPM del informe DCVG (plantilla `DCVG_PPM_.xlsx`).

Hasta ahora el PPM solo se generaba en PAP/CIPS, así que la carpeta `05_PPM`
del paquete de entrega salía vacía en los DCVG.

La plantilla tiene dos hojas:
  · `DCVG`         — una fila por registro del recorrido (postes, defectos y
                     hallazgos), con las columnas de poste (P On/P Off) y las
                     de defecto (P_RE, OL_RE, PORC_IR, carácter, clasificación).
  · `RESISTIVIDAD` — una fila por medida Wenner, con las ρ ya calculadas.
"""
import os

import openpyxl
import pytest

from ppm_generator import PPMDcvgGenerator

INFO = {'tramo': 'Ramal Ansermanuevo', 'contrato': '551007370', 'distrito': 'D08',
        'tipo_ducto': 'Ramal', 'fecha': '07/08/2026', 'tipo_inspeccion': 'DCVG'}

POSTES = [{'tipo': 'Poste de potencial', 'pk_m': 0, 'on': -1600.0, 'off': -1100.0,
           'lat': 4.50, 'lon': -75.70},
          {'tipo': 'Poste de potencial', 'pk_m': 4000, 'on': -1500.0, 'off': -1000.0,
           'lat': 4.55, 'lon': -75.75}]

# pulso = 500 mV en ambos postes -> P/RE 500; OL/RE 95.7 -> 19,14 %IR
DEFECTOS = [{'pk_m': 2000, 'ol_re': 95.7, 'caracter': 'CC', 'profundidad': 190,
             'lat': 4.52, 'lon': -75.72, 'comentarios': 'defecto en soldadura'}]

RESIST = [{'pk_m': 1000, 'r1': 3.8, 'r2': 2.3, 'r3': 2.0, 'lat': 4.51,
           'lon': -75.71, 'sector': 'Potrero'}]

HALLAZGOS = [{'abscisa_val': 3000, 'descripcion': 'Cruce de vía',
              'observaciones': 'Cruce de vía', 'lat': 4.53, 'lon': -75.73}]


def _ppm(tmp_path, **kw):
    gen = PPMDcvgGenerator()
    out = os.path.join(tmp_path, "ppm.xlsx")
    gen.generate(kw.get('info', INFO), kw.get('postes', POSTES),
                 kw.get('defectos', DEFECTOS), kw.get('resistividades', RESIST),
                 hallazgos=kw.get('hallazgos', HALLAZGOS), output_path=out)
    return openpyxl.load_workbook(out), gen


def _fila(ws, abscisa):
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=7).value == abscisa:
            return r
    raise AssertionError(f"no hay fila con abscisa {abscisa}")


def test_cabecera_comun(tmp_path):
    ws = _ppm(tmp_path)[0]["DCVG"]
    r = _fila(ws, 0)
    assert ws.cell(row=r, column=1).value == "R_ANS"        # ENGROUTEID
    assert ws.cell(row=r, column=2).value == "551007370"    # contrato
    assert ws.cell(row=r, column=3).value == "D08"          # distrito
    assert ws.cell(row=r, column=4).value == "Ramal"        # tipo de tramo
    assert ws.cell(row=r, column=5).value == "Ramal Ansermanuevo"
    assert ws.cell(row=r, column=6).value == "07/08/2026"


def test_fila_de_poste(tmp_path):
    ws = _ppm(tmp_path)[0]["DCVG"]
    r = _fila(ws, 0)
    assert ws.cell(row=r, column=9).value == 4.50           # latitud
    assert ws.cell(row=r, column=13).value == -1600.0       # P On
    assert ws.cell(row=r, column=14).value == -1100.0       # P Off
    assert ws.cell(row=r, column=16).value in (None, "")    # OL_RE: no aplica


def test_fila_de_defecto(tmp_path):
    ws = _ppm(tmp_path)[0]["DCVG"]
    r = _fila(ws, 2000)
    assert ws.cell(row=r, column=15).value == pytest.approx(500, abs=1)   # P_RE
    assert ws.cell(row=r, column=16).value == 95.7                        # OL_RE
    # PORC_IR va en FRACCIÓN: la celda tiene formato 0.00%
    assert ws.cell(row=r, column=17).value == pytest.approx(0.1914, abs=0.001)
    assert ws.cell(row=r, column=19).value == "CC"                        # carácter
    # 95,7 / 500 = 19,14 % -> 'Pequeño' (Muy Pequeño es <= 15 %)
    assert ws.cell(row=r, column=20).value == "Pequeño"                   # clasificación
    assert ws.cell(row=r, column=21).value == 190                         # profundidad
    assert "soldadura" in str(ws.cell(row=r, column=22).value).lower()


def test_incluye_los_hallazgos(tmp_path):
    ws = _ppm(tmp_path)[0]["DCVG"]
    r = _fila(ws, 3000)
    assert "cruce" in str(ws.cell(row=r, column=22).value).lower()
    assert ws.cell(row=r, column=13).value in (None, "")    # sin medición


def test_ordenado_por_abscisa_y_marca_inicio_y_fin(tmp_path):
    ws = _ppm(tmp_path)[0]["DCVG"]
    absc = [ws.cell(row=r, column=7).value for r in range(2, 6)]
    assert absc == [0, 2000, 3000, 4000]
    assert ws.cell(row=2, column=12).value == "Inicio Inspección"
    assert ws.cell(row=5, column=12).value == "Fin Inspección"
    assert ws.cell(row=3, column=12).value in (None, "")


def test_sin_dato_queda_vacio(tmp_path):
    """Altitud y cama anódica temporal no las captura el FastField."""
    ws = _ppm(tmp_path)[0]["DCVG"]
    r = _fila(ws, 0)
    assert ws.cell(row=r, column=11).value in (None, "")    # altitud
    assert ws.cell(row=r, column=18).value in (None, "")    # cama anódica


def test_hoja_resistividad(tmp_path):
    ws = _ppm(tmp_path)[0]["RESISTIVIDAD"]
    r = _fila(ws, 1000)
    assert ws.cell(row=r, column=1).value == "R_ANS"
    assert ws.cell(row=r, column=5).value == "Ramal Ansermanuevo"
    import math
    assert ws.cell(row=r, column=12).value == pytest.approx(2 * math.pi * 3.8 * 100, rel=1e-6)
    assert ws.cell(row=r, column=13).value == pytest.approx(2 * math.pi * 2.3 * 200, rel=1e-6)
    assert ws.cell(row=r, column=14).value == pytest.approx(2 * math.pi * 2.0 * 300, rel=1e-6)
    # capa 1-2 m: ((R1*R2)/|R2-R1|)*2*pi*100
    esperado = ((3.8 * 2.3) / abs(2.3 - 3.8)) * 2 * math.pi * 100
    assert ws.cell(row=r, column=16).value == pytest.approx(esperado, rel=1e-6)


def test_no_arrastra_datos_de_la_plantilla(tmp_path):
    """Si la plantilla se guardó con datos, no deben quedar en el PPM."""
    wb, _g = _ppm(tmp_path, defectos=[], hallazgos=[], resistividades=[])
    ws = wb["DCVG"]
    filas = [r for r in range(2, ws.max_row + 1)
             if any(ws.cell(row=r, column=c).value not in (None, "")
                    for c in range(1, 23))]
    assert len(filas) == 2, "solo deberían quedar los 2 postes"


def test_sin_datos_no_rompe(tmp_path):
    wb, _g = _ppm(tmp_path, postes=[], defectos=[], resistividades=[], hallazgos=[])
    assert wb["DCVG"].cell(row=2, column=1).value in (None, "")
