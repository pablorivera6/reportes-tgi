"""CONCLUSIONES y RECOMENDACIONES: cada plantilla las tiene en filas distintas.

  Plantilla   CONCLUSIONES   RECOMENDACIONES   ELABORÓ
  DCVG        69             84                98
  PAP         83             94                99
  CIPS        81             90                93

`fill_conclusiones` buscaba la etiqueta a partir de la fila 70, así que en DCVG
(fila 69) no la encontraba y escribía en la 88 — encima de las recomendaciones.
Ahora la sección se ubica por su título y el texto no puede invadir la sección
siguiente ni el bloque de firmas.
"""
import os

import openpyxl
import pytest

from generator import ReportGenerator, resource_path

CONC = [f"Conclusión número {i}" for i in range(1, 3)]
RECO = [f"Recomendación número {i}" for i in range(1, 3)]

# plantilla -> (fila CONCLUSIONES, fila RECOMENDACIONES, cupo de recomendaciones)
# El espacio lo fija la propia plantilla: entre el título y la sección
# siguiente (o el bloque de firmas). CIPS solo deja 2 filas de recomendaciones.
LAYOUT = {"DCVG_REP.xlsx": (69, 84, 13),
          "EN BLANCO.xlsx": (83, 94, 4),
          "CIPS EN BLANCO.xlsx": (81, 90, 2)}


def _informe(tmp_path, plantilla, conc=CONC, reco=RECO, gen_out=None):
    gen = ReportGenerator(resource_path(plantilla))
    gen.fill_conclusiones(conc)
    gen.fill_recomendaciones(reco)
    out = os.path.join(tmp_path, plantilla)
    gen.save(out)
    if gen_out is not None:
        gen_out.append(gen)
    return openpyxl.load_workbook(out)["Informe"]


@pytest.mark.parametrize("plantilla", list(LAYOUT))
def test_van_debajo_de_su_titulo(tmp_path, plantilla):
    f_conc, f_reco, _cupo = LAYOUT[plantilla]
    ws = _informe(tmp_path, plantilla)
    assert ws.cell(row=f_conc, column=1).value == "CONCLUSIONES"
    assert ws.cell(row=f_reco, column=1).value == "RECOMENDACIONES"
    for i, c in enumerate(CONC):
        assert c in str(ws.cell(row=f_conc + 1 + i, column=1).value), \
            f"{plantilla}: la conclusión {i + 1} no quedó bajo su título"
    for i, r in enumerate(RECO):
        assert r in str(ws.cell(row=f_reco + 1 + i, column=1).value), \
            f"{plantilla}: la recomendación {i + 1} no quedó bajo su título"


def test_dcvg_no_escribe_encima_de_las_recomendaciones(tmp_path):
    ws = _informe(tmp_path, "DCVG_REP.xlsx")
    # antes las conclusiones caían en la fila 88, dentro del bloque siguiente
    assert "Conclusión" not in str(ws.cell(row=88, column=1).value or "")
    assert ws.cell(row=84, column=1).value == "RECOMENDACIONES"


@pytest.mark.parametrize("plantilla", list(LAYOUT))
def test_una_lista_larga_no_invade_la_seccion_siguiente(tmp_path, plantilla):
    f_conc, f_reco, _cupo = LAYOUT[plantilla]
    muchas = [f"Conclusión larga {i}" for i in range(30)]
    ws = _informe(tmp_path, plantilla, conc=muchas, reco=RECO)
    assert ws.cell(row=f_reco, column=1).value == "RECOMENDACIONES", \
        "las conclusiones pisaron el título de RECOMENDACIONES"
    for i, r in enumerate(RECO):
        assert r in str(ws.cell(row=f_reco + 1 + i, column=1).value)


@pytest.mark.parametrize("plantilla", list(LAYOUT))
def test_las_recomendaciones_no_pisan_las_firmas(tmp_path, plantilla):
    _f, f_reco, _cupo = LAYOUT[plantilla]
    muchas = [f"Recomendación larga {i}" for i in range(30)]
    ws = _informe(tmp_path, plantilla, conc=CONC, reco=muchas)
    fila_firmas = next(r for r in range(f_reco, ws.max_row + 1)
                       for c in range(1, 30)
                       if str(ws.cell(row=r, column=c).value or '').strip().upper()
                       == 'ELABORÓ')
    for c in range(1, 30):
        assert "Recomendación" not in str(
            ws.cell(row=fila_firmas, column=c).value or "")


def test_sin_conclusiones_no_toca_nada(tmp_path):
    ws = _informe(tmp_path, "DCVG_REP.xlsx", conc=[], reco=[])
    assert ws.cell(row=69, column=1).value == "CONCLUSIONES"
    assert ws.cell(row=70, column=1).value in (None, "")


@pytest.mark.parametrize("plantilla", list(LAYOUT))
def test_avisa_lo_que_no_cabe(tmp_path, plantilla):
    """El espacio lo fija la plantilla; lo que no entra se reporta en vez de
    perderse en silencio."""
    _fc, _fr, cupo = LAYOUT[plantilla]
    gens = []
    muchas = [f"Recomendación {i}" for i in range(cupo + 3)]
    _informe(tmp_path, plantilla, reco=muchas, gen_out=gens)
    assert gens[0].recomendaciones_omitidas == 3
    assert gens[0].conclusiones_omitidas == 0
