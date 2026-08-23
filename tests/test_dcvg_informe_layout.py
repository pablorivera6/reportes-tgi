"""La hoja 'Informe' de la plantilla DCVG NO tiene la misma distribución que la
de PAP/CIPS:

           PAP/CIPS                     DCVG
  OBJETIVO   fila 11 (texto 12-14)      fila 11 (texto en 12, TEXTO FIJO)
  DOCUMENTOS fila 16 (refs 17-20)       fila 13 (refs 14-16)
  EQUIPOS    fila 23 (lista 24-28)      fila 18 (lista 19-23)
  valores    columnas G / V / AF        columnas G / U / AE
  fila 7                No de OT                    Contratista
  fila 8                Contratista                 OT

Como el generador tenía las filas y columnas de PAP quemadas, el informe DCVG
salía con los datos del formato de ejemplo (Ramal Villa María, sus equipos) y
además pisaba los títulos de sección. Estos tests fijan la distribución de
cada plantilla.
"""
import os

import openpyxl
import pytest

from generator import ReportGenerator, resource_path

INFO = {'fecha': '12/03/2025', 'gasoducto': 'Mariquita-Cali',
        'tramo': 'Ramal Armenia', 'inspector': 'Juan Perez',
        'contrato': '551007370', 'ot': '1300013506', 'contratista': 'PCC',
        'serial_equipo': 'IBTVM-99999', 'fecha_calibracion': '01/01/2025',
        'tipo_recubrimiento': 'Tricapa', 'diametro': '6', 'ciclo': '2025',
        'distrito': 'D07', 'tipo_inspeccion': 'DCVG'}

EQUIPOS = ["Datalogger: iBTVM - 12345", "Electrodo: Mc Miller - 1",
           "GPS: Garmin - 2", "Multímetro: Fluke - 3", "Cámara: Canon - 4",
           "Osciloscopio: Tek - 5"]


def _informe(tmp_path, plantilla, info=None, equipos=None, nombre="i.xlsx"):
    gen = ReportGenerator(resource_path(plantilla))
    gen.fill_general_info(dict(info or INFO))
    if equipos:
        gen.fill_equipos_utilizados(equipos)
    out = os.path.join(tmp_path, nombre)
    gen.save(out)
    return openpyxl.load_workbook(out)["Informe"]


# ── DCVG ──────────────────────────────────────────────────────────────────────

def test_objetivo_dcvg_lleva_el_tramo_de_esta_inspeccion(tmp_path):
    ws = _informe(tmp_path, "DCVG_REP.xlsx")
    obj = str(ws["A12"].value)
    assert "Villa María" not in obj, "quedó el ramal del formato de ejemplo"
    assert "Ramal Armenia" in obj
    assert "Mariquita-Cali" in obj
    assert "distrito 7" in obj
    # se conserva la redacción del formato
    assert obj.startswith("Realizar un diagnóstico del estado general")
    assert "Clasificar puntos estratégicos" in obj


def test_valores_del_encabezado_dcvg_en_las_columnas_u_y_ae(tmp_path):
    ws = _informe(tmp_path, "DCVG_REP.xlsx")
    assert ws["G6"].value == '12/03/2025'          # fecha
    assert ws["G8"].value == 'Ramal Armenia'       # tramo
    assert ws["U6"].value == 'IBTVM-99999'         # serial
    assert ws["U7"].value == '01/01/2025'          # fecha calibración
    assert ws["U8"].value == 'Tricapa'             # recubrimiento
    assert ws["U9"].value == '6'                   # diámetro
    # OJO: en la plantilla DCVG las filas 7 y 8 de esta columna están al revés
    # que en PAP/CIPS (aquí 7=Contratista y 8=OT), por eso se mapea por etiqueta
    assert ws["AE6"].value == '551007370'          # No de contrato
    assert ws["AE7"].value == 'PCC'                # Contratista
    assert ws["AE8"].value == '1300013506'         # OT
    assert ws["AE9"].value == '2025'               # Ciclo


def test_equipos_dcvg_reemplazan_los_del_ejemplo(tmp_path):
    ws = _informe(tmp_path, "DCVG_REP.xlsx", equipos=EQUIPOS)
    bloque = [str(ws.cell(row=r, column=c).value or '')
              for r in range(19, 25) for c in (1, 19)]
    texto = " ".join(bloque)
    assert "Datalogger: iBTVM - 12345" in texto
    assert "Radiodetection-1211" not in texto, "quedaron los equipos del ejemplo"
    assert "IBTVM-06582" not in texto


def test_no_se_pisan_los_titulos_de_seccion_dcvg(tmp_path):
    ws = _informe(tmp_path, "DCVG_REP.xlsx", equipos=EQUIPOS)
    assert ws["A25"].value == 'DESCRIPCIÓN DE LA LÍNEA OBJETO DE ESTUDIO'
    assert ws["A27"].value == 'ANTECEDENTES'
    assert ws["A18"].value == 'EQUIPOS UTILIZADOS'


def test_procedimiento_no_pisa_los_equipos_dcvg(tmp_path):
    ws = _informe(tmp_path, "DCVG_REP.xlsx")
    # el PR-I-06 va en el bloque de documentos de referencia (14-17),
    # no encima de la lista de equipos (19-23)
    refs = " ".join(str(ws.cell(row=r, column=1).value or '') for r in range(14, 18))
    assert "PR-I-06" in refs and "DCVG" in refs
    assert "Localizador de tubería" in str(ws["A20"].value or ''), \
        "el procedimiento pisó un equipo"


# ── PAP / CIPS: no se debe romper lo que ya funcionaba ───────────────────────

@pytest.mark.parametrize("plantilla", ["EN BLANCO.xlsx", "CIPS EN BLANCO.xlsx"])
def test_pap_cips_mantienen_su_distribucion(tmp_path, plantilla):
    info = dict(INFO, tipo_inspeccion='PAP')
    ws = _informe(tmp_path, plantilla, info=info, equipos=EQUIPOS,
                  nombre=plantilla)
    assert ws["G6"].value == '12/03/2025'
    assert ws["V6"].value == 'IBTVM-99999'     # serial (columna V en PAP/CIPS)
    assert ws["V8"].value == 'Tricapa'
    assert ws["AF6"].value == '551007370'      # contrato
    assert ws["AF8"].value == 'PCC'
    assert "PR-I-06" in str(ws["A20"].value)   # procedimiento en su fila
    assert "Datalogger: iBTVM - 12345" in str(ws["A24"].value)
    assert ws["A23"].value == 'EQUIPOS UTILIZADOS'
    assert ws["A29"].value == 'DESCRIPCIÓN DE LA LÍNEA OBJETO DE ESTUDIO'
    # el objetivo de PAP/CIPS es una fórmula del formato: no se toca
    assert str(ws["A12"].value).startswith("=")
