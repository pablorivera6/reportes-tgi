"""Comentarios de la Gráfica VDC del informe CIPS: el template no debe traer
anotaciones de informes viejos, y fill_graficas_cips debe limpiarlas y
escribir las del survey actual."""
import openpyxl

from generator import ReportGenerator, resource_path


def _comentarios(ws):
    out = []
    for r in range(49, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in (4, 5, 6)]
        if any(v not in (None, '') for v in vals):
            out.append(vals)
    return out


def test_template_cips_sin_comentarios_viejos():
    wb = openpyxl.load_workbook(resource_path("CIPS EN BLANCO.xlsx"))
    restos = _comentarios(wb["Gráfica VDC "])
    assert not restos, f"El template trae comentarios viejos: {restos[:3]}"


def test_fill_graficas_cips_escribe_comentarios_del_survey():
    gen = ReportGenerator(resource_path("CIPS EN BLANCO.xlsx"))
    ws = gen.wb["Gráfica VDC "]
    # simular template sucio
    ws.cell(row=49, column=6, value="ANOTACION DE OTRO INFORME")
    ws.cell(row=80, column=6, value="OTRA VIEJA")

    datos = [{'abscisa_val': 100, 'observaciones': 'valvula pk 0+100',
              'on_mv': -1100, 'off_mv': -900},
             {'abscisa_val': 500, 'observaciones': '', 'on_mv': -1100,
              'off_mv': -900},
             {'abscisa_val': 900, 'observaciones': 'cruce via',
              'on_mv': -1100, 'off_mv': -900}]
    gen.fill_cips(datos)
    gen.fill_graficas_cips(datos, {})

    coms = _comentarios(ws)
    textos = [c[2] for c in coms]
    assert textos == ['valvula pk 0+100', 'cruce via'], f"comentarios: {textos}"
    assert coms[0][0] == 100 and coms[1][0] == 900
