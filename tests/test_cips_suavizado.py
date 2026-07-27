"""La hoja CIPS y su gráfica deben mostrar el potencial SUAVIZADO
(on_limpio/off_limpio), no el crudo — igual que la app original proceso-cips,
que grafica y reporta la data limpia. Los picos aislados no deben aparecer."""
import os

import openpyxl

from generator import ReportGenerator, resource_path


def test_columnas_EF_llevan_el_suavizado():
    gen = ReportGenerator(resource_path("CIPS EN BLANCO.xlsx"))
    # un punto con pico crudo fuerte pero limpio suave
    datos = [
        {'abscisa_val': 0, 'on_mv': -1600.0, 'off_mv': -1000.0,
         'on_limpio': -1600.0, 'off_limpio': -1000.0},
        {'abscisa_val': 5, 'on_mv': -3200.0, 'off_mv': -300.0,   # PICO crudo
         'on_limpio': -1605.0, 'off_limpio': -1002.0},           # suavizado
        {'abscisa_val': 10, 'on_mv': -1610.0, 'off_mv': -1004.0,
         'on_limpio': -1610.0, 'off_limpio': -1004.0},
    ]
    gen.fill_cips(datos)
    ws = gen.wb['Potenciales CIPS']
    # fila del pico (i=1 -> fila 13): E/F deben ser el suavizado, no el crudo
    assert ws.cell(row=13, column=5).value == -1605.0, "E debe ser on_limpio"
    assert ws.cell(row=13, column=6).value == -1002.0, "F debe ser off_limpio"
    # G/H [CORREGIDO] siguen vacías
    assert ws.cell(row=13, column=7).value in (None, '')
    assert ws.cell(row=13, column=8).value in (None, '')
    # IR con el suavizado
    assert round(ws.cell(row=13, column=18).value, 1) == -603.0


def test_fallback_a_crudo_si_no_hay_limpio():
    gen = ReportGenerator(resource_path("CIPS EN BLANCO.xlsx"))
    gen.fill_cips([{'abscisa_val': 0, 'on_mv': -1500.0, 'off_mv': -950.0}])
    ws = gen.wb['Potenciales CIPS']
    assert ws.cell(row=12, column=5).value == -1500.0
    assert ws.cell(row=12, column=6).value == -950.0
