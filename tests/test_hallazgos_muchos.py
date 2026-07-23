"""La plantilla trae 500 filas de datos pre-formateadas antes del bloque de
firmas, así que muchos hallazgos NO rompen el formato: cada fila queda con
borde y todas sus columnas, y las firmas quedan intactas debajo."""
import os

import openpyxl

from generator import ReportGenerator, resource_path


def _find_firmas(ws):
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=3).value == 'ELABORÓ':
            return r
    return None


def test_muchos_hallazgos_no_rompen_formato(tmp_path):
    n = 200
    gen = ReportGenerator(resource_path("CIPS EN BLANCO.xlsx"))
    info = {'gasoducto': 'Mariquita-Cali', 'tramo': 'Filandia',
            'fecha': '27/06/2026'}
    hall = [{'abscisa_val': 100 + i * 30, 'tipo': 'Observación de campo',
             'descripcion': f'punto {i}', 'lat': 4.6 + i * 0.001, 'lon': -75.7}
            for i in range(n)]
    gen.fill_hallazgos(hall, info)
    out = os.path.join(tmp_path, "h.xlsx")
    gen.save(out)

    ws = openpyxl.load_workbook(out)['Hallazgos']
    start = 18
    for i in range(n):
        r = start + i
        assert ws.cell(row=r, column=1).value == i + 1, f"item fila {r}"
        assert ws.cell(row=r, column=2).value == 100 + i * 30, f"abscisa fila {r}"
        assert ws.cell(row=r, column=5).value == 'Mariquita-Cali', f"gasoducto fila {r}"
        assert ws.cell(row=r, column=6).value == 'Filandia', f"tramo fila {r}"
        b = ws.cell(row=r, column=2).border
        assert b and b.left and b.left.style, f"sin borde en fila {r}"
    firmas = _find_firmas(ws)
    assert firmas is not None and firmas > start + n, f"firmas en {firmas}"
    assert ws.cell(row=firmas + 1, column=1).value == 'Nombre'


def test_pocos_hallazgos_limpia_sobrantes(tmp_path):
    gen = ReportGenerator(resource_path("CIPS EN BLANCO.xlsx"))
    gen.fill_hallazgos([{'abscisa_val': 10, 'tipo': 'Cruce', 'descripcion': 'x',
                         'lat': 4.0, 'lon': -73.0}], {'gasoducto': 'G', 'tramo': 'T'})
    out = os.path.join(tmp_path, "p.xlsx")
    gen.save(out)
    ws = openpyxl.load_workbook(out)['Hallazgos']
    assert ws.cell(row=18, column=1).value == 1
    # fila 19 en adelante debe quedar vacía (no residuos)
    assert ws.cell(row=19, column=1).value in (None, '')
    assert ws.cell(row=19, column=2).value in (None, '')
