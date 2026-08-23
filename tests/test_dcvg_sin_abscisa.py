"""Registros DCVG sin abscisa: NO se pierden.

El técnico a veces deja el PK vacío en FastField. Antes esos defectos/postes
se descartaban y el ingeniero tenía que volver al archivo de campo a buscarlos.
Ahora la fila se escribe igual, en el punto de la secuencia donde va (anclada
al registro anterior del archivo de campo), con la celda de abscisa VACÍA y
resaltada en amarillo, y con las fórmulas dependientes del PK escritas pero
neutralizadas: al escribir la abscisa, Excel completa distancia, P/RE,
severidad y clasificación solo.
"""
import os

import openpyxl

from generator import ReportGenerator, resource_path

AMARILLO = "FFFF00"


def _gen():
    return ReportGenerator(resource_path("DCVG_REP.xlsx"))


def _guardar(gen, tmp_path, nombre="dcvg.xlsx"):
    out = os.path.join(tmp_path, nombre)
    gen.save(out)
    return openpyxl.load_workbook(out)


def _es_amarilla(celda):
    f = celda.fill
    return bool(f and f.fill_type == "solid"
                and str(f.start_color.rgb or "").upper().endswith(AMARILLO))


def _postes():
    return [
        {"tipo": "Poste de potencial", "pk_m": 0, "on": -1682.0, "off": -1169.0,
         "lat": 4.53, "lon": -75.74},
        {"tipo": "Poste de potencial", "pk_m": 6000, "on": -1600.0, "off": -1100.0,
         "lat": 4.50, "lon": -75.75},
    ]


def test_defecto_sin_abscisa_queda_en_su_secuencia(tmp_path):
    """El defecto sin PK va justo después del defecto que lo precede en el
    archivo de campo, no al final ni descartado."""
    defectos = [
        {"pk_m": 3000, "forma_n": 54.8, "forma_s": 23.8, "forma_e": 26.1,
         "forma_o": 52.9, "ol_re": 95.7, "profundidad": 190, "caracter": "CC",
         "lat": 4.52, "lon": -75.745, "comentarios": "cruce"},
        # registrado a continuación en campo, pero sin PK
        {"pk_m": None, "forma_n": 10.0, "forma_s": 11.0, "forma_e": 12.0,
         "forma_o": 13.0, "ol_re": 40.0, "profundidad": 150, "caracter": "CA",
         "lat": 4.51, "lon": -75.746, "comentarios": "sin pk"},
    ]
    gen = _gen()
    gen.fill_dcvg(_postes(), defectos)
    ws = _guardar(gen, tmp_path)["Inspección DCVG"]

    # 4 filas: poste 0, defecto 3000, defecto sin PK, poste 6000
    assert ws.cell(row=8, column=4).value == 0
    assert ws.cell(row=9, column=4).value == 3000
    assert ws.cell(row=10, column=4).value in (None, "")   # sin abscisa
    assert ws.cell(row=11, column=4).value == 6000

    r = 10
    # los datos del defecto SÍ están
    assert ws.cell(row=r, column=2).value == "Defecto"           # B
    assert ws.cell(row=r, column=8).value == 10.0                # H forma N
    assert ws.cell(row=r, column=9).value == 12.0                # I forma E
    assert ws.cell(row=r, column=10).value == 11.0               # J forma S
    assert ws.cell(row=r, column=11).value == 13.0               # K forma O
    assert ws.cell(row=r, column=12).value == "CA"               # L carácter
    assert ws.cell(row=r, column=13).value == 40.0               # M OL/RE
    assert ws.cell(row=r, column=18).value == 150                # R profundidad
    assert ws.cell(row=r, column=5).value == 4.51                # E lat
    assert ws.cell(row=r, column=6).value == -75.746             # F lon
    # la celda de abscisa queda resaltada para completarla
    assert _es_amarilla(ws.cell(row=r, column=4))
    # y solo esa (la del defecto completo no se pinta)
    assert not _es_amarilla(ws.cell(row=9, column=4))


def test_formulas_se_activan_al_escribir_el_pk(tmp_path):
    """C (distancia), Q (P/RE) y la severidad quedan escritas pero en blanco
    mientras falta el PK: van envueltas en IF(D...="";"";...)."""
    defectos = [
        {"pk_m": 3000, "ol_re": 95.7, "caracter": "CC", "profundidad": 190,
         "lat": 4.52, "lon": -75.745},
        {"pk_m": None, "ol_re": 40.0, "caracter": "CA", "profundidad": 150,
         "lat": 4.51, "lon": -75.746},
    ]
    gen = _gen()
    gen.fill_dcvg(_postes(), defectos)
    ws = _guardar(gen, tmp_path)["Inspección DCVG"]
    r = 10   # defecto sin PK

    c = str(ws.cell(row=r, column=3).value)          # C distancia
    assert c.startswith('=IF(D10="","",') and 'D10-$D$8' in c

    q = str(ws.cell(row=r, column=17).value)         # Q P/RE interpolado
    assert q.startswith('=IF(D10="","",') and "P8" in q and "P11" in q

    # CA -> severidad en T(20); protegida contra Q vacío
    t = str(ws.cell(row=r, column=20).value)
    assert t.startswith('=IF(Q10="","",') and "M10/Q10" in t
    # clasificación V referida a la misma columna de severidad
    v = str(ws.cell(row=r, column=22).value)
    assert "T10" in v and "Muy Pequeño" in v


def test_sin_abscisa_al_inicio_se_ancla_al_siguiente(tmp_path):
    """Si el primer registro del archivo no trae PK, se ancla al siguiente que
    sí lo trae (queda ANTES de él, no al final de la hoja)."""
    defectos = [
        {"pk_m": None, "ol_re": 40.0, "caracter": "CA", "comentarios": "primero"},
        {"pk_m": 3000, "ol_re": 95.7, "caracter": "CC", "comentarios": "segundo"},
    ]
    gen = _gen()
    gen.fill_dcvg([], defectos)
    ws = _guardar(gen, tmp_path)["Inspección DCVG"]
    assert ws.cell(row=8, column=4).value in (None, "")     # el sin PK
    assert ws.cell(row=8, column=13).value == 40.0
    assert ws.cell(row=9, column=4).value == 3000


def test_poste_y_hallazgo_sin_abscisa_se_conservan(tmp_path):
    postes = [
        {"tipo": "Poste de potencial", "pk_m": 0, "on": -1682.0, "off": -1169.0},
        {"tipo": "Caja de empalme", "pk_m": None, "on": -1500.0, "off": -1000.0},
        {"tipo": "Poste de potencial", "pk_m": 6000, "on": -1600.0, "off": -1100.0},
    ]
    hall = [{"abscisa_val": None, "descripcion": "Cruce vía", "tipo": "Cruce"}]
    gen = _gen()
    gen.fill_dcvg(postes, [], hallazgos=hall)
    ws = _guardar(gen, tmp_path)["Inspección DCVG"]

    fp = next(r for r in range(8, 15)
              if ws.cell(row=r, column=2).value == "Caja de empalme")
    assert ws.cell(row=fp, column=4).value in (None, "")
    assert ws.cell(row=fp, column=14).value == -1500.0     # N ON
    assert ws.cell(row=fp, column=15).value == -1000.0     # O OFF
    assert str(ws.cell(row=fp, column=16).value) == f"=ABS(N{fp}-O{fp})"  # P pulso
    assert _es_amarilla(ws.cell(row=fp, column=4))

    fh = next(r for r in range(8, 15)
              if ws.cell(row=r, column=2).value == "Cruce vía")
    assert ws.cell(row=fh, column=4).value in (None, "")
    assert _es_amarilla(ws.cell(row=fh, column=4))


def test_contadores(tmp_path):
    defectos = [{"pk_m": 3000, "ol_re": 95.7, "caracter": "CC"},
                {"pk_m": None, "ol_re": 40.0, "caracter": "CA"}]
    gen = _gen()
    gen.fill_dcvg(_postes(), defectos)
    gen.save(os.path.join(tmp_path, "c.xlsx"))
    assert gen.dcvg_sin_abscisa == 1     # marcados para completar
    assert gen.dcvg_omitidos == 0        # ya no se descarta nada
    assert gen.dcvg_filas == 4           # 2 postes + 2 defectos


def test_resistividad_sin_abscisa_resaltada(tmp_path):
    """Las resistividades sin PK ya salían al final; ahora además llevan la
    celda de abscisa resaltada para completarla."""
    resist = [
        {"pk_m": 5000, "sector": "Potrero", "profundidad": 190, "lat": 4.5,
         "lon": -75.7, "r1": 3.8, "r2": 2.3, "r3": 2.0},
        {"pk_m": None, "sector": "Sin pk", "profundidad": 180, "lat": 4.6,
         "lon": -75.8, "r1": 6.1, "r2": 3.2, "r3": 2.8},
    ]
    gen = _gen()
    gen.fill_resistividad(resist)
    ws = _guardar(gen, tmp_path, "r.xlsx")["Resistividad"]
    assert ws.cell(row=9, column=1).value == 5000
    assert ws.cell(row=10, column=1).value in (None, "")
    assert ws.cell(row=10, column=6).value == 6.1        # F R1: el dato está
    assert _es_amarilla(ws.cell(row=10, column=1))
    assert not _es_amarilla(ws.cell(row=9, column=1))
