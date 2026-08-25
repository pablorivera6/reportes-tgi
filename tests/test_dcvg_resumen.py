"""RESUMEN DE INDICACIONES y CLASIFICACIÓN RESISTIVIDAD (hoja Informe, DCVG).

La plantilla trae los rangos de las fórmulas escritos a mano y descuadrados
entre sí: V8:V37 en una celda y V8:V237 en la de al lado; U9:U25 junto a
U9:U244, U9:U246 y U9:U247. Con inspecciones largas, esas tablas contaban solo
una parte de la data.

Además dos criterios no coinciden con lo que el propio informe escribe:
  · carácter  'A-A'    → el informe escribe 'AA'      (contaba siempre 0)
  · severidad 'Mediana'→ el informe escribe 'Mediano' (contaba siempre 0)
"""
import os
import re

import openpyxl

from generator import ReportGenerator, resource_path

RANGO = re.compile(r"(?:'Inspección DCVG'|Resistividad)!(\$?[A-Z]{1,3}\$?)(\d+)"
                   r":(\$?[A-Z]{1,3}\$?)(\d+)")


def _postes(n=3):
    return [{"tipo": "Poste", "pk_m": i * 2000, "on": -1600.0, "off": -1100.0}
            for i in range(n)]


def _defectos(n=40):
    car = ["AA", "CA", "CC"]
    return [{"pk_m": 100 + i * 100, "ol_re": 50.0 + i * 8, "caracter": car[i % 3],
             "profundidad": 180} for i in range(n)]


def _resist(n=60):
    return [{"pk_m": i * 250, "r1": 3.8, "r2": 2.3, "r3": 2.0, "sector": "P"}
            for i in range(n)]


def _informe(tmp_path, n_def=40, n_res=60):
    gen = ReportGenerator(resource_path("DCVG_REP.xlsx"))
    postes, defectos, resist = _postes(), _defectos(n_def), _resist(n_res)
    gen.fill_dcvg(postes, defectos, resist)
    gen.fill_resistividad(resist)
    gen.ajustar_resumenes_dcvg(gen.dcvg_filas, len(resist))
    out = os.path.join(tmp_path, "r.xlsx")
    gen.save(out)
    return openpyxl.load_workbook(out)["Informe"], gen


def _formulas(ws, filas):
    for r in filas:
        for c in range(1, 40):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.startswith("="):
                yield ws.cell(row=r, column=c).coordinate, v


def test_rangos_del_resumen_cubren_toda_la_inspeccion(tmp_path):
    ws, gen = _informe(tmp_path)
    fin = 7 + gen.dcvg_filas          # los datos empiezan en la fila 8
    hallados = 0
    for coord, f in _formulas(ws, range(55, 58)):
        for ini_col, ini, fin_col, fin_txt in RANGO.findall(f):
            assert int(ini) == 8, f"{coord}: el rango empieza en {ini}"
            assert int(fin_txt) == fin, \
                f"{coord}: rango hasta {fin_txt}, la data llega a {fin}"
            hallados += 1
    assert hallados >= 20, "no se revisaron las fórmulas del resumen"


def test_rangos_de_resistividad_cubren_toda_la_data(tmp_path):
    ws, _ = _informe(tmp_path, n_res=60)
    fin = 8 + 60                      # los datos empiezan en la fila 9
    hallados = 0
    for coord, f in _formulas(ws, range(62, 68)):
        for _ic, ini, _fc, fin_txt in RANGO.findall(f):
            assert int(ini) == 9, f"{coord}: empieza en {ini}"
            assert int(fin_txt) == fin, \
                f"{coord}: rango hasta {fin_txt}, hay 60 resistividades"
            hallados += 1
    assert hallados >= 18


def test_criterio_del_caracter_aa(tmp_path):
    """El informe escribe 'AA'; la plantilla contaba 'A-A' (siempre 0)."""
    ws, _ = _informe(tmp_path)
    todas = " ".join(f for _c, f in _formulas(ws, range(55, 58)))
    assert '"A-A"' not in todas
    assert todas.count('"AA"') >= 4      # las 4 severidades de la fila AA


def test_criterio_de_severidad_mediana(tmp_path):
    """El informe escribe 'Mediano'; la plantilla contaba 'Mediana'."""
    ws, _ = _informe(tmp_path)
    todas = " ".join(f for _c, f in _formulas(ws, range(55, 58)))
    assert '"Mediana"' not in todas
    assert '"Mediano"' in todas
    # no se toca la corrosividad, que sí dice 'Medianamente'
    resist = " ".join(f for _c, f in _formulas(ws, range(62, 68)))
    assert "Medianamente corrosivo" in resist


def test_promedios_de_resistividad_tambien_se_ajustan(tmp_path):
    ws, _ = _informe(tmp_path, n_res=60)
    prom = ws["K67"].value
    assert "AVERAGE" in str(prom) and ":R68" in str(prom).replace("$", "")


def test_sin_datos_no_rompe(tmp_path):
    gen = ReportGenerator(resource_path("DCVG_REP.xlsx"))
    gen.fill_dcvg([], [])
    gen.ajustar_resumenes_dcvg(0, 0)
    out = os.path.join(tmp_path, "vacio.xlsx")
    gen.save(out)
    ws = openpyxl.load_workbook(out)["Informe"]
    assert str(ws["H55"].value).startswith("=COUNTIFS")
