"""PPM: el template debe estar limpio, generate() no debe arrastrar filas
viejas del template, y el PPM debe llevar la misma data del informe
(incluida la CIPS cuando el flujo es CIPS)."""
import os

import openpyxl

from generator import resource_path
from ppm_generator import PPMGenerator


def _fila_con_datos(ws, row):
    return any(ws.cell(row=row, column=c).value not in (None, '')
               for c in range(1, 19))


def test_template_ppm_sin_datos_viejos():
    wb = openpyxl.load_workbook(resource_path("PPM.XLSX"))
    ws = wb['CIPS - PAP']
    sucias = [r for r in range(2, ws.max_row + 1) if _fila_con_datos(ws, r)]
    assert not sucias, f"El template PPM trae datos viejos en filas {sucias}"


def test_generate_limpia_sobrantes_del_template(tmp_path):
    # Ensuciar una copia del template simula el archivo base con datos
    sucio = os.path.join(tmp_path, "PPM_sucio.xlsx")
    wb = openpyxl.load_workbook(resource_path("PPM.XLSX"))
    ws = wb['CIPS - PAP']
    for r in range(2, 15):
        ws.cell(row=r, column=1, value="R_VIEJO")
        ws.cell(row=r, column=7, value=r * 100)
    wb.save(sucio)

    pots = [{'abscisa': i * 50, 'on_mv': -1100, 'off_mv': -900,
             'lat': 4.0, 'lon': -73.0, 'observaciones': ''} for i in range(3)]
    out = os.path.join(tmp_path, "ppm_out.xlsx")
    PPMGenerator(sucio).generate({'route_id': 'R_NUEVO'}, pots, [], out)

    ws2 = openpyxl.load_workbook(out)['CIPS - PAP']
    assert ws2.cell(row=2, column=1).value == 'R_NUEVO'
    assert ws2.cell(row=4, column=1).value == 'R_NUEVO'
    restos = [r for r in range(5, 20) if _fila_con_datos(ws2, r)]
    assert not restos, f"Quedaron filas viejas del template: {restos}"


def test_generate_incluye_data_cips(tmp_path):
    cips = [{'abscisa_val': i * 10, 'on_mv': -1050.0, 'off_mv': -910.0,
             'on_limpio': -1050.0, 'off_limpio': -910.0,
             'lat': 4.1, 'lon': -73.1, 'observaciones': f'c{i}',
             'referencia': ''} for i in range(4)]
    out = os.path.join(tmp_path, "ppm_cips.xlsx")
    PPMGenerator().generate({'route_id': 'R_X'}, [], [], out, cips=cips)

    ws = openpyxl.load_workbook(out)['CIPS - PAP']
    absc = [ws.cell(row=r, column=7).value for r in range(2, 6)]
    assert absc == [0, 10, 20, 30], f"abscisas CIPS en PPM: {absc}"
    assert ws.cell(row=2, column=12).value == -1050.0   # On
    assert ws.cell(row=2, column=13).value == -910.0    # Off
    assert not _fila_con_datos(ws, 6)
