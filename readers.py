import openpyxl
from datetime import datetime
import re
from typing import List, Dict, Optional
import os
import pandas as pd

class FastFieldReader:
    def __init__(self):
        pass
        
    def _parse_abscisa(self, text: str) -> int:
        if not text: return 0
        match = re.search(r'(\d+)\+(\d+)', str(text))
        if match:
            return int(match.group(1)) * 1000 + int(match.group(2))
        try:
            return int(text)
        except ValueError:
            return 0

    def _parse_gps(self, text: str) -> tuple[Optional[float], Optional[float]]:
        if not text: return None, None
        parts = str(text).split(',')
        if len(parts) >= 2:
            try:
                return float(parts[0]), float(parts[1])
            except ValueError:
                pass
        return None, None

    def _safe_float(self, val) -> Optional[float]:
        if val is None or str(val).strip() == '': return None
        try:
            return float(str(val).replace(',', '.'))
        except ValueError:
            return None

    def _safe_int(self, val) -> Optional[int]:
        f = self._safe_float(val)
        return int(round(f)) if f is not None else None

    def _safe_mv(self, val) -> Optional[int]:
        f = self._safe_float(val)
        if f is not None:
            # If absolute value is between 0 and 10, it was likely entered in Volts instead of mV
            if 0 < abs(f) < 10:
                f = f * 1000
            return int(round(f))
        return None

    def read(self, filepath: str) -> dict:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        
        # Read general info from Root
        ws_root = wb['Root'] if 'Root' in wb.sheetnames else None
        tecnico, fecha = '', ''
        if ws_root:
            tecnico = str(ws_root.cell(row=2, column=7).value or '')
            fecha_val = ws_root.cell(row=2, column=9).value
            if isinstance(fecha_val, datetime):
                fecha = fecha_val.strftime('%d/%m/%Y')
            else:
                fecha = str(fecha_val or '')

        ws = wb['subform_1'] if 'subform_1' in wb.sheetnames else wb.active
        
        data = {
            'route_id': '',
            'contrato': '',
            'tramo': '',
            'tipo_tramo': '',
            'distrito': '',
            'tecnico': tecnico,
            'fecha': fecha,
            'potenciales': []
        }

        headers = {}
        on_cols = []
        off_cols = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            for idx, val in enumerate(row):
                if val is not None:
                    v = str(val).strip().lower()
                    # Remove special characters to make matching easier
                    v_clean = v.replace('á', 'a').replace('é', 'e').replace('í', 'i').replace('ó', 'o').replace('ú', 'u')
                    headers[v_clean] = idx
                    if 'on [mv]' in v:
                        on_cols.append(idx)
                    elif 'off [mv]' in v:
                        off_cols.append(idx)
            break

        is_old_format = 'engrouteid' in headers

        def get_val(row, *col_names):
            for name in col_names:
                if name in headers:
                    idx = headers[name]
                    if idx < len(row):
                        return row[idx]
            for name in col_names:
                for h_name, idx in headers.items():
                    if name in h_name:
                        if idx < len(row):
                            return row[idx]
            return None

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            if not any(row): continue
            
            import re
            
            raw_tramo = get_val(row, 'ramal en lista', 'tramo tgi', 'tramo')
            row_tramo = re.sub(r'\s*\(?PK.*', '', str(raw_tramo or '')).strip()
            row_route_id = str(get_val(row, 'engrouteid', 'route_id') or row_tramo)
            
            if i == 0:
                data['contrato'] = str(get_val(row, 'cliente', 'contrato') or '')
                data['tramo'] = row_tramo
                data['route_id'] = row_route_id
                data['tipo_tramo'] = str(get_val(row, 'tipo de tramo', 'tipo_tramo') or '')
                data['distrito'] = str(get_val(row, 'distrito') or '')
                
            ref = str(get_val(row, 'referencia geografica', 'direccion') or '').capitalize()
            abscisa_str = str(get_val(row, 'abscisado', 'abscisa') or '')
            lat, lon = self._parse_gps(get_val(row, 'localizacion gps'))
            
            # In new format, lat/lon might be in dedicated columns
            if lat is None or lon is None:
                lat = self._safe_float(get_val(row, 'latitud', 'lat'))
                lon = self._safe_float(get_val(row, 'longitud', 'lon'))
            
            # Fetch all 5 on/off pairs
            def get_on_off(idx):
                on_v = self._safe_mv(row[on_cols[idx]]) if idx < len(on_cols) else None
                off_v = self._safe_mv(row[off_cols[idx]]) if idx < len(off_cols) else None
                return on_v, off_v
            
            on_mv, off_mv = get_on_off(0)
            on_mv_corregido, off_mv_corregido = get_on_off(1)
            on_mv_neg2, off_mv_neg2 = get_on_off(2)
            on_mv_foraneo1, off_mv_foraneo1 = get_on_off(3)
            on_mv_foraneo2, off_mv_foraneo2 = get_on_off(4)
            
            pot_nat = self._safe_mv(get_val(row, 'potencial natural [mv]'))
            polar = self._safe_int(get_val(row, 'polarizacion [mv]'))
            vac = self._safe_float(get_val(row, 'voltaje ac', 'voltaje ac [mv]'))
            resistencia = self._safe_float(get_val(row, 'resistencia entre neg1-neg2 [ohm]'))
            ir_on_off_val = get_val(row, 'ir on-off [mv]')
            
            pintura = str(get_val(row, 'estado pintura') or '')
            conexiones = str(get_val(row, 'estado conexiones') or '')
            verticalidad = str(get_val(row, 'estado verticalidad') or '')
            observaciones = str(get_val(row, 'observaciones') or '')

            abscisa_val = self._parse_abscisa(abscisa_str)
            # Retornar el número tal cual para que el excel no reciba el texto con '+' y pueda aplicar su propio formato personalizado
            abscisa_str = str(abscisa_val) if abscisa_val else ''
            
            if vac is not None and vac > 1.0:
                vac = vac / 1000.0
                
            ir_on_off = None
            if on_mv is not None and off_mv is not None:
                ir_on_off = on_mv - off_mv
            elif ir_on_off_val is not None:
                ir_on_off = self._safe_int(ir_on_off_val)
                
            if 'Bueno' in conexiones: conexiones = 'Bueno [Mide Correctamente]'
            elif 'Malo' in conexiones: conexiones = 'Malo [No mide en los 2 cables]'
            elif 'Regular' in conexiones: conexiones = 'Regular [Mide en 1 sólo cable]'
            
            tipo_mant = 'NO APLICA'
            obs_lower = observaciones.lower()
            if 'corte concreto' in obs_lower or 'reposicion de concreto' in obs_lower or 'corte y reposicion' in obs_lower or 'corte de concreto' in obs_lower:
                tipo_mant = 'VI'
            elif 'hibrido' in obs_lower or 'híbrido' in obs_lower:
                tipo_mant = 'V'
            elif 'concreto' in obs_lower:
                tipo_mant = 'IV'
            elif 'poste nuevo' in obs_lower or 'nuevo poste' in obs_lower:
                tipo_mant = 'III'
            elif 'malo' in conexiones.lower() or 'regular' in conexiones.lower() or 'caido' in verticalidad.lower() or 'caído' in verticalidad.lower():
                tipo_mant = 'II'
            elif 'malo' in pintura.lower() or 'regular' in pintura.lower():
                tipo_mant = 'I'
                
            if ref == 'Poste de potencial': ref = 'Poste de Potencial'
            if ref == 'Poste abscisado': ref = 'Poste de Abscisado'

            data['potenciales'].append({
                'tramo': row_tramo,
                'route_id': row_route_id,
                'abscisa': abscisa_val,
                'abscisa_str': abscisa_str,
                'fecha': fecha,
                'ref_geografica': ref,
                'on_mv': on_mv,
                'off_mv': off_mv,
                'on_mv_corregido': on_mv_corregido,
                'off_mv_corregido': off_mv_corregido,
                'on_mv_neg2': on_mv_neg2,
                'off_mv_neg2': off_mv_neg2,
                'on_mv_foraneo1': on_mv_foraneo1,
                'off_mv_foraneo1': off_mv_foraneo1,
                'on_mv_foraneo2': on_mv_foraneo2,
                'off_mv_foraneo2': off_mv_foraneo2,
                'potencial_natural': pot_nat,
                'polarizacion': polar,
                'vac': vac,
                'resistencia': resistencia,
                'ir_on_off': ir_on_off,
                'lat': lat,
                'lon': lon,
                'pintura': pintura,
                'conexiones': conexiones,
                'verticalidad': verticalidad,
                'tipo_mant': tipo_mant,
                'observaciones': observaciones
            })
            
        data['potenciales'].sort(key=lambda x: x['abscisa'])
        return data


class EquipoReader:
    def __init__(self):
        pass

    def read(self, filepath: str) -> dict:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        
        info = {
            'pipeline': '', 'work_order': '', 'technician': '', 'date': '',
            'survey_name': '', 'cycle_on_ms': 0, 'cycle_off_ms': 0, 'start_station': 0
        }
        if 'Survey Info' in wb.sheetnames:
            ws_info = wb['Survey Info']
            for row in ws_info.iter_rows(min_row=2, values_only=True):
                k, v = str(row[0] or '').strip(), row[1]
                if k == 'Name of P/L': info['pipeline'] = str(v or '')
                elif k == 'Work Order #': info['work_order'] = str(v or '')
                elif k == 'Technician Name': info['technician'] = str(v or '')
                elif k == 'date / time': info['date'] = str(v or '')
                elif k == 'SurveyName': info['survey_name'] = str(v or '')
                elif k == 'Rectifier Cycle On Time': info['cycle_on_ms'] = int(v or 0)
                elif k == 'Rectifier Cycle Off Time': info['cycle_off_ms'] = int(v or 0)
                elif k == 'Survey Location':
                    try: info['start_station'] = int(str(v or '0').replace('+',''))
                    except: pass
                    
        hallazgos = []
        dcp_points = []
        
        if 'DCP Data' in wb.sheetnames:
            ws_dcp = wb['DCP Data']
            for row in ws_dcp.iter_rows(min_row=2, values_only=True):
                if not any(row): continue
                
                station = 0
                try: station = int(row[1] or 0)
                except: pass
                
                comment = str(row[6] or '').strip()
                lat = float(row[9] or 0) if row[9] else None
                lon = float(row[10] or 0) if row[10] else None
                alt = float(row[11] or 0) if row[11] else None
                
                is_hallazgo = False
                if comment and lat and lon:
                    comment_lower = comment.lower()
                    skip = bool(re.search(r'pk\s*\d+\+\d+', comment_lower) or 'poste' in comment_lower)
                    
                    if not skip:
                        parts = comment_lower.split(' y ')
                        for part in parts:
                            tipo = ''
                            if 'via' in part or 'vía' in part: tipo = 'Cruce de Vía'
                            elif 'rio' in part or 'río' in part: tipo = 'Cruce de Río'
                            elif 'caño' in part: tipo = 'Caño'
                            elif 'at' in part or 'alta tension' in part: tipo = 'Línea AT (Alta Tensión)'
                            
                            if tipo:
                                is_hallazgo = True
                                hallazgos.append({
                                    'tipo': tipo,
                                    'descripcion': part.capitalize(),
                                    'lat': lat,
                                    'lon': lon,
                                    'alt': alt,
                                    'abscisa': None,
                                    'fecha': info['date']
                                })
                
                dcp_points.append({
                    'station': station,
                    'comment': comment,
                    'lat': lat,
                    'lon': lon,
                    'alt': alt,
                    'is_hallazgo': is_hallazgo
                })
                
        return {
            'survey_info': info,
            'hallazgos': hallazgos,
            'dcp_points': dcp_points
        }


class RectificadorReader:
    def __init__(self):
        pass

    def read(self, filepath: str) -> dict:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        # Find URPC sheet
        ws = None
        for name in wb.sheetnames:
            if 'URPC' in name.upper():
                ws = wb[name]
                break
        
        if not ws:
            return {}
            
        data = {
            'nombre': str(ws.cell(row=6, column=3).value or ''),
            'gasoducto': str(ws.cell(row=6, column=8).value or ''),
            'modelo': str(ws.cell(row=10, column=7).value or ''),
            'voltaje_nominal': ws.cell(row=10, column=12).value,
            'corriente_nominal': ws.cell(row=11, column=12).value,
            'taps_gruesos': str(ws.cell(row=11, column=7).value or ''),
            'taps_finos': str(ws.cell(row=12, column=7).value or ''),
            'serial': str(ws.cell(row=15, column=3).value or ''),
            'ultima_inspeccion': {},
            'conexion_estructura': {}
        }
        
        # Find last operation row
        last_op_row = None
        for row in range(20, 30):
            if ws.cell(row=row, column=1).value:
                last_op_row = row
            else:
                break
                
        if last_op_row:
            data['ultima_inspeccion'] = {
                'fecha': str(ws.cell(row=last_op_row, column=1).value or ''),
                'taps': str(ws.cell(row=last_op_row, column=6).value or ''),
                'vac_entrada': ws.cell(row=last_op_row, column=7).value,
                'vdc_salida': ws.cell(row=last_op_row, column=9).value,
                'idc_salida': ws.cell(row=last_op_row, column=10).value,
                'eficiencia': ws.cell(row=last_op_row, column=14).value,
                'disponibilidad_v': ws.cell(row=last_op_row, column=15).value,
                'disponibilidad_i': ws.cell(row=last_op_row, column=15).value, # Default same
            }
            
        # Find last connection row
        last_conn_row = None
        for row in range(33, 40):
            if ws.cell(row=row, column=2).value:
                last_conn_row = row
            else:
                break
                
        if last_conn_row:
            data['conexion_estructura'] = {
                'estructura': str(ws.cell(row=last_conn_row, column=2).value or ''),
                'corriente': ws.cell(row=last_conn_row, column=4).value,
                'pot_on': ws.cell(row=last_conn_row, column=5).value,
                'pot_off': ws.cell(row=last_conn_row, column=6).value,
            }
            
        return data

class AislamientoReader:
    def __init__(self):
        pass

    def read_files(self, filepaths: list) -> list:
        aislamientos = []
        try:
            for filepath in filepaths:
                if filepath.endswith('.xlsx') and not os.path.basename(filepath).startswith('~'):

                    try:
                        df = pd.read_excel(filepath)
                        if df.empty: continue
                        row = df.iloc[0]
                        
                        def get_val(col_name):
                            if col_name in df.columns:
                                val = row[col_name]
                                if pd.isna(val): return ""
                                return val
                            return ""
                            
                        # Desglosar ubicacion GPS
                        lat, lon = "", ""
                        ubicacion = str(get_val('Ubicacion GPS'))
                        if ',' in ubicacion:
                            parts = ubicacion.split(',')
                            lat = parts[0].strip()
                            if len(parts) > 1:
                                lon = parts[1].strip()

                        # Calcular diferencia si no viene
                        pot_on_arriba = get_val('Potencial On (mVcse)')
                        pot_on_abajo = get_val('Potencial On (mVcse).1')
                        diff = get_val('Diferencia aguas arriba-aguas abajo')
                        
                        if diff == "":
                            try:
                                if pot_on_arriba != "" and pot_on_abajo != "":
                                    diff = abs(float(pot_on_arriba) - float(pot_on_abajo))
                            except:
                                pass
                                
                        aislamiento = {
                            'abscisado': get_val('Abscisado'),
                            'tag': get_val('Tag'),
                            'clase': get_val('Class (ANSI B 16.5)'),
                            'diametro': get_val('Diametro Nominal (Pulgadas)'),
                            'presion': get_val('Presion (PSI)'),
                            'temperatura': get_val('Temperatura (C)'),
                            'tipo_brida': get_val('Tipo de brida '),
                            'numero_pernos': get_val('Numero de pernos'),
                            'diametro_pernos': get_val('Diametro pernos (pulgadas)'),
                            'tipo_aislamiento': get_val('Tipo de aislamiento electrico'),
                            'porcentaje_aislamiento': get_val('% Aislamiento electrico entre caras'),
                            'pot_on_arriba': pot_on_arriba,
                            'pot_off_arriba': get_val('Potencial Off (mVcse)'),
                            'pot_on_abajo': pot_on_abajo,
                            'pot_off_abajo': get_val('Potencial Off (mVcse).1'),
                            'diferencia': diff,
                            'diagnostico': get_val('Diagnostico'),
                            'latitud': lat,
                            'longitud': lon,
                            'observaciones': get_val('Observaciones')
                        }
                        aislamientos.append(aislamiento)
                    except Exception as e:
                        print(f"Error procesando archivo de aislamiento {filepath}: {e}")
        except Exception as e:
            import traceback
            print(f"Error leyendo aislamientos: {e}")
            traceback.print_exc()
            
        return aislamientos

    def read_folder(self, folder_path: str) -> list:
        filepaths = []
        for filename in os.listdir(folder_path):
            if filename.endswith('.xlsx') and not filename.startswith('~'):
                filepaths.append(os.path.join(folder_path, filename))
        return self.read_files(filepaths)
