import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import os
from generator import resource_path
from ortografia import corregir_campo

class PPMGenerator:
    def __init__(self, template_path: str = None):
        if template_path is None:
            template_path = resource_path("PPM.XLSX")
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"No se encontró la plantilla PPM en: {template_path}")
        self.template_path = template_path
        self.wb = openpyxl.load_workbook(template_path)
        self.ws = self.wb['CIPS - PAP'] if 'CIPS - PAP' in self.wb.sheetnames else self.wb.active

    def _safe_write(self, row, col, value):
        try:
            cell = self.ws.cell(row=row, column=col)
            cell.value = value
            # Optionally set a basic style here if needed
        except Exception as e:
            pass

    def _limpiar_filas_datos(self):
        """Borra cualquier dato remanente del template (filas 2+): si el
        archivo base se guardó con datos de una inspección anterior, no deben
        arrastrarse al PPM generado."""
        for row in self.ws.iter_rows(min_row=2, max_row=self.ws.max_row,
                                     min_col=1, max_col=18):
            for cell in row:
                if cell.value is not None:
                    cell.value = None

    def generate(self, info: dict, potenciales: list, aislamientos: list,
                 output_path: str, cips: list = None):
        # The template has headers in row 1, data starts at row 2
        self._limpiar_filas_datos()

        # We need to map everything into rows
        current_row = 2

        # 1. Add potential points
        for pot in potenciales:
            self._write_row(current_row, info, pot)
            current_row += 1

        # 1b. Add CIPS points (mismo contenido que la hoja CIPS del informe)
        for c in (cips or []):
            on = c.get('on_mv')
            off = c.get('off_mv')
            ir = None
            if c.get('on_limpio') is not None and c.get('off_limpio') is not None:
                ir = c['on_limpio'] - c['off_limpio']
            elif on is not None and off is not None:
                ir = on - off
            self._write_row(current_row, info, {
                'abscisa': c.get('abscisa_val'),
                'lat': c.get('lat', ''),
                'lon': c.get('lon', ''),
                'alt': '',
                'on_mv': on,
                'off_mv': off,
                'potencial_natural': None,
                'polarizacion': None,
                'observaciones': c.get('observaciones', ''),
                'ir_on_off': ir,
                'fecha': c.get('fecha'),
            })
            current_row += 1

        # 2. Add isolation points
        for ais in aislamientos:
            pot_pseudo = {
                'abscisa': ais.get('abscisado', 0),
                'lat': ais.get('latitud', ''),
                'lon': ais.get('longitud', ''),
                'alt': '',
                'on_mv': ais.get('pot_on_arriba'),
                'off_mv': ais.get('pot_off_arriba'),
                'potencial_natural': None,
                'polarizacion': None,
                'observaciones': ais.get('observaciones', ''),
                'ir_on_off': None
            }
            # Compute IR drop if both exist
            if pot_pseudo['on_mv'] is not None and pot_pseudo['off_mv'] is not None:
                try:
                    pot_pseudo['ir_on_off'] = float(pot_pseudo['on_mv']) - float(pot_pseudo['off_mv'])
                except:
                    pass
                    
            self._write_row(current_row, info, pot_pseudo)
            current_row += 1
            
        self.wb.save(output_path)

    def _write_row(self, row, info, data):
        # ['ENGROUTEID', 'No Contrato', 'Distrito', 'Tipo de Tramo', 'Tramo', 
        #  'Fecha de Inspección', 'ABCISA', 'Center line', 'Latitud', 'Longitud', 
        #  'Altitud', 'P On mV', 'P Off mV', 'P Natural mV', 'Polarizacion mV', 
        #  'Dirección de la inspección', 'Comentario', 'IR On-Off mV']
        
        self._safe_write(row, 1, info.get('route_id', ''))
        self._safe_write(row, 2, info.get('contrato', ''))
        self._safe_write(row, 3, info.get('distrito', ''))
        self._safe_write(row, 4, info.get('tipo_ducto', ''))
        self._safe_write(row, 5, info.get('tramo', ''))
        # Fecha del día en que se tomó el punto (si viene por dato); si no, la
        # fecha general del informe.
        self._safe_write(row, 6, data.get('fecha') or info.get('fecha', ''))
        
        # ABCISA is usually an integer in meters? We'll write exactly what's there (should be integer normally)
        abscisa = data.get('abscisa', '')
        if isinstance(abscisa, str):
            # If abscisa is "km+m", let's extract the integer meter value if possible, or leave as is
            pass
        self._safe_write(row, 7, abscisa)
        
        self._safe_write(row, 8, None) # Center line
        self._safe_write(row, 9, data.get('lat', ''))
        self._safe_write(row, 10, data.get('lon', ''))
        self._safe_write(row, 11, data.get('alt', ''))
        self._safe_write(row, 12, data.get('on_mv'))
        self._safe_write(row, 13, data.get('off_mv'))
        self._safe_write(row, 14, data.get('potencial_natural'))
        self._safe_write(row, 15, data.get('polarizacion'))
        
        # Default Ascendente for now?
        self._safe_write(row, 16, 'Ascendente')
        
        self._safe_write(row, 17, corregir_campo(data.get('observaciones', '')))
        self._safe_write(row, 18, data.get('ir_on_off'))


class PPMDcvgGenerator:
    """PPM del informe DCVG (plantilla `DCVG_PPM_.xlsx`).

    Dos hojas:
      · `DCVG`         — un renglón por registro del recorrido, ordenados por
        abscisa: postes (P On/P Off), defectos (P_RE, OL_RE, PORC_IR, carácter,
        clasificación, profundidad) y hallazgos (solo su descripción).
      · `RESISTIVIDAD` — un renglón por medida Wenner, con las resistividades ya
        calculadas (las mismas fórmulas de la hoja Resistividad del informe).

    `PORC_IR` va en FRACCIÓN porque la celda tiene formato `0.00%`, igual que la
    severidad del informe. Altitud y 'Cama anódica temporal' quedan vacías: el
    FastField no captura esos datos.
    """
    HOJA_DCVG = 'DCVG'
    HOJA_RESIST = 'RESISTIVIDAD'
    HOJA_LISTAS = 'Hoja2'
    COLS_DCVG = 22
    COLS_RESIST = 18

    def __init__(self, template_path: str = None):
        template_path = template_path or resource_path("DCVG_PPM_.xlsx")
        if not os.path.exists(template_path):
            raise FileNotFoundError(
                f"No se encontró la plantilla del PPM de DCVG en: {template_path}")
        self.wb = openpyxl.load_workbook(template_path)

    # ── helpers ──────────────────────────────────────────────────────────────
    @staticmethod
    def _w(ws, row, col, value):
        try:
            ws.cell(row=row, column=col).value = value
        except (AttributeError, ValueError):
            pass

    def _limpiar(self, ws, columnas):
        """Quita los datos que la plantilla pudiera traer de otra inspección."""
        ultima = min(ws.max_row, 5000)
        for fila in ws.iter_rows(min_row=2, max_row=ultima, min_col=1,
                                 max_col=columnas):
            for celda in fila:
                if celda.value is not None:
                    celda.value = None

    def engrouteid(self, tramo):
        """Código de ruta (R_ANS, T_NRSA…). Se busca en el catálogo que la
        propia plantilla trae en `Hoja2`; si no está, se arma con la sigla del
        listado de infraestructura."""
        from nombres import mismo_tramo, sigla_tramo
        if self.HOJA_LISTAS in self.wb.sheetnames:
            ws = self.wb[self.HOJA_LISTAS]
            for r in range(2, ws.max_row + 1):
                nombre = ws.cell(row=r, column=4).value
                codigo = ws.cell(row=r, column=5).value
                if not codigo:
                    continue
                if nombre and mismo_tramo(tramo, nombre):
                    return str(codigo).strip()
        sigla, letra, hallado = sigla_tramo(tramo)
        return f"{letra}_{sigla}" if (hallado and letra) else ""

    def _cabecera(self, ws, row, info, fecha=None):
        self._w(ws, row, 1, self.engrouteid(info.get('tramo', '')))
        self._w(ws, row, 2, info.get('contrato', ''))
        self._w(ws, row, 3, info.get('distrito', ''))
        self._w(ws, row, 4, info.get('tipo_ducto', ''))
        self._w(ws, row, 5, info.get('tramo', ''))
        self._w(ws, row, 6, fecha or info.get('fecha', ''))

    # ── generación ───────────────────────────────────────────────────────────
    def generate(self, info: dict, postes: list, defectos: list,
                 resistividades: list = None, hallazgos: list = None,
                 output_path: str = None):
        self._llenar_dcvg(info, postes or [], defectos or [], hallazgos or [])
        self._llenar_resistividad(info, resistividades or [])
        if output_path:
            self.wb.save(output_path)
        return output_path

    def _llenar_dcvg(self, info, postes, defectos, hallazgos):
        from generator import ReportGenerator
        import db as _db

        ws = self.wb[self.HOJA_DCVG]
        self._limpiar(ws, self.COLS_DCVG)

        sev = {id(d): s for d, s in
               zip(defectos, _db._severidad_dcvg(postes, defectos))}

        # mismo orden que la hoja 'Inspección DCVG' del informe: los registros
        # sin abscisa se anclan al vecino del archivo de campo
        filas, seq = [], 0
        for tipo, items, clave in (('poste', postes, 'pk_m'),
                                   ('defecto', defectos, 'pk_m'),
                                   ('hallazgo', hallazgos, 'abscisa_val')):
            for ancla, pos, it in ReportGenerator._anclar_sin_abscisa(items, clave):
                filas.append((ancla, pos, {'poste': 0, 'defecto': 1,
                                           'hallazgo': 2}[tipo], seq, tipo, it))
                seq += 1
        filas.sort(key=lambda f: f[:4])
        self.n_dcvg = len(filas)

        for i, (_a, _p, _t, _s, tipo, r) in enumerate(filas):
            row = 2 + i
            absc = r.get('pk_m') if tipo != 'hallazgo' else r.get('abscisa_val')
            self._cabecera(ws, row, info, r.get('fecha'))
            self._w(ws, row, 7, absc)
            self._w(ws, row, 9, r.get('lat'))
            self._w(ws, row, 10, r.get('lon'))
            if i == 0:
                self._w(ws, row, 12, "Inicio Inspección")
            elif i == len(filas) - 1:
                self._w(ws, row, 12, "Fin Inspección")

            if tipo == 'poste':
                self._w(ws, row, 13, r.get('on'))
                self._w(ws, row, 14, r.get('off'))
                self._w(ws, row, 22, corregir_campo(r.get('tipo', '')))
            elif tipo == 'defecto':
                s = sev.get(id(r), {})
                self._w(ws, row, 15, s.get('p_re'))
                self._w(ws, row, 16, r.get('ol_re'))
                pct = s.get('severidad_pct')
                if pct is not None:
                    self._w(ws, row, 17, pct / 100.0)   # la celda es 0.00%
                self._w(ws, row, 19, r.get('caracter', ''))
                self._w(ws, row, 20, s.get('clasificacion') or
                        r.get('clasificacion_campo', ''))
                self._w(ws, row, 21, r.get('profundidad'))
                self._w(ws, row, 22, corregir_campo(r.get('comentarios', '')))
            else:
                self._w(ws, row, 22, corregir_campo(
                    r.get('descripcion') or r.get('observaciones') or ''))

    def _llenar_resistividad(self, info, resistividades):
        import math
        ws = self.wb[self.HOJA_RESIST]
        self._limpiar(ws, self.COLS_RESIST)
        datos = sorted([r for r in resistividades if r.get('pk_m') is not None],
                       key=lambda r: r['pk_m'])
        datos += [r for r in resistividades if r.get('pk_m') is None]
        self.n_resist = len(datos)

        def _capa(ra, rb, a_cm):
            if ra is None or rb is None or abs(rb - ra) == 0:
                return None
            return ((ra * rb) / abs(rb - ra)) * 2 * math.pi * a_cm

        for i, r in enumerate(datos):
            row = 2 + i
            self._cabecera(ws, row, info)
            self._w(ws, row, 7, r.get('pk_m'))
            self._w(ws, row, 9, r.get('lat'))
            self._w(ws, row, 10, r.get('lon'))
            r1, r2, r3 = r.get('r1'), r.get('r2'), r.get('r3')
            for col, (res, a_cm) in zip((12, 13, 14),
                                        ((r1, 100), (r2, 200), (r3, 300))):
                if res is not None:
                    self._w(ws, row, col, 2 * math.pi * res * a_cm)
            if r1 is not None:
                self._w(ws, row, 15, 2 * math.pi * r1 * 100)   # capa 0-1 m
            self._w(ws, row, 16, _capa(r1, r2, 100))           # capa 1-2 m
            self._w(ws, row, 17, _capa(r2, r3, 100))           # capa 2-3 m
            self._w(ws, row, 18, corregir_campo(r.get('sector', '')))
