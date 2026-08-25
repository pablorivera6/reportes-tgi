"""
TGI Report Generator - Motor de generación de informes Excel
Llena la plantilla EN BLANCO.xlsx con los datos procesados
"""
import openpyxl
import re
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from copy import copy
from datetime import datetime
from ortografia import corregir_campo
from typing import Optional
import os
import sys
import math


def resource_path(relative_path: str) -> str:
    """Get absolute path to resource, works for dev and for PyInstaller"""
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), relative_path)


class ReportGenerator:
    """Generates inspection reports by filling the EN BLANCO.xlsx template"""

    def __init__(self, template_path: Optional[str] = None):
        """Initializes the generator with a template"""
        if template_path is None:
            # Revert to standard EN BLANCO.xlsx which now contains the Potenciales CIPS sheet manually
            template_path = resource_path("EN BLANCO.xlsx")
        self.wb = openpyxl.load_workbook(template_path)
        # Algunas plantillas (p.ej. DCVG) no traen todas las hojas; tolerar.
        def _hoja(nombre):
            return self.wb[nombre] if nombre in self.wb.sheetnames else None
        self.ws_informe = _hoja('Informe')
        self.ws_potenciales = _hoja('Potenciales PAP')
        self.ws_hallazgos = _hoja('Hallazgos')
        # Gráficas - handle encoding variations
        self._init_sheet_refs()

    def _init_sheet_refs(self):
        """Initialize sheet references handling encoding"""
        sheet_map = {}
        for name in self.wb.sheetnames:
            sheet_map[name.lower()] = name
        
        def find_sheet(keywords):
            for name, original in [(n.lower(), n) for n in self.wb.sheetnames]:
                if all(k.lower() in name for k in keywords):
                    return self.wb[original]
            return None

        self.ws_grafica_vdc = find_sheet(['fica', 'vdc']) or find_sheet(['fica', 'VDC'])
        self.ws_grafica_interf = find_sheet(['fica', 'nterferencia'])
        self.ws_grafica_vac = find_sheet(['fica', 'vac']) or find_sheet(['fica', 'VAC'])
        self.ws_marco_h = find_sheet(['marco', 'h'])
        self.ws_ce = find_sheet(['ce'])
        self.ws_anodos = find_sheet(['nodos'])
        self.ws_aislamientos = find_sheet(['aislamiento'])
        self.ws_cupones_ir = find_sheet(['cupones', 'ir', 'free'])
        self.ws_cupones_grav = find_sheet(['cupones', 'gravim'])
        self.ws_pe = find_sheet(['pe'])
        self.ws_tramos_aereos = find_sheet(['tramos', 'reos'])
        self.ws_tramos_no_insp = find_sheet(['tramos', 'no'])

    def _safe_write(self, ws, row: int, col: int, value):
        """Write value to cell, preserving formatting of merged cells"""
        try:
            cell = ws.cell(row=row, column=col)
            cell.value = value
        except (AttributeError, ValueError):
            pass

    # Resaltado de las celdas que el ingeniero debe completar a mano (una
    # abscisa que el técnico no registró en campo).
    _COLOR_POR_COMPLETAR = "FFFFFF00"

    def _marcar_por_completar(self, ws, row: int, col: int):
        """Pinta de amarillo una celda que quedó vacía a propósito."""
        try:
            ws.cell(row=row, column=col).fill = PatternFill(
                start_color=self._COLOR_POR_COMPLETAR,
                end_color=self._COLOR_POR_COMPLETAR, fill_type="solid")
        except (AttributeError, ValueError):
            pass

    @staticmethod
    def _anclar_sin_abscisa(items: list, clave: str) -> list:
        """[(ancla, pos, registro)] para ordenar por abscisa SIN perder los
        registros que no la traen.

        El archivo de campo va en el orden en que el técnico hizo el recorrido,
        así que un registro sin abscisa va donde su vecino: se ancla a la
        abscisa del anterior que sí la tenga (pos=1 → queda justo después) o,
        si es el primero, a la del siguiente (pos=-1 → justo antes). pos=0 = el
        registro sí trae abscisa.
        """
        vals = [it.get(clave) for it in items]
        out = []
        for i, v in enumerate(vals):
            if v is not None:
                out.append((v, 0, items[i]))
                continue
            prev = next((x for x in reversed(vals[:i]) if x is not None), None)
            if prev is not None:
                out.append((prev, 1, items[i]))
            else:
                sig = next((x for x in vals[i + 1:] if x is not None), None)
                out.append((sig if sig is not None else 0, -1, items[i]))
        return out

    def _copy_row_style(self, ws, source_row: int, target_row: int, min_col: int, max_col: int):
        """Copy cell styles from source row to target row to preserve template formatting"""
        for col in range(min_col, max_col + 1):
            source_cell = ws.cell(row=source_row, column=col)
            target_cell = ws.cell(row=target_row, column=col)
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.alignment = copy(source_cell.alignment)

    # ── Distribución de la hoja 'Informe' ────────────────────────────────
    # La plantilla DCVG coloca las secciones y las columnas de valor en sitios
    # distintos a las de PAP/CIPS. En vez de quemar filas y columnas, se
    # localizan por la etiqueta y por las celdas combinadas.

    SECCIONES_INFORME = ('OBJETIVO', 'DOCUMENTOS DE REFERENCIA',
                         'EQUIPOS UTILIZADOS',
                         'DESCRIPCIÓN DE LA LÍNEA OBJETO DE ESTUDIO',
                         'ANTECEDENTES', 'SISTEMA INSPECCIONADO',
                         'CONCLUSIONES', 'RECOMENDACIONES')

    @staticmethod
    def _txt(v):
        return re.sub(r'\s+', ' ', str(v or '')).strip().upper()

    @staticmethod
    def _etiqueta(v):
        """Etiqueta comparable: sin tildes, minúsculas, sin espacios de más."""
        import unicodedata
        t = ''.join(c for c in unicodedata.normalize('NFD', str(v or ''))
                    if unicodedata.category(c) != 'Mn')
        return re.sub(r'\s+', ' ', t).strip().lower()

    # Etiqueta del encabezado -> campo de Datos Generales. El ORDEN importa:
    # 'Fecha calibración' antes que 'Fecha' y 'Contratista' antes que
    # 'contrato'. Se mapea por etiqueta y no por fila porque la plantilla DCVG
    # tiene Contratista y OT en filas cambiadas frente a PAP/CIPS.
    CAMPOS_ENCABEZADO = (
        (r'^fecha\s*calibraci', 'fecha_calibracion'),
        (r'^fecha', 'fecha'),
        (r'^serial', 'serial_equipo'),
        (r'^contratista', 'contratista'),
        (r'contrato', 'contrato'),
        (r'^(no\.?\s*de\s*)?ot\b', 'ot'),
        (r'^gasoducto', 'gasoducto'),
        (r'^tramo', 'tramo'),
        (r'recubrimiento', 'tipo_recubrimiento'),
        (r'^inspector', 'inspector'),
        (r'^diametro', 'diametro'),
        (r'^ciclo', 'ciclo'),
    )

    @classmethod
    def _campo_de_etiqueta(cls, etiqueta):
        e = cls._etiqueta(etiqueta)
        if not e:
            return None
        for patron, campo in cls.CAMPOS_ENCABEZADO:
            if re.search(patron, e):
                return campo
        return None

    def _fila_seccion(self, ws, etiqueta, hasta=80):
        """Fila del título de sección (col A). None si no está."""
        objetivo = self._txt(etiqueta)
        for r in range(1, min(ws.max_row, hasta) + 1):
            if self._txt(ws.cell(row=r, column=1).value) == objetivo:
                return r
        return None

    def _bloque_seccion(self, ws, etiqueta, hasta=80):
        """(primera, última) fila de contenido de una sección, o None."""
        ini = self._fila_seccion(ws, etiqueta, hasta)
        if ini is None:
            return None
        otras = [self._fila_seccion(ws, s, hasta) for s in self.SECCIONES_INFORME]
        siguientes = [r for r in otras if r is not None and r > ini]
        fin = (min(siguientes) - 1) if siguientes else min(ws.max_row, ini + 6)
        return (ini + 1, fin) if fin >= ini + 1 else None

    def _fila_firmas(self, ws, hasta=140):
        """Fila del bloque de firmas ('ELABORÓ'), en cualquier columna."""
        for r in range(1, min(ws.max_row, hasta) + 1):
            for c in range(1, 30):
                if self._etiqueta(ws.cell(row=r, column=c).value) == 'elaboro':
                    return r
        return None

    def _bloque_texto(self, ws, etiqueta, hasta=140):
        """(primera, última) fila donde cabe el texto de CONCLUSIONES o
        RECOMENDACIONES. Cada plantilla las tiene en filas distintas (DCVG 69 y
        84; PAP 83 y 94; CIPS 81 y 90), así que se ubican por el título y se
        acotan con la sección siguiente o con el bloque de firmas."""
        ini = self._fila_seccion(ws, etiqueta, hasta)
        if ini is None:
            return None
        siguientes = [r for r in
                      [self._fila_seccion(ws, s, hasta) for s in self.SECCIONES_INFORME]
                      + [self._fila_firmas(ws, hasta)]
                      if r is not None and r > ini]
        fin = (min(siguientes) - 1) if siguientes else min(ws.max_row, ini + 10)
        return (ini + 1, fin) if fin >= ini + 1 else None

    def _cols_valor(self, ws, fila, defecto=(7, 22, 32)):
        """Columnas de VALOR de una fila del encabezado. La fila alterna
        etiqueta/valor en celdas combinadas: PAP/CIPS usa G/V/AF y DCVG G/U/AE,
        así que se toman las anclas de merge en posición impar."""
        anclas = sorted({m.min_col for m in ws.merged_cells.ranges
                         if m.min_row == fila and m.max_row == fila})
        if len(anclas) >= 6:
            return (anclas[1], anclas[3], anclas[5])
        return defecto

    # Objetivo del informe DCVG: la plantilla lo trae como texto fijo con el
    # ramal del ejemplo (a diferencia de PAP/CIPS, que lo arman por fórmula).
    OBJETIVO_DCVG = (
        "Realizar un diagnóstico del estado general del recubrimiento, bajo la "
        "técnica de inspección DCVG para la tubería enterrada e inspección "
        "visual para la tubería aérea {linea}. Clasificar puntos estratégicos "
        "susceptibles a falla a corto plazo del recubrimiento, para definir "
        "programas de mantenimiento y reparación.")

    def _fill_objetivo_dcvg(self, ws, data):
        """Reescribe el OBJETIVO con el tramo/gasoducto/distrito de ESTA
        inspección. Solo aplica cuando es texto fijo (DCVG); si la plantilla lo
        calcula por fórmula (PAP/CIPS) no se toca."""
        bloque = self._bloque_seccion(ws, 'OBJETIVO')
        if not bloque:
            return
        fila = next((r for r in range(bloque[0], bloque[1] + 1)
                     if ws.cell(row=r, column=1).value not in (None, '')), None)
        actual = ws.cell(row=fila, column=1).value if fila else None
        if not isinstance(actual, str) or actual.startswith('='):
            return                      # objetivo por fórmula: es del formato
        tramo = str(data.get('tramo') or '').strip()
        if not tramo:
            return
        linea = f"del {tramo}"
        gas = str(data.get('gasoducto') or '').strip()
        if gas:
            linea += f", el cual hace parte del gasoducto {gas}"
        distrito = re.sub(r'^[Dd]0*', '', str(data.get('distrito') or '').strip())
        if distrito:
            linea += f" perteneciente al distrito {distrito}"
        self._safe_write(ws, fila, 1, self.OBJETIVO_DCVG.format(linea=linea))

    def _fill_descripcion_linea(self, ws, data):
        """Redacta la DESCRIPCIÓN DE LA LÍNEA OBJETO DE ESTUDIO del informe
        DCVG, que en esa plantilla viene vacía. Sigue la estructura del formato
        PAP. Solo para DCVG: en PAP/CIPS ese párrafo lo escribe el ingeniero."""
        if self._etiqueta(data.get('tipo_inspeccion')) != 'dcvg':
            return
        tramo = str(data.get('tramo') or '').strip()
        if not tramo:
            return
        bloque = self._bloque_seccion(ws, 'DESCRIPCIÓN DE LA LÍNEA OBJETO DE ESTUDIO')
        if not bloque:
            return
        partes = [f"El {tramo}"]
        gas = str(data.get('gasoducto') or '').strip()
        if gas:
            partes[0] += f" perteneciente al Gasoducto {gas}"
        try:
            km = float(data.get('longitud_km') or 0)
        except (TypeError, ValueError):
            km = 0
        if km > 0:
            partes[0] += (f", cuenta con una longitud de {self._num_es(km, 1)} Km "
                          f"aproximadamente")
        partes[0] += "."

        rec = str(data.get('tipo_recubrimiento') or '').strip()
        diam = re.sub(r'[^0-9.,/]', '', str(data.get('diametro') or '')).strip(' .,')
        detalle = []
        if rec:
            detalle.append(f"un recubrimiento {rec}")
        if diam:
            detalle.append(f"un Diámetro de {diam} in")
        if detalle:
            partes.append("La Tubería cuenta con " + self._unir(detalle) + ".")

        rect = str(data.get('rectificadores_tgi') or '').strip()
        if rect and 'ESCRIBIR' not in rect.upper():
            partes.append(f"Tiene como mecanismo contra la corrosión externa un "
                          f"sistema de corriente impresa por las URPC de {rect}.")
        self._safe_write(ws, bloque[0], 1, " ".join(partes))

    def fill_general_info(self, data: dict):
        """Fill Informe sheet rows 6-9 with general information
        
        data keys: fecha, gasoducto, tramo, inspector, serial_equipo, 
                   fecha_calibracion, tipo_recubrimiento, diametro,
                   contrato, ot, contratista, ciclo
        """
        ws = self.ws_informe
        # Filas 6-9: pares etiqueta/valor en celdas combinadas. Cada valor se
        # escribe en la celda que sigue a SU etiqueta (ver CAMPOS_ENCABEZADO):
        # así funciona con cualquiera de las tres plantillas.
        respaldo = {   # por si la fila no trae celdas combinadas
            6: ('fecha', 'serial_equipo', 'contrato'),
            7: ('gasoducto', 'fecha_calibracion', 'ot'),
            8: ('tramo', 'tipo_recubrimiento', 'contratista'),
            9: ('inspector', 'diametro', 'ciclo'),
        }
        for fila in (6, 7, 8, 9):
            anclas = sorted({m.min_col for m in ws.merged_cells.ranges
                             if m.min_row == fila and m.max_row == fila})
            escrito = False
            for i in range(0, len(anclas) - 1, 2):
                campo = self._campo_de_etiqueta(
                    ws.cell(row=fila, column=anclas[i]).value)
                if campo:
                    self._safe_write(ws, fila, anclas[i + 1], data.get(campo, ''))
                    escrito = True
            if not escrito:
                for col, clave in zip(self._cols_valor(ws, fila), respaldo[fila]):
                    self._safe_write(ws, fila, col, data.get(clave, ''))

        # Procedimiento, dentro del bloque de DOCUMENTOS DE REFERENCIA (en la
        # plantilla DCVG ese bloque está 6 filas más arriba que en PAP/CIPS,
        # donde la fila 20 caía encima de la lista de equipos).
        tipo_inspeccion = data.get('tipo_inspeccion', 'CIPS')
        texto_pr = (f"PR-I-06 PROCEDIMIENTO PARA ENCENDIDO, CALIBRACIÓN E "
                    f"INSPECCIÓN {tipo_inspeccion} DE SPC")
        bloque = self._bloque_seccion(ws, 'DOCUMENTOS DE REFERENCIA')
        fila_pr = 20
        if bloque:
            ini, fin = bloque
            fila_pr = next((r for r in range(ini, fin + 1)
                            if str(ws.cell(row=r, column=1).value or '')
                            .strip().upper().startswith('PR-I-06')), None)
            if fila_pr is None:
                fila_pr = next((r for r in range(ini, fin + 1)
                                if ws.cell(row=r, column=1).value in (None, '')),
                               fin)
        self._safe_write(ws, fila_pr, 1, texto_pr)

        self._fill_objetivo_dcvg(ws, data)
        self._fill_descripcion_linea(ws, data)

    def fill_equipos_utilizados(self, equipos_list: list):
        """Fill the EQUIPOS UTILIZADOS section (rows 24-28)
        equipos_list is a list of strings
        """
        ws = self.ws_informe
        if not equipos_list:
            return
        # El bloque de equipos está en filas distintas según la plantilla
        # (PAP/CIPS 24-28, DCVG 19-23): se ubica por el título de la sección.
        bloque = self._bloque_seccion(ws, 'EQUIPOS UTILIZADOS') or (24, 28)
        ini, fin = bloque
        cupo = fin - ini + 1
        for r in range(ini, fin + 1):          # limpiar los del formato
            self._safe_write(ws, r, 1, "")
            self._safe_write(ws, r, 19, "")
        for i, eq in enumerate(equipos_list):
            if i < cupo:
                self._safe_write(ws, ini + i, 1, eq)
            elif i < cupo * 2:
                self._safe_write(ws, ini + (i - cupo), 19, eq)

    def fill_sistema_inspeccionado(self, data: dict, potenciales: list):
        """Fill system inspection section (rows 38-46)
        
        data keys: tipo_inspeccion, detalle, justificacion, uso_tierra,
                   amenaza, tipo_ducto, tipo_spc, topografia,
                   altura_inicio, altura_fin, resumen_justificacion
        """
        ws = self.ws_informe
        
        # Calculate start/end points from potenciales
        if potenciales:
            sorted_pot = sorted(potenciales, key=lambda p: p.get('abscisa', 0))
            first = sorted_pot[0]
            last = sorted_pot[-1]
            
            # Row 39 - Punto Inicial
            self._safe_write(ws, 39, 7, first.get('abscisa', 0))           # G39
            self._safe_write(ws, 39, 15, first.get('lat'))                  # O39
            self._safe_write(ws, 39, 23, first.get('lon'))                  # W39
            self._safe_write(ws, 39, 29, data.get('altura_inicio', ''))     # AC39
            
            # Row 40 - Punto Final
            self._safe_write(ws, 40, 7, last.get('abscisa', 0))            # G40
            self._safe_write(ws, 40, 15, last.get('lat'))                   # O40
            self._safe_write(ws, 40, 23, last.get('lon'))                   # W40
            self._safe_write(ws, 40, 29, data.get('altura_fin', ''))        # AC40

        # Row 41
        self._safe_write(ws, 41, 7, data.get('tipo_inspeccion', 'Inspección PAP'))
        self._safe_write(ws, 41, 15, data.get('detalle', 'Normal'))
        self._safe_write(ws, 41, 23, data.get('justificacion', 'Monitoreo'))
        self._safe_write(ws, 41, 29, data.get('uso_tierra', ''))
        
        # Row 42
        self._safe_write(ws, 42, 7, data.get('amenaza', 'CORROSIÓN EXTERNA'))
        self._safe_write(ws, 42, 15, data.get('tipo_ducto', ''))
        self._safe_write(ws, 42, 23, data.get('tipo_spc', ''))
        self._safe_write(ws, 42, 29, data.get('topografia', ''))

        # Calculate lengths (rows 43-45)
        if potenciales:
            sorted_pot = sorted(potenciales, key=lambda p: p.get('abscisa', 0))
            total_length_km = (sorted_pot[-1].get('abscisa', 0) - sorted_pot[0].get('abscisa', 0)) / 1000.0
            
            # Count protected/unprotected/overprotected
            offs = [p['off_mv'] for p in potenciales if p.get('off_mv') is not None]
            n_total = len(offs) if offs else 1
            n_protected = sum(1 for v in offs if v <= -850)
            n_unprotected = sum(1 for v in offs if v > -850)
            n_overprotected = sum(1 for v in offs if v <= -1200)
            
            pct_protected = n_protected / n_total if n_total > 0 else 0
            pct_unprotected = n_unprotected / n_total if n_total > 0 else 0
            pct_overprotected = n_overprotected / n_total if n_total > 0 else 0
            
            len_protected = total_length_km * pct_protected
            len_unprotected = total_length_km * pct_unprotected
            len_overprotected = total_length_km * pct_overprotected
            
            # Row 43 - Aerial (default 0 for PAP)
            self._safe_write(ws, 43, 7, 0)     # G43 - Long total aérea
            self._safe_write(ws, 43, 15, 0)    # O43 - Long aérea inspeccionada
            self._safe_write(ws, 43, 23, 0)    # W43 - Aérea no insp
            self._safe_write(ws, 43, 29, 0)    # AC43 - Enterrada no insp
            
            # Row 44 - Buried
            self._safe_write(ws, 44, 7, total_length_km)     # G44
            self._safe_write(ws, 44, 15, total_length_km)    # O44
            self._safe_write(ws, 44, 23, len_protected)      # W44
            self._safe_write(ws, 44, 29, len_unprotected)    # AC44
            
            # Row 45 - Percentages
            self._safe_write(ws, 45, 7, len_overprotected)   # G45
            self._safe_write(ws, 45, 15, pct_protected)      # O45
            self._safe_write(ws, 45, 23, pct_unprotected)    # W45
            self._safe_write(ws, 45, 29, pct_overprotected)  # AC45
            
            # Row 31 - Descripción de la Línea
            tipo_tramo = data.get('tipo_tramo', 'Tramo')
            tramo = data.get('tramo', '')
            gasoducto = data.get('gasoducto', '')
            recubrimiento = data.get('tipo_recubrimiento', '')
            diametro = data.get('diametro', '')
            rectificadores_tgi = data.get('rectificadores_tgi', '[ESCRIBIR RECTIFICADORES TGI]')
            
            descripcion = (f"El {tipo_tramo} {tramo} perteneciente al Gasoducto {gasoducto}, "
                           f"cuenta con una longitud de {total_length_km:.1f} Km aproximadamente. "
                           f"La Tubería cuenta con un recubrimiento {recubrimiento} y un Diámetro de {diametro} in, "
                           f"tiene como mecanismo contra la corrosión externa un sistema de corriente impresa "
                           f"por las URPC de {rectificadores_tgi} propiedad de TGI. Adicional, las URPC's "
                           f"[ESCRIBIR RECTIFICADORES CENIT] propiedad de CENIT, tienen influencia sobre el Ramal.")
            self._safe_write(ws, 30, 1, descripcion)
        
        # Row 46 - Resumen justificación
        self._safe_write(ws, 46, 7, data.get('resumen_justificacion', ''))

    def fill_monitoreo(self, data: dict):
        """Fill monitoring section (rows 48-51)
        
        data keys: criterio, descripcion_criterio, ciclo_on, ciclo_off,
                   datos_por_km, pct_rechazados, clima
        """
        ws = self.ws_informe
        # Row 50
        self._safe_write(ws, 50, 7, data.get('criterio', '6.2.1.3 (-850mVCSE)'))
        self._safe_write(ws, 50, 15, data.get('descripcion_criterio', 
            'Potencial polarizado más electronegativo que -850mVCSE'))
        self._safe_write(ws, 50, 23, data.get('ciclo_on', 1.6))
        self._safe_write(ws, 50, 29, data.get('ciclo_off', 0.4))
        # Row 51
        self._safe_write(ws, 51, 7, data.get('datos_por_km', 0))
        self._safe_write(ws, 51, 15, data.get('pct_rechazados', 0))
        self._safe_write(ws, 51, 23, data.get('clima', ''))

    def fill_potenciales_pap(self, potenciales: list, fecha: str = ''):
        """Fill Potenciales PAP data table starting at row 12
        
        Each potencial dict has: abscisa, fecha, ref_geografica, on_mv, off_mv,
        on_mv_corregido, off_mv_corregido, on_mv_neg2, off_mv_neg2,
        on_mv_foraneo1, off_mv_foraneo1, on_mv_foraneo2, off_mv_foraneo2,
        potencial_natural, polarizacion, vac, resistencia, ir_on_off,
        lat, lon, alt, pintura, conexiones, verticalidad, tipo_mant, observaciones
        """
        ws = self.ws_potenciales
        sorted_pot = sorted(potenciales, key=lambda p: p.get('abscisa', 0))
        
        if len(sorted_pot) > 1:
            ws.insert_rows(13, amount=len(sorted_pot)-1)
            # Copiar estilo (solo si no son demasiados, o hacerlo rapido)
            if len(sorted_pot) < 1000:
                for i in range(1, len(sorted_pot)):
                    self._copy_row_style(ws, 12, 12 + i, 1, 27)

        for i, p in enumerate(sorted_pot):
            row = 12 + i
                
            self._safe_write(ws, row, 1, i + 1)                              # A - ITEM
            self._safe_write(ws, row, 2, p.get('abscisa', ''))                # B - ABSCISADO
            self._safe_write(ws, row, 3, p.get('fecha', fecha))               # C - FECHA
            self._safe_write(ws, row, 4, corregir_campo(p.get('ref_geografica', '')))  # D - REF GEOG
            self._safe_write(ws, row, 5, p.get('on_mv'))                      # E - ON NEG1
            self._safe_write(ws, row, 6, p.get('off_mv'))                     # F - OFF NEG1
            self._safe_write(ws, row, 7, p.get('on_mv_corregido'))            # G - ON CORR
            self._safe_write(ws, row, 8, p.get('off_mv_corregido'))           # H - OFF CORR
            self._safe_write(ws, row, 9, p.get('on_mv_neg2'))                 # I - ON NEG2
            self._safe_write(ws, row, 10, p.get('off_mv_neg2'))               # J - OFF NEG2
            self._safe_write(ws, row, 11, p.get('on_mv_foraneo1'))            # K
            self._safe_write(ws, row, 12, p.get('off_mv_foraneo1'))           # L
            self._safe_write(ws, row, 13, p.get('on_mv_foraneo2'))            # M
            self._safe_write(ws, row, 14, p.get('off_mv_foraneo2'))           # N
            self._safe_write(ws, row, 15, p.get('potencial_natural'))         # O
            self._safe_write(ws, row, 16, p.get('polarizacion'))              # P
            self._safe_write(ws, row, 17, p.get('vac'))                       # Q - VAC
            self._safe_write(ws, row, 18, p.get('resistencia'))               # R - Resistencia
            self._safe_write(ws, row, 19, p.get('ir_on_off'))                 # S - IR ON-OFF
            self._safe_write(ws, row, 20, p.get('lat'))                       # T - LAT
            self._safe_write(ws, row, 21, p.get('lon'))                       # U - LON
            self._safe_write(ws, row, 22, p.get('alt'))                       # V - Altura
            self._safe_write(ws, row, 23, p.get('pintura'))                   # W - Pintura
            self._safe_write(ws, row, 24, p.get('conexiones'))                # X - Conexiones
            self._safe_write(ws, row, 25, p.get('verticalidad'))              # Y - Verticalidad
            self._safe_write(ws, row, 26, p.get('tipo_mant'))                 # Z - Tipo mant
            self._safe_write(ws, row, 27, corregir_campo(p.get('observaciones')))  # AA - Obs

    def fill_cips(self, cips_data: list):
        self.cips_truncados = 0
        if not cips_data:
            return

        ws_cips = self.wb['Potenciales CIPS']

        # Capacidad: al final de la hoja hay un bloque de resumen/firmas con
        # celdas combinadas (no se puede escribir sobre ellas). La data va de
        # la fila 12 hasta justo antes de ese bloque. El bloque se detecta por
        # el primer rango combinado en/bajo la fila 12 (rápido).
        merges_datos = [m.min_row for m in ws_cips.merged_cells.ranges
                        if m.min_row >= 12]
        capacidad = (min(merges_datos) - 12) if merges_datos else len(cips_data)
        if len(cips_data) > capacidad:
            self.cips_truncados = len(cips_data) - capacidad
            cips_data = cips_data[:capacidad]

        # Modelo de formato (fila 12) para extender las filas que la plantilla
        # no trae pre-formateadas (surveys largos tipo OCENSA). El formato
        # numérico "K 000+000" de la columna B marca las filas ya listas.
        _kfmt = ws_cips.cell(row=12, column=2).number_format
        _modelo = {c: ws_cips.cell(row=12, column=c) for c in range(1, 22)}

        # Start inserting at row 12
        for i, data in enumerate(cips_data):
            row_idx = i + 12

            # Si la fila no está pre-formateada, copiar el estilo del modelo y
            # las fórmulas por fila (V/W/X, empty-safe) para que la tabla se vea
            # bien y las estadísticas la cuenten.
            if ws_cips.cell(row=row_idx, column=2).number_format != _kfmt:
                for c, src in _modelo.items():
                    if src.has_style:
                        tc = ws_cips.cell(row=row_idx, column=c)
                        tc.font = copy(src.font)
                        tc.border = copy(src.border)
                        tc.fill = copy(src.fill)
                        tc.number_format = src.number_format
                        tc.alignment = copy(src.alignment)
                ws_cips.cell(row=row_idx, column=22).value = \
                    f'=IF(F{row_idx}="","",IF(F{row_idx}<=-850,1,0))'
                ws_cips.cell(row=row_idx, column=23).value = \
                    f'=IF(F{row_idx}="","",IF(F{row_idx}>-1200,1,0))'
                ws_cips.cell(row=row_idx, column=24).value = \
                    f'=CONCATENATE(T{row_idx},",",S{row_idx})'
            
            # Abscisa: escribir el VALOR NUMÉRICO en metros. La celda de la
            # columna B tiene formato personalizado \K\ 000\+000 que lo muestra
            # como "K 000+007", y la serie X (numérica) de las gráficas VDC /
            # Interferencia lee esta columna: si se escribe texto, Excel no
            # encuentra puntos y la gráfica queda vacía. Se usa 'abscisa_val'
            # (metros del motor LRS) y 0 es un valor válido, no "sin dato".
            abscisa = data.get('abscisa_val', data.get('abscisa'))
            if abscisa is not None and pd.notna(abscisa):
                abscisa = int(round(abscisa)) if isinstance(abscisa, (int, float)) else abscisa
            else:
                abscisa = None

            self._safe_write(ws_cips, row_idx, 1, i + 1)
            self._safe_write(ws_cips, row_idx, 2, abscisa)
            # Fecha del día en que se tomó el dato (columna C = FECHA)
            self._safe_write(ws_cips, row_idx, 3, data.get('fecha', ''))
            self._safe_write(ws_cips, row_idx, 4, corregir_campo(data.get('referencia', '')))
            
            # POTENCIAL NEGATIVO 1 TGI (E/F): se escribe el potencial SUAVIZADO
            # (on_limpio/off_limpio del motor LRS: mediana móvil ventana 15 que
            # reemplaza los picos > 250 mV), igual que la app original
            # proceso-cips —que grafica y reporta la data limpia, no la cruda—.
            # La gráfica CIPS lee E/F, así que aquí es donde debe ir el suavizado.
            on_e = data.get('on_limpio')
            off_e = data.get('off_limpio')
            if on_e is None or (isinstance(on_e, float) and pd.isna(on_e)):
                on_e = data.get('on_mv', '')
            if off_e is None or (isinstance(off_e, float) and pd.isna(off_e)):
                off_e = data.get('off_mv', '')
            self._safe_write(ws_cips, row_idx, 5, on_e)
            self._safe_write(ws_cips, row_idx, 6, off_e)

            # Columnas G/H "POTENCIAL NEGATIVO 1 TGI [CORREGIDO]" se dejan
            # VACÍAS por pedido del usuario: el informe lleva únicamente los
            # potenciales de la columna POTENCIAL NEGATIVO 1 TGI (E/F).
            
            # POTENCIAL NATURAL, POLARIZACIÓN (leave empty for now unless calculated)
            
            self._safe_write(ws_cips, row_idx, 11, data.get('vac', ''))
            
            self._safe_write(ws_cips, row_idx, 12, data.get('metal_on', ''))
            self._safe_write(ws_cips, row_idx, 13, data.get('metal_off', ''))
            
            self._safe_write(ws_cips, row_idx, 14, data.get('far_on', ''))
            self._safe_write(ws_cips, row_idx, 15, data.get('far_off', ''))
            
            self._safe_write(ws_cips, row_idx, 16, data.get('near_on', ''))
            self._safe_write(ws_cips, row_idx, 17, data.get('near_off', ''))
            
            # IR ON-OFF con los mismos potenciales suavizados de E/F
            if isinstance(on_e, (int, float)) and isinstance(off_e, (int, float)):
                self._safe_write(ws_cips, row_idx, 18, on_e - off_e)
                
            self._safe_write(ws_cips, row_idx, 19, data.get('lat', ''))
            self._safe_write(ws_cips, row_idx, 20, data.get('lon', ''))
            self._safe_write(ws_cips, row_idx, 21, corregir_campo(data.get('observaciones', '')))

    def fill_graficas(self, potenciales: list, info: dict):
        """Fill chart data for VDC, Interferencia, and VAC graphs"""
        sorted_pot = sorted(potenciales, key=lambda p: p.get('abscisa', 0))
        if not sorted_pot:
            return
            
        max_abscisa = sorted_pot[-1].get('abscisa', 0)
        fecha = info.get('fecha', '')
        gasoducto = info.get('gasoducto', '')
        tramo = info.get('tramo', '')
        tipo_ducto = info.get('tipo_ducto', '')
        longitud = info.get('longitud_km', 0)
        diametro = info.get('diametro', '')
        recubrimiento = info.get('tipo_recubrimiento', '')

        # --- Gráfica VDC ---
        if self.ws_grafica_vdc:
            ws = self.ws_grafica_vdc
            # Info box (rows 31-36)
            self._safe_write(ws, 31, 4, fecha)
            self._safe_write(ws, 32, 4, gasoducto)
            self._safe_write(ws, 33, 4, f'{tipo_ducto} {tramo}')
            self._safe_write(ws, 34, 4, longitud)
            self._safe_write(ws, 35, 4, diametro)
            self._safe_write(ws, 36, 4, recubrimiento)
            
            # Criteria lines (rows 41-43)
            self._safe_write(ws, 42, 5, 0)
            self._safe_write(ws, 43, 5, max_abscisa)
            self._safe_write(ws, 42, 6, -850)
            self._safe_write(ws, 43, 6, -850)
            self._safe_write(ws, 42, 7, -1200)
            self._safe_write(ws, 43, 7, -1200)
            self._safe_write(ws, 42, 8, 100)
            self._safe_write(ws, 43, 8, 100)
            
            # Comments for chart annotations (rows 47+)
            self._safe_write(ws, 47, 5, 'Comentarios gráfica')
            self._safe_write(ws, 48, 5, 'Abscisa')
            self._safe_write(ws, 48, 6, 'Posición Y coment.')
            self._safe_write(ws, 48, 7, 'Comentario')
            for i, p in enumerate(sorted_pot):
                row = 49 + i
                self._safe_write(ws, row, 5, p.get('abscisa', 0))
                self._safe_write(ws, row, 6, -2100)  # Y position for annotation
                obs = p.get('observaciones', '')
                self._safe_write(ws, row, 7, corregir_campo(obs if obs else p.get('ref_geografica', '')))
            
            # Observations text
            offs = [p['off_mv'] for p in potenciales if p.get('off_mv') is not None]
            n_total = len(offs) if offs else 1
            n_protected = sum(1 for v in offs if v <= -850)
            pct = round(n_protected / n_total * 100) if n_total > 0 else 0
            n_over = sum(1 for v in offs if v <= -1200)
            pct_over = round(n_over / n_total * 100) if n_total > 0 else 0
            
            obs_vdc = (f'Los potenciales de protección catódica (Instant Off), registrados mediante '
                      f'la técnica de Inspección PAP realizada a la línea {tipo_ducto} {tramo}, '
                      f'cumple en un {pct}% de la longitud inspeccionada el criterio establecido '
                      f'en el numeral 6.2.1.3 de la norma NACE SP0169 "un potencial estructura '
                      f'electrolito de -850 mV o mas negativo, medido respecto a un electrodo '
                      f'de referencia de cobre sulfato de cobre [CSE]"')
            self._safe_write(ws, 32, 6, obs_vdc)
            
            obs_1200 = (f'Los potenciales de protección catódica (Instant Off), registrados mediante '
                       f'la técnica de Inspección PAP realizada a la línea {tipo_ducto} {tramo}, '
                       f'registró que el {pct_over:03d}% de la longitud inspeccionada presenta un '
                       f'potencial estructura electrolito mas electronegativo de -1200 mV[CSE].')
            self._safe_write(ws, 35, 6, obs_1200)

        # --- Gráfica Interferencia ---
        if self.ws_grafica_interf:
            ws = self.ws_grafica_interf
            self._safe_write(ws, 31, 4, fecha)
            self._safe_write(ws, 32, 4, gasoducto)
            self._safe_write(ws, 33, 4, f'{tipo_ducto} {tramo}')
            self._safe_write(ws, 34, 4, longitud)
            self._safe_write(ws, 35, 4, diametro)
            self._safe_write(ws, 36, 4, recubrimiento)
            # Criteria 50mV line
            self._safe_write(ws, 41, 5, 0)
            self._safe_write(ws, 42, 5, max_abscisa)
            self._safe_write(ws, 41, 6, 50)
            self._safe_write(ws, 42, 6, 50)
            # Observation
            obs_interf = (f'A lo largo de la totalidad del tramo inspeccionado del '
                         f'{tipo_ducto} {tramo} no se evidencian inversiones de potencial.')
            self._safe_write(ws, 32, 6, obs_interf)

        # --- Gráfica VAC ---
        if self.ws_grafica_vac:
            ws = self.ws_grafica_vac
            self._safe_write(ws, 31, 4, fecha)
            self._safe_write(ws, 32, 4, gasoducto)
            self._safe_write(ws, 33, 4, f'{tipo_ducto} {tramo}')
            self._safe_write(ws, 34, 4, longitud)
            self._safe_write(ws, 35, 4, diametro)
            self._safe_write(ws, 36, 4, recubrimiento)
            # Criteria 15 VAC
            self._safe_write(ws, 42, 5, 0)
            self._safe_write(ws, 43, 5, max_abscisa)
            self._safe_write(ws, 42, 6, 15)
            self._safe_write(ws, 43, 6, 15)
            # Observation
            vacs = [p.get('vac', 0) for p in potenciales if p.get('vac') is not None]
            cumple = all(v <= 15 for v in vacs) if vacs else True
            supera_text = 'no superan' if cumple else 'superan'
            obs_vac = (f'Los potenciales AC registrados en {tipo_ducto} {tramo} {supera_text} '
                      f'el limite establecido en la norma NACE SP0177-19 numeral 5.2.1.1 '
                      f'"Los límites de seguridad los determinará un personal calificado y '
                      f'estos no deben superar los 15VAC con respecto a una tierra local, '
                      f'en este caso al electrodo de Cu/CUSO4" en los '
                      f'{longitud:03.0f} Km inspeccionados.')
            self._safe_write(ws, 32, 6, obs_vac)

        self.ajustar_graficas(potenciales)

    def fill_hallazgos(self, hallazgos: list, info: dict):
        """Fill Hallazgos sheet
        
        info keys: fecha, gasoducto, tramo, tipo_inspeccion, contrato, contratista, ot, inspector
        Each hallazgo: abscisa_inicio, abscisa_fin, longitud, gasoducto, tramo,
                       lat_inicio, lon_inicio, lat_fin, lon_fin, fecha, tipo, descripcion
        """
        ws = self.ws_hallazgos
        # Header info (rows 6-9)
        self._safe_write(ws, 6, 3, info.get('fecha', ''))
        self._safe_write(ws, 6, 12, info.get('tipo_inspeccion', 'Inspección PAP'))
        self._safe_write(ws, 7, 3, info.get('gasoducto', ''))
        self._safe_write(ws, 7, 12, info.get('contrato', ''))
        self._safe_write(ws, 8, 3, info.get('tramo', ''))
        self._safe_write(ws, 8, 12, info.get('contratista', ''))
        self._safe_write(ws, 9, 3, info.get('inspector', ''))
        self._safe_write(ws, 9, 12, info.get('ot', ''))

        if not hallazgos:
            return

        # Ordenar de menor a mayor abscisa (los sin abscisa quedan al final).
        def _absc(h):
            v = h.get('abscisa_val', h.get('abscisa_inicio', h.get('abscisa')))
            try:
                return (v is None, float(v))
            except (TypeError, ValueError):
                return (True, 0.0)
        hallazgos = sorted(hallazgos, key=_absc)

        start_row = 18

        # La plantilla ya trae 500 filas de datos formateadas antes del bloque
        # de firmas, así que aquí solo se escribe (no se insertan filas). La
        # capacidad = filas de datos disponibles antes de las firmas.
        fila_firmas = next((r for r in range(start_row, ws.max_row + 1)
                            if ws.cell(row=r, column=3).value == 'ELABORÓ'), None)
        if fila_firmas is not None:
            tope_bloque = min([m.min_row for m in ws.merged_cells.ranges
                               if m.min_row >= start_row] + [fila_firmas])
        else:
            tope_bloque = start_row + len(hallazgos) + 1
        capacidad = tope_bloque - start_row

        hallazgos = hallazgos[:capacidad]   # no invadir el bloque de firmas

        for i, h in enumerate(hallazgos):
            row = start_row + i

            # Prefer abscisa_val for numeric formatting if available
            abs_ini = h.get('abscisa_val', h.get('abscisa_inicio', h.get('abscisa', '')))
            abs_fin = h.get('abscisa_fin', '')
            
            self._safe_write(ws, row, 1, i + 1)                               # A - ITEM
            self._safe_write(ws, row, 2, abs_ini)                              # B
            self._safe_write(ws, row, 3, abs_fin)                              # C
            self._safe_write(ws, row, 4, h.get('longitud', ''))                # D
            self._safe_write(ws, row, 5, h.get('gasoducto') or info.get('gasoducto', ''))  # E
            self._safe_write(ws, row, 6, h.get('tramo') or info.get('tramo', ''))           # F
            self._safe_write(ws, row, 7, h.get('lat_inicio', h.get('lat', '')))  # G
            self._safe_write(ws, row, 8, h.get('lon_inicio', h.get('lon', '')))  # H
            self._safe_write(ws, row, 9, h.get('lat_fin', ''))                   # I
            self._safe_write(ws, row, 10, h.get('lon_fin', ''))                  # J
            # fecha del punto; si el hallazgo no la trae, la general del informe
            self._safe_write(ws, row, 11, h.get('fecha') or info.get('fecha', ''))  # K
            self._safe_write(ws, row, 12, corregir_campo(h.get('tipo', '')))     # L
            self._safe_write(ws, row, 13, corregir_campo(h.get('descripcion', '')))  # M

        # Limpiar las filas de datos no usadas (hasta el bloque de firmas), por
        # si el template traía texto de ejemplo o de una corrida anterior.
        for r in range(start_row + len(hallazgos), tope_bloque):
            for c in range(1, 14):
                self._safe_write(ws, r, c, '')

    # Bloque de rectificadores: mismas columnas lógicas en las tres plantillas,
    # pero en posiciones distintas (PAP disponibilidad en Q/T, DCVG en O/Q…).
    TITULO_RECTIFICADORES = 'PARÁMETROS OPERATIVOS EN RECTIFICADORES'

    def _buscar_texto(self, ws, texto, hasta=140):
        """(fila, columna) de la primera celda cuyo texto empieza por `texto`."""
        objetivo = self._etiqueta(texto)
        for r in range(1, min(ws.max_row, hasta) + 1):
            for c in range(1, 6):
                if self._etiqueta(ws.cell(row=r, column=c).value).startswith(objetivo):
                    return (r, c)
        return (None, None)

    def _mapa_rectificadores(self, ws):
        """(fila_datos, tope, {campo: columna}) del bloque de URPC, leído de sus
        dos filas de encabezado (grupos y subcolumnas)."""
        fila_tit, _c = self._buscar_texto(ws, self.TITULO_RECTIFICADORES)
        if fila_tit is None:
            return (80, None, {})       # respaldo: distribución de PAP
        f_grupo, f_sub = fila_tit + 1, fila_tit + 2
        inicio = f_sub + 1

        def celdas(fila):
            return [(c, self._etiqueta(ws.cell(row=fila, column=c).value))
                    for c in range(1, 40)
                    if ws.cell(row=fila, column=c).value not in (None, '')]

        grupos = celdas(f_grupo)
        subs = celdas(f_sub)

        def col_grupo(nombre):
            return next((c for c, t in grupos if t.startswith(nombre)), None)

        def subs_de(nombre):
            ini = col_grupo(nombre)
            if ini is None:
                return []
            sig = [c for c, _t in grupos if c > ini]
            fin = min(sig) if sig else 99
            return [(c, t) for c, t in subs if ini <= c < fin]

        def sub_con(nombre, palabra):
            return next((c for c, t in subs_de(nombre) if palabra in t), None)

        mapa = {
            'nombre': col_grupo('urpc'),
            'voltaje_nominal': sub_con('datos nominales', 'voltaje'),
            'corriente_nominal': sub_con('datos nominales', 'corriente'),
            'vdc_salida': sub_con('datos operacionales', 'voltaje'),
            'idc_salida': sub_con('datos operacionales', 'corriente'),
            'disponibilidad_v': sub_con('disponibilidad', 'voltaje'),
            'disponibilidad_i': sub_con('disponibilidad', 'corriente'),
            'taps': col_grupo('taps'),
        }
        for clave, grupo in (('neg', 'potencial on-instant off'),
                             ('neg_a', 'corriente negativos')):
            cols = [c for c, _t in subs_de(grupo)] or [col_grupo(grupo)]
            for i, c in enumerate(cols[:3]):
                mapa[f'{clave}{i + 1}'] = c

        # el bloque termina donde empieza la siguiente sección
        tope = None
        for r in range(inicio, min(ws.max_row, inicio + 40) + 1):
            if any(ws.cell(row=r, column=c).value not in (None, '')
                   for c in (1, 2)) and r > inicio:
                tope = r - 1
                break
        return (inicio, tope, mapa)

    def fill_rectificadores(self, rectificadores: list):
        """Llena 'PARÁMETROS OPERATIVOS EN RECTIFICADORES' (URPC) de la hoja
        Informe. Las filas y columnas se leen del propio encabezado del bloque,
        porque cada plantilla lo tiene en un sitio distinto.

        Cada rect: nombre, voltaje_nominal, corriente_nominal, ultima_inspeccion
        {vdc_salida, idc_salida, disponibilidad_v, disponibilidad_i, taps},
        conexion_estructura {pot_on, pot_off, corriente}.
        `self.rect_omitidos` = los que no cupieron en el bloque."""
        self.rect_omitidos = 0
        if not rectificadores:
            return
        ws = self.ws_informe
        inicio, tope, mapa = self._mapa_rectificadores(ws)
        if not mapa.get('nombre'):
            return
        cupo = (tope - inicio + 1) if tope else len(rectificadores)
        if len(rectificadores) > cupo:
            self.rect_omitidos = len(rectificadores) - cupo
            rectificadores = rectificadores[:cupo]

        def poner(row, campo, valor):
            col = mapa.get(campo)
            if col:
                self._safe_write(ws, row, col, valor)

        for i, r in enumerate(rectificadores):
            row = inicio + i
            if i > 0:
                self._copy_row_style(ws, inicio, row, 2, 34)
            poner(row, 'nombre', r.get('nombre', ''))
            poner(row, 'voltaje_nominal', r.get('voltaje_nominal'))
            poner(row, 'corriente_nominal', r.get('corriente_nominal'))
            ultima = r.get('ultima_inspeccion', {})
            for campo in ('vdc_salida', 'idc_salida', 'disponibilidad_v',
                          'disponibilidad_i'):
                poner(row, campo, ultima.get(campo))
            poner(row, 'taps', ultima.get('taps', ''))

            conexion = r.get('conexion_estructura', {})
            pot_on, pot_off = conexion.get('pot_on', ''), conexion.get('pot_off', '')
            poner(row, 'neg1', f'ON: {pot_on}\nOFF: {pot_off}'
                  if (pot_on and pot_off) else '-')
            poner(row, 'neg2', '-')
            poner(row, 'neg3', '-')
            poner(row, 'neg_a1', conexion.get('corriente', ''))
            poner(row, 'neg_a2', '-')
            poner(row, 'neg_a3', '-')

    def fill_aislamientos(self, aislamientos: list):
        """Fill Aislamientos sheet data starting at row 13
        
        Each aislamiento: abscisado, tag, clase, diametro, presion, temperatura,
                          tipo_brida, num_pernos, diam_pernos, tipo_aislamiento,
                          pct_aislamiento, pot_on_arriba, pot_off_arriba,
                          pot_on_abajo, pot_off_abajo, dif_on, dif_off,
                          diagnostico, lat, lon, observaciones
        """
        if not self.ws_aislamientos or not aislamientos:
            return
        ws = self.ws_aislamientos
        
        n_prefilled = 6
        start_row = 13
        
        if len(aislamientos) > n_prefilled:
            ws.insert_rows(start_row + n_prefilled, len(aislamientos) - n_prefilled)
            for r in range(start_row + n_prefilled, start_row + len(aislamientos)):
                self._copy_row_style(ws, start_row, r, 1, 22)
        
        for i, a in enumerate(aislamientos):
            row = start_row + i
                
            self._safe_write(ws, row, 1, i + 1)
            self._safe_write(ws, row, 2, a.get('abscisa_val', a.get('abscisado', '')))
            self._safe_write(ws, row, 3, a.get('tag', '-'))
            self._safe_write(ws, row, 4, a.get('clase', ''))
            self._safe_write(ws, row, 5, a.get('diametro', ''))
            self._safe_write(ws, row, 6, a.get('presion', '-'))
            self._safe_write(ws, row, 7, a.get('temperatura', '-'))
            self._safe_write(ws, row, 8, a.get('tipo_brida', ''))
            self._safe_write(ws, row, 9, a.get('num_pernos', ''))
            self._safe_write(ws, row, 10, a.get('diam_pernos', ''))
            self._safe_write(ws, row, 11, a.get('tipo_aislamiento', ''))
            self._safe_write(ws, row, 12, a.get('pct_aislamiento', ''))
            self._safe_write(ws, row, 13, a.get('pot_on_arriba'))
            self._safe_write(ws, row, 14, a.get('pot_off_arriba'))
            self._safe_write(ws, row, 15, a.get('pot_on_abajo'))
            self._safe_write(ws, row, 16, a.get('pot_off_abajo'))
            self._safe_write(ws, row, 17, a.get('dif_on'))
            self._safe_write(ws, row, 18, a.get('dif_off'))
            self._safe_write(ws, row, 19, a.get('diagnostico', ''))
            self._safe_write(ws, row, 20, a.get('lat'))
            self._safe_write(ws, row, 21, a.get('lon'))
            self._safe_write(ws, row, 22, corregir_campo(a.get('observaciones', '')))
            
        # Clear unused prefilled rows
        if len(aislamientos) < n_prefilled:
            for r in range(start_row + len(aislamientos), start_row + n_prefilled):
                for c in range(1, 23):
                    self._safe_write(ws, r, c, '')

    def fill_inspecciones(self, marco_h: list = None, ce: list = None,
                          anodos: list = None, cupones_ir: list = None,
                          cupones_grav: list = None, pe: list = None,
                          tramos_aereos: list = None, tramos_no_insp: list = None):
        """Fill special inspection sheets with data or leave defaults"""
        
        # Marco H
        if marco_h and self.ws_marco_h:
            ws = self.ws_marco_h
            # Clear default text
            self._safe_write(ws, 12, 16, '')
            self._safe_write(ws, 13, 16, '')
            
            for i, m in enumerate(marco_h):
                row = 12 + i
                self._safe_write(ws, row, 1, i + 1)
                self._safe_write(ws, row, 2, m.get('abscisado', ''))
                self._safe_write(ws, row, 3, m.get('fecha', ''))
                self._safe_write(ws, row, 4, m.get('pot_on_gasoducto', ''))
                self._safe_write(ws, row, 5, m.get('pot_off_gasoducto', ''))
                self._safe_write(ws, row, 6, m.get('pot_on_marco', ''))
                self._safe_write(ws, row, 7, m.get('pot_off_marco', ''))
                self._safe_write(ws, row, 8, m.get('aislado', ''))
                self._safe_write(ws, row, 9, m.get('dif_on', ''))
                self._safe_write(ws, row, 10, m.get('dif_off', ''))
                self._safe_write(ws, row, 11, 'Aislado' if m.get('aislado') else 'Corto')
                self._safe_write(ws, row, 12, m.get('estado_aislante', 'Buen Estado'))
                self._safe_write(ws, row, 13, m.get('lat', ''))
                self._safe_write(ws, row, 14, m.get('lon', ''))
                self._safe_write(ws, row, 15, m.get('estado_pintura', 'Bueno'))
                self._safe_write(ws, row, 16, corregir_campo(m.get('observaciones', '')))

        # Inventario Tramos Aéreos
        if tramos_aereos and self.ws_tramos_aereos:
            ws = self.ws_tramos_aereos
            self._safe_write(ws, 12, 1, '')  # Clear default
            for i, t in enumerate(tramos_aereos):
                row = 12 + i
                self._safe_write(ws, row, 1, i + 1)
                self._safe_write(ws, row, 2, t.get('inicio_abscisa', ''))
                self._safe_write(ws, row, 3, t.get('fin_abscisa', ''))
                self._safe_write(ws, row, 4, t.get('longitud', ''))
                self._safe_write(ws, row, 5, t.get('gasoducto', ''))
                self._safe_write(ws, row, 6, t.get('tramo', ''))
                self._safe_write(ws, row, 7, t.get('lat_inicio'))
                self._safe_write(ws, row, 8, t.get('lon_inicio'))
                self._safe_write(ws, row, 9, t.get('lat_fin'))
                self._safe_write(ws, row, 10, t.get('lon_fin'))
                self._safe_write(ws, row, 11, t.get('fecha', ''))
                self._safe_write(ws, row, 12, corregir_campo(t.get('observaciones', '')))

        # Tramos no inspeccionados
        if tramos_no_insp and self.ws_tramos_no_insp:
            ws = self.ws_tramos_no_insp
            for i, t in enumerate(tramos_no_insp):
                row = 12 + i
                self._safe_write(ws, row, 1, i + 1)
                self._safe_write(ws, row, 2, t.get('abscisa_inicio', ''))
                self._safe_write(ws, row, 3, t.get('abscisa_fin', ''))
                self._safe_write(ws, row, 4, t.get('longitud', ''))
                self._safe_write(ws, row, 5, t.get('gasoducto', ''))
                self._safe_write(ws, row, 6, t.get('tramo', ''))
                self._safe_write(ws, row, 7, t.get('lat_inicio'))
                self._safe_write(ws, row, 8, t.get('lon_inicio'))
                self._safe_write(ws, row, 9, t.get('lat_fin'))
                self._safe_write(ws, row, 10, t.get('lon_fin'))
                self._safe_write(ws, row, 11, t.get('fecha', ''))
                self._safe_write(ws, row, 12, t.get('justificacion', ''))

    def _escribir_bloque_texto(self, etiqueta, textos):
        """Escribe una lista de párrafos bajo el título de su sección, sin
        pasarse del espacio disponible. Devuelve cuántos no cupieron."""
        if not textos:
            return 0
        ws = self.ws_informe
        bloque = self._bloque_texto(ws, etiqueta)
        if not bloque:
            return len(textos)
        ini, fin = bloque
        cupo = fin - ini + 1
        for i, txt in enumerate(textos[:cupo]):
            self._safe_write(ws, ini + i, 1, f"• {txt}")
        return max(0, len(textos) - cupo)

    def fill_conclusiones(self, conclusiones: list):
        """Escribe las conclusiones bajo el título CONCLUSIONES de la hoja
        Informe (la fila cambia según la plantilla).
        `self.conclusiones_omitidas` = las que no cupieron."""
        self.conclusiones_omitidas = self._escribir_bloque_texto(
            'CONCLUSIONES', conclusiones)

    def fill_recomendaciones(self, recomendaciones: list):
        """Ídem para RECOMENDACIONES. `self.recomendaciones_omitidas`."""
        self.recomendaciones_omitidas = self._escribir_bloque_texto(
            'RECOMENDACIONES', recomendaciones)

    def fill_firmas(self, elaboro: dict, reviso: dict, aprobo: dict):
        """Fill signatures in ALL sheets
        
        Each dict has: nombre, cargo, empresa
        """
        # Informe sheet
        ws = self.ws_informe
        self._safe_write(ws, 104, 4, elaboro.get('nombre', ''))
        self._safe_write(ws, 105, 4, elaboro.get('cargo', ''))
        self._safe_write(ws, 106, 4, elaboro.get('empresa', ''))
        self._safe_write(ws, 104, 15, reviso.get('nombre', ''))
        self._safe_write(ws, 105, 15, reviso.get('cargo', ''))
        self._safe_write(ws, 106, 15, reviso.get('empresa', ''))
        self._safe_write(ws, 104, 24, aprobo.get('nombre', ''))
        self._safe_write(ws, 105, 24, aprobo.get('cargo', ''))
        self._safe_write(ws, 106, 24, aprobo.get('empresa', ''))

        # Potenciales PAP
        ws = self.ws_potenciales
        start_row = 77
        for r in range(12, 500):
            val = ws.cell(row=r, column=4).value
            if val and isinstance(val, str) and 'ELABORÓ' in val.upper():
                start_row = r + 1
                break
        self._safe_write(ws, start_row, 4, elaboro.get('nombre', ''))
        self._safe_write(ws, start_row + 1, 4, elaboro.get('cargo', ''))
        self._safe_write(ws, start_row + 2, 4, elaboro.get('empresa', ''))
        self._safe_write(ws, start_row, 15, reviso.get('nombre', ''))
        self._safe_write(ws, start_row + 1, 15, reviso.get('cargo', ''))
        self._safe_write(ws, start_row + 2, 15, reviso.get('empresa', ''))
        self._safe_write(ws, start_row, 24, aprobo.get('nombre', ''))
        self._safe_write(ws, start_row + 1, 24, aprobo.get('cargo', ''))
        self._safe_write(ws, start_row + 2, 24, aprobo.get('empresa', ''))

        # Hallazgos
        if self.ws_hallazgos:
            ws = self.ws_hallazgos
            start_row = 26
            for r in range(18, 500):
                val = ws.cell(row=r, column=3).value
                if val and isinstance(val, str) and 'ELABORÓ' in val.upper():
                    start_row = r + 1
                    break
            self._safe_write(ws, start_row, 3, elaboro.get('nombre', ''))
            self._safe_write(ws, start_row + 1, 3, elaboro.get('cargo', ''))
            self._safe_write(ws, start_row + 2, 3, elaboro.get('empresa', ''))
            self._safe_write(ws, start_row, 7, reviso.get('nombre', ''))
            self._safe_write(ws, start_row + 1, 7, reviso.get('cargo', ''))
            self._safe_write(ws, start_row + 2, 7, reviso.get('empresa', ''))
            self._safe_write(ws, start_row, 12, aprobo.get('nombre', ''))
            self._safe_write(ws, start_row + 1, 12, aprobo.get('cargo', ''))
            self._safe_write(ws, start_row + 2, 12, aprobo.get('empresa', ''))

        # Aislamientos
        if self.ws_aislamientos:
            ws = self.ws_aislamientos
            start_row = 19
            for r in range(15, 200):
                val = ws.cell(row=r, column=1).value
                if val and isinstance(val, str) and 'NOMBRE' in val.upper():
                    start_row = r
                    break
                    
            self._safe_write(ws, start_row, 3, elaboro.get('nombre', ''))
            self._safe_write(ws, start_row + 1, 3, elaboro.get('cargo', ''))
            self._safe_write(ws, start_row + 2, 3, elaboro.get('empresa', ''))
            self._safe_write(ws, start_row, 8, reviso.get('nombre', ''))
            self._safe_write(ws, start_row + 1, 8, reviso.get('cargo', ''))
            self._safe_write(ws, start_row + 2, 8, reviso.get('empresa', ''))
            self._safe_write(ws, start_row, 18, aprobo.get('nombre', ''))
            self._safe_write(ws, start_row + 1, 18, aprobo.get('cargo', ''))
            self._safe_write(ws, start_row + 2, 18, aprobo.get('empresa', ''))


    def fill_aislamientos(self, aislamientos: list):
        """Fill Aislamientos data table starting at row 13"""
        ws = self.ws_aislamientos
        if not ws or not aislamientos:
            return
            
        for i, a in enumerate(aislamientos):
            row = 13 + i
            if i > 0:
                ws.insert_rows(row)
                self._copy_row_style(ws, 13, row, 1, 22)
                
            self._safe_write(ws, row, 1, i + 1)                              # A - ÍTEM
            self._safe_write(ws, row, 2, a.get('abscisado', ''))              # B - ABSCISADO
            self._safe_write(ws, row, 3, a.get('tag', ''))                    # C - TAG
            self._safe_write(ws, row, 4, a.get('clase', ''))                  # D - CLASS
            self._safe_write(ws, row, 5, a.get('diametro', ''))               # E - DIÁMETRO NOMINAL
            self._safe_write(ws, row, 6, a.get('presion', ''))                # F - PRESIÓN
            self._safe_write(ws, row, 7, a.get('temperatura', ''))            # G - TEMPERATURA
            self._safe_write(ws, row, 8, a.get('tipo_brida', ''))             # H - TIPO DE BRIDA
            self._safe_write(ws, row, 9, a.get('numero_pernos', ''))          # I - NÚMERO DE PERNOS
            self._safe_write(ws, row, 10, a.get('diametro_pernos', ''))       # J - DIÁMETRO DE PERNOS
            self._safe_write(ws, row, 11, a.get('tipo_aislamiento', ''))      # K - TIPO DE AISLAMIENTO
            self._safe_write(ws, row, 12, a.get('porcentaje_aislamiento', '')) # L - % AISLAMIENTO
            self._safe_write(ws, row, 13, a.get('pot_on_arriba', ''))         # M - AGUAS ARRIBA POT ON
            self._safe_write(ws, row, 14, a.get('pot_off_arriba', ''))        # N - AGUAS ARRIBA POT OFF
            self._safe_write(ws, row, 15, a.get('pot_on_abajo', ''))          # O - AGUAS ABAJO POT ON
            self._safe_write(ws, row, 16, a.get('pot_off_abajo', ''))         # P - AGUAS ABAJO POT OFF
            self._safe_write(ws, row, 17, a.get('diferencia', ''))            # Q - DIFERENCIA
            self._safe_write(ws, row, 18, "")                                 # R - DIFERENCIA INSTANT OFF
            self._safe_write(ws, row, 19, corregir_campo(a.get('diagnostico', '')))  # S - DIAGNÓSTICO
            self._safe_write(ws, row, 20, a.get('latitud', ''))               # T - LATITUD
            self._safe_write(ws, row, 21, a.get('longitud', ''))              # U - LONGITUD
            self._safe_write(ws, row, 22, corregir_campo(a.get('observaciones', '')))  # V - OBSERVACIONES


    def fill_comentario_huella(self, comentario: str):
        """Fill oscilloscopic footprint comment in Informe row 74"""
        self._safe_write(self.ws_informe, 74, 6, comentario)

    @staticmethod
    def _nice_floor(v, step):
        return math.floor(v / step) * step

    @staticmethod
    def _nice_ceil(v, step):
        return math.ceil(v / step) * step

    def ajustar_graficas(self, potenciales):
        """Ajusta rango de series y ejes de las 3 graficas a los datos reales.

        - Series 'Potenciales PAP' de fila 12 (primer poste) a 11+N.
        - Eje X: 0 hasta la abscisa maxima (redondeada a 100 m).
        - Eje Y: min/max de los datos + 10% margen, manteniendo visibles las
          lineas de criterio. Elimina las series #REF! rotas del template.
        """
        import re
        if not potenciales:
            return
        n = len(potenciales)
        last = 11 + n
        max_absc = max((p.get('abscisa') or 0) for p in potenciales)
        x_max = self._nice_ceil(max_absc, 100) if max_absc > 0 else 100

        cfgs = [
            (self.ws_grafica_vdc,    ['on_mv', 'off_mv'], [-850, -1200], 100),
            (self.ws_grafica_interf, ['ir_on_off'],       [50],          10),
            (self.ws_grafica_vac,    ['vac'],             [15],          2),
        ]

        def _ref_ok(s):
            for part in (s.xVal, s.yVal):
                f = part.numRef.f if (part and part.numRef) else None
                if f is None or '#REF!' in f:
                    return False
            return True

        for ws, keys, criterios, step in cfgs:
            if ws is None or not getattr(ws, '_charts', None):
                continue
            chart = ws._charts[0]

            series_validas = [s for s in chart.series if _ref_ok(s)]
            for s in series_validas:
                fx = s.xVal.numRef.f
                if 'Potenciales PAP' in fx:
                    s.xVal.numRef.f = re.sub(r'\$B\$\d+:\$B\$\d+',
                                             f'$B$12:$B${last}', fx)
                    col = re.search(r'\$([A-Z]+)\$\d+:', s.yVal.numRef.f).group(1)
                    s.yVal.numRef.f = f"'Potenciales PAP'!${col}$12:${col}${last}"
                    if s.xVal.numRef.numCache:
                        s.xVal.numRef.numCache = None
                    if s.yVal.numRef.numCache:
                        s.yVal.numRef.numCache = None
            chart.series = series_validas

            chart.x_axis.scaling.min = 0
            chart.x_axis.scaling.max = x_max

            vals = [p.get(k) for k in keys for p in potenciales if p.get(k) is not None]
            if vals:
                lo = min(vals + criterios)
                hi = max(vals + criterios)
                span = (hi - lo) or 1
                margin = span * 0.1
                chart.y_axis.scaling.min = self._nice_floor(lo - margin, step)
                chart.y_axis.scaling.max = self._nice_ceil(hi + margin, step)

    def fill_graficas_cips(self, cips_data, info=None):
        """Ajusta las gráficas del template CIPS (VDC e Interferencia).

        Estas gráficas leen la hoja 'Potenciales CIPS' (eje X = abscisa en
        columna B; Y = On/Off en E/F e IR en R). Las líneas de criterio
        (-850, -1200, 100, 50 mV) ya vienen precargadas en el template, así que
        aquí solo se: (1) recorta el rango de las series de datos de las ~29 000
        filas del template a las N filas realmente escritas, (2) fija la abscisa
        máxima real en la línea de criterio y en el eje X. Sin esto, la serie X
        abarca miles de celdas vacías y el eje queda dominado por el 38 000 m
        que trae el template por defecto.
        """
        import re
        if not cips_data:
            return
        n = len(cips_data)
        last = 11 + n
        abscisas = [d.get('abscisa_val', d.get('abscisa')) for d in cips_data]
        abscisas = [a for a in abscisas if a is not None and pd.notna(a)]
        max_absc = max(abscisas) if abscisas else 0
        x_max = self._nice_ceil(max_absc, 100) if max_absc > 0 else 100

        # Endpoint de la abscisa en cada línea de criterio (celda ya existente).
        criterio_absc = [
            ('Gráfica VDC ', 44, 4),          # D44
            ('Gráfica Interferencia', 42, 5),  # E42
        ]
        for hoja, fila, col in criterio_absc:
            if hoja in self.wb.sheetnames:
                self._safe_write(self.wb[hoja], fila, col, x_max)

        # Comentarios de la gráfica (filas 49+ de 'Gráfica VDC '): limpiar
        # cualquier remanente del template y escribir los comentarios reales
        # del survey (etiquetas DCP del técnico). Si quedaran filas viejas, el
        # informe mostraría anotaciones de OTRA inspección.
        if 'Gráfica VDC ' in self.wb.sheetnames:
            ws_vdc = self.wb['Gráfica VDC ']
            for r in range(49, max(ws_vdc.max_row, 49) + 1):
                for c in (4, 5, 6):
                    if ws_vdc.cell(row=r, column=c).value is not None:
                        ws_vdc.cell(row=r, column=c).value = None
            fila = 49
            for d in cips_data:
                obs = str(d.get('observaciones') or d.get('referencia') or '').strip()
                absc = d.get('abscisa_val')
                if not obs or obs.lower() == 'nan' or absc is None:
                    continue
                self._safe_write(ws_vdc, fila, 4, absc)
                self._safe_write(ws_vdc, fila, 5, -2000)
                self._safe_write(ws_vdc, fila, 6, corregir_campo(obs))
                fila += 1

        for hoja in ('Gráfica VDC ', 'Gráfica Interferencia'):
            if hoja not in self.wb.sheetnames:
                continue
            ws = self.wb[hoja]
            if not getattr(ws, '_charts', None):
                continue
            chart = ws._charts[0]
            for s in chart.series:
                fx = s.xVal.numRef.f if (s.xVal and s.xVal.numRef) else None
                if not fx or 'Potenciales CIPS' not in fx:
                    continue
                s.xVal.numRef.f = re.sub(r'\$B\$\d+:\$B\$\d+',
                                         f'$B$12:$B${last}', fx)
                m = re.search(r'\$([A-Z]+)\$\d+:', s.yVal.numRef.f)
                if m:
                    col = m.group(1)
                    s.yVal.numRef.f = f"'Potenciales CIPS'!${col}$12:${col}${last}"
                if s.xVal.numRef.numCache:
                    s.xVal.numRef.numCache = None
                if s.yVal.numRef.numCache:
                    s.yVal.numRef.numCache = None
            chart.x_axis.scaling.min = 0
            chart.x_axis.scaling.max = x_max

    # ── DCVG ──────────────────────────────────────────────────────────────────

    def fill_dcvg(self, postes: list, defectos: list, resistividades: list = None,
                  hallazgos: list = None):
        """Llena la hoja 'Inspección DCVG' con postes + defectos + hallazgos
        intercalados por abscisa (en secuencia del recorrido). Postes traen
        ON/OFF (potencial estructura) y su pulso P=ABS(N-O). Los defectos traen
        forma (N→12/E→3/S→6/O→9), carácter, OL/RE, profundidad; la severidad
        %IR = OL/RE (col S/T/U según carácter AA/CA/CC como fracción para el
        formato '0%'), y la clasificación (V) por umbrales. Los hallazgos
        (cruces, tramos enmontados…) van como filas de referencia con su
        abscisa y descripción.

        Los registros a los que el técnico no les puso abscisa NO se pierden:
        se escriben en el punto de la secuencia donde van (anclados al registro
        anterior del archivo de campo), con la celda D vacía y resaltada en
        amarillo, y con las fórmulas que dependen del PK (C, Q, severidad)
        neutralizadas con IF — al escribir la abscisa, Excel las completa solo.
        `self.dcvg_sin_abscisa` cuenta esas filas; `self.dcvg_omitidos` solo
        cuenta lo que no cupo en la hoja, y `self.dcvg_filas` las escritas."""
        self.dcvg_omitidos = 0
        self.dcvg_sin_abscisa = 0
        self.dcvg_filas = 0
        if 'Inspección DCVG' not in self.wb.sheetnames:
            return
        ws = self.wb['Inspección DCVG']
        start = 8
        tope = next((r for r in range(start, ws.max_row + 1)
                     if ws.cell(row=r, column=3).value
                     and 'ELABOR' in str(ws.cell(row=r, column=3).value).upper()),
                    ws.max_row)
        capacidad = tope - start

        # (tipo, registro, ancla, pos, secuencia): los registros sin abscisa se
        # anclan al vecino del archivo de campo para caer en su punto del
        # recorrido (ver _anclar_sin_abscisa).
        filas = []
        seq = 0
        for tipo, items, clave in (('poste', postes, 'pk_m'),
                                   ('defecto', defectos, 'pk_m'),
                                   ('hallazgo', hallazgos, 'abscisa_val')):
            for ancla, pos, it in self._anclar_sin_abscisa(items or [], clave):
                filas.append((tipo, it, ancla, pos, seq))
                seq += 1
        # orden por abscisa; a igual abscisa, postes/defectos antes que hallazgos
        _orden = {'poste': 0, 'defecto': 1, 'hallazgo': 2}
        filas.sort(key=lambda t: (t[2], t[3], _orden[t[0]], t[4]))
        if len(filas) > capacidad:
            self.dcvg_omitidos += len(filas) - capacidad
            filas = filas[:capacidad]
        if not filas:
            return
        self.dcvg_filas = len(filas)

        # filas (1-based en Excel) de los postes que TIENEN pulso (ON y OFF):
        # el P/RE de cada defecto se interpola entre el pulso anterior y el
        # posterior, así que solo cuentan los postes con pulso real.
        fila_poste = [start + i for i, f in enumerate(filas)
                      if f[0] == 'poste' and f[1].get('on') is not None
                      and f[1].get('off') is not None]

        _COLSEV = {'AA': 19, 'CA': 20, 'CC': 21}   # S/T/U
        from openpyxl.utils import get_column_letter as _gcl

        for i, (tipo, r, _anc, _pos, _sq) in enumerate(filas):
            row = start + i
            absc = r.get('pk_m') if tipo != 'hallazgo' else r.get('abscisa_val')
            falta_absc = absc is None
            if falta_absc:
                self.dcvg_sin_abscisa += 1
            self._safe_write(ws, row, 1, i + 1)                              # A item
            self._safe_write(ws, row, 4, absc)                              # D abscisa
            if falta_absc:
                self._marcar_por_completar(ws, row, 4)
            if row > start:
                dist = f"D{row}-$D${start}"
                self._safe_write(ws, row, 3,                                 # C distancia
                                 f'=IF(D{row}="","",{dist})' if falta_absc
                                 else f"={dist}")
            self._safe_write(ws, row, 5, r.get('lat'))                       # E
            self._safe_write(ws, row, 6, r.get('lon'))                       # F
            if tipo == 'hallazgo':
                desc = corregir_campo(r.get('descripcion')
                                      or r.get('observaciones') or 'Hallazgo')
                self._safe_write(ws, row, 2, desc)                          # B referencia
                self._safe_write(ws, row, 24, desc)                        # X observaciones
            elif tipo == 'poste':
                self._safe_write(ws, row, 2, corregir_campo(r.get('tipo', 'Poste')))  # B
                self._safe_write(ws, row, 14, r.get('on'))                   # N ON
                self._safe_write(ws, row, 15, r.get('off'))                  # O OFF
                if r.get('on') is not None and r.get('off') is not None:
                    self._safe_write(ws, row, 16, f"=ABS(N{row}-O{row})")    # P pulso
                self._safe_write(ws, row, 24, corregir_campo(r.get('tipo', '')))  # X
            else:
                self._safe_write(ws, row, 2, "Defecto")                      # B
                self._safe_write(ws, row, 8, r.get('forma_n'))               # H 12
                self._safe_write(ws, row, 9, r.get('forma_e'))               # I 3
                self._safe_write(ws, row, 10, r.get('forma_s'))              # J 6
                self._safe_write(ws, row, 11, r.get('forma_o'))              # K 9
                car = r.get('caracter', '')
                self._safe_write(ws, row, 12, car)                           # L carácter
                self._safe_write(ws, row, 13, r.get('ol_re'))                # M OL/RE
                self._safe_write(ws, row, 18, r.get('profundidad'))          # R profundidad
                # P/RE (Q): pulso interpolado entre los postes que rodean
                pa = max([fp for fp in fila_poste if fp < row], default=None)
                ps = min([fp for fp in fila_poste if fp > row], default=None)
                tiene_pre = bool(pa or ps)
                if pa and ps:
                    interp = f"((P{ps}-P{pa})/(D{ps}-D{pa})*(D{row}-D{pa}))+P{pa}"
                elif pa or ps:
                    interp = f"P{pa or ps}"                                  # extremo
                else:
                    interp = None
                if interp:
                    # sin abscisa la interpolación daría #DIV/0!: se deja
                    # escrita pero en blanco hasta que se ponga el PK.
                    self._safe_write(ws, row, 17,
                                     f'=IF(D{row}="","",{interp})' if falta_absc
                                     else f"={interp}")
                # SEVERIDAD %IR = OL/RE ÷ P/RE (=M/Q). La celda S/T/U tiene
                # formato '0%', así que la fracción M/Q se muestra como
                # porcentaje. La clasificación compara contra fracciones
                # (0.15/0.35/0.60).
                col_sev = _COLSEV.get(car)
                if col_sev and r.get('ol_re') is not None and tiene_pre:
                    self._safe_write(ws, row, col_sev,
                                     f'=IF(Q{row}="","",M{row}/Q{row})'
                                     if falta_absc else f"=M{row}/Q{row}")
                    letra = _gcl(col_sev)
                    self._safe_write(ws, row, 22,
                        f'=IF({letra}{row}="","",IF({letra}{row}<=0.15,"Muy Pequeño",'
                        f'IF({letra}{row}<=0.35,"Pequeño",IF({letra}{row}<=0.6,'
                        f'"Mediano","Grande"))))')
                # W resistividad más cercana por abscisa (necesita el PK)
                if resistividades and not falta_absc:
                    cerc = min(resistividades,
                               key=lambda x: abs((x.get('pk_m') or 1e12) - absc))
                    partes = [f"{n}m {int(cerc[k])}" for n, k in
                              (("1", "r1"), ("2", "r2"), ("3", "r3"))
                              if cerc.get(k) is not None]
                    if partes:
                        self._safe_write(ws, row, 23, "\n".join(partes))     # W
                obs = r.get('comentarios') or "Defecto"
                self._safe_write(ws, row, 24, corregir_campo(obs))           # X

        # limpiar filas de datos sobrantes hasta el bloque de firmas
        for row in range(start + len(filas), tope):
            for c in range(1, 25):
                self._safe_write(ws, row, c, '')

    def fill_resistividad(self, resistividades: list):
        """Llena la hoja 'Resistividad' desde la fila 9 (fila 8 = encabezados).
        Las columnas ρ y la clasificación de corrosividad ya son fórmulas del
        template; aquí se escriben A=abscisa, B=sector, C/D=lat/lon,
        E=profundidad, F/H/J = Resistencia 1/2/3 m. Las que no traen abscisa
        se escriben igual (al final) con la celda A resaltada en amarillo."""
        if not resistividades or 'Resistividad' not in self.wb.sheetnames:
            return
        import re
        # Ordenar por abscisa ascendente (arriba->abajo). Las que no traen
        # abscisa quedan al final, en su orden original.
        resistividades = sorted(
            resistividades,
            key=lambda d: (d.get('pk_m') is None, d.get('pk_m') or 0))
        ws = self.wb['Resistividad']
        start = 9
        # última fila con fórmula ρ (modelo a replicar si hay más puntos)
        modelo = start
        for r in range(start, ws.max_row + 1):
            v = ws.cell(row=r, column=12).value
            if isinstance(v, str) and v.startswith('='):
                modelo = r
            else:
                break
        for i, d in enumerate(resistividades):
            row = start + i
            # si la fila no trae las fórmulas ρ (más puntos que el template),
            # copiarlas del modelo con la fila ajustada
            if not (isinstance(ws.cell(row=row, column=12).value, str)
                    and str(ws.cell(row=row, column=12).value).startswith('=')):
                self._copy_row_style(ws, modelo, row, 1, 24)
                for c in range(12, 25):
                    fv = ws.cell(row=modelo, column=c).value
                    if isinstance(fv, str) and fv.startswith('='):
                        ws.cell(row=row, column=c).value = re.sub(
                            r'(\$?[A-Z]+\$?)' + str(modelo) + r'\b',
                            lambda mm: mm.group(1) + str(row), fv)
            self._safe_write(ws, row, 1, d.get('pk_m'))            # A abscisa
            if d.get('pk_m') is None:
                # el técnico no la registró: el dato se conserva igual y la
                # celda queda resaltada para completarla
                self._marcar_por_completar(ws, row, 1)
            self._safe_write(ws, row, 2, corregir_campo(d.get('sector', '')))  # B
            self._safe_write(ws, row, 3, d.get('lat'))             # C
            self._safe_write(ws, row, 4, d.get('lon'))             # D
            self._safe_write(ws, row, 5, d.get('profundidad'))     # E
            self._safe_write(ws, row, 6, d.get('r1'))              # F R1
            self._safe_write(ws, row, 8, d.get('r2'))              # H R2
            self._safe_write(ws, row, 10, d.get('r3'))             # J R3

    # Números en letras para el texto de las observaciones ("tres (3)").
    _LETRAS = ['cero', 'una', 'dos', 'tres', 'cuatro', 'cinco', 'seis', 'siete',
               'ocho', 'nueve', 'diez', 'once', 'doce', 'trece', 'catorce',
               'quince', 'dieciséis', 'diecisiete', 'dieciocho', 'diecinueve',
               'veinte']

    @classmethod
    def _n_letras(cls, n):
        """'3' -> 'tres (3)'."""
        letra = cls._LETRAS[n] if 0 <= n < len(cls._LETRAS) else str(n)
        return f"{letra} ({n})"

    @staticmethod
    def _num_es(x, dec=2):
        """Número con coma decimal, como se escribe en el informe."""
        return f"{x:,.{dec}f}".replace(",", "@").replace(".", ",").replace("@", ".")

    # Corrosividad del suelo por resistividad (mismos cortes que las fórmulas
    # de la hoja Resistividad). Cada clase con su forma para el rango
    # ("desde X hasta Y") y para la lista ("a zonas X").
    _CORROSIVIDAD = [
        (500, "muy corrosivas", "muy corrosivas"),
        (1000, "corrosivas", "corrosivas"),
        (2000, "moderadamente corrosivas", "moderadamente corrosivas"),
        (10000, "medianamente corrosivas", "medianamente corrosivas"),
        (float('inf'), "corrosividad despreciable", "de corrosividad despreciable"),
    ]

    @classmethod
    def _clase_corrosividad(cls, rho):
        for tope, nombre, _zona in cls._CORROSIVIDAD:
            if rho <= tope:
                return nombre
        return cls._CORROSIVIDAD[-1][1]

    @classmethod
    def _cant_indicaciones(cls, n, total, articulo=True):
        """'las tres (3) indicaciones' / 'una (1) indicación' / 'dos (2)
        indicaciones', según si son todas y según el número."""
        if n == 1:
            return "una (1) indicación"
        art = "las " if (articulo and n == total and total > 1) else ""
        return f"{art}{cls._n_letras(n)} indicaciones"

    @staticmethod
    def _articulo_tramo(tramo):
        """'Ramal Armenia' -> 'el Ramal Armenia'; 'Armenia' -> 'Armenia'."""
        t = str(tramo or '').strip()
        primera = t.split()[0].lower() if t else ''
        if primera in ('ramal', 'troncal', 'loop', 'gasoducto', 'tramo'):
            return f"el {t}"
        return t

    def _texto_obs_dcvg(self, info, postes, defectos):
        """Redacta las observaciones de la gráfica DCVG con los datos de ESTA
        inspección: cuántas indicaciones, densidad por km, severidad %IR y
        carácter."""
        tramo = self._articulo_tramo(info.get('tramo')) or 'el tramo'
        n = len(defectos or [])
        if not n:
            return (f"En {tramo} no se identificaron indicaciones en la "
                    f"inspección DCVG realizada.")
        import db as _db
        sev = _db._severidad_dcvg(postes or [], defectos)
        # longitud inspeccionada: la de Datos Generales o la del recorrido
        try:
            km = float(info.get('longitud_km') or 0)
        except (TypeError, ValueError):
            km = 0
        if km <= 0:
            absc = [x.get('pk_m') for x in (list(postes or []) + list(defectos))
                    if x.get('pk_m') is not None]
            km = (max(absc) - min(absc)) / 1000 if len(absc) > 1 else 0
        verbo = "se identificó" if n == 1 else "se identificaron"
        partes = [f"En {tramo} {verbo} "
                  f"{self._cant_indicaciones(n, n, articulo=False)}"]
        if km > 0:
            partes[0] += (f", con una densidad de defectos de "
                          f"{self._num_es(n / km)} Indicaciones/Km")
        partes[0] += "."

        # severidad %IR por rango (mismos cortes que la clasificación)
        rangos = [("menor a 15%", lambda p: p <= 15),
                  ("entre 15% y 35%", lambda p: 15 < p <= 35),
                  ("entre 35% y 60%", lambda p: 35 < p <= 60),
                  ("mayor a 60%", lambda p: p > 60)]
        pct = [s.get('severidad_pct') for s in sev]
        frases = []
        for etq, cond in rangos:
            k = sum(1 for p in pct if p is not None and cond(p))
            if k:
                frases.append(f"{self._cant_indicaciones(k, n)} "
                              f"{'tiene' if k == 1 else 'tienen'} un índice de "
                              f"severidad IR {etq}")
        if frases:
            partes.append("En donde " + self._unir(frases) + ".")
        sin_sev = sum(1 for p in pct if p is None)
        if sin_sev:
            partes.append(f"Para {self._cant_indicaciones(sin_sev, n)} no fue "
                          f"posible calcular el índice de severidad IR.")

        # carácter de las indicaciones
        cars = {}
        for d in defectos:
            c = str(d.get('caracter') or '').strip().upper()
            if c:
                cars[c] = cars.get(c, 0) + 1
        if cars:
            det = [f"{self._cant_indicaciones(v, n)} "
                   f"{'es' if v == 1 else 'son'} de carácter {k}"
                   for k, v in sorted(cars.items())]
            partes.append("De lo anterior se puede observar que "
                          + self._unir(det) + ".")
        return "\n".join(partes)

    @staticmethod
    def _unir(items):
        """['a', 'b', 'c'] -> 'a, b y c'."""
        items = [i for i in items if i]
        if len(items) <= 1:
            return items[0] if items else ""
        return ", ".join(items[:-1]) + " y " + items[-1]

    def _texto_obs_resistividad(self, resistividades):
        """Redacta las observaciones de la gráfica de resistividad con las
        medidas de ESTA inspección (espaciamiento y grado de corrosividad)."""
        datos = [r for r in (resistividades or [])
                 if any(r.get(k) is not None for k in ('r1', 'r2', 'r3'))]
        if not datos:
            return ""
        # espaciamiento típico entre medidas
        absc = sorted(r['pk_m'] for r in datos if r.get('pk_m') is not None)
        pasos = [b - a for a, b in zip(absc, absc[1:]) if b > a]
        paso = int(round(sorted(pasos)[len(pasos) // 2] / 10.0) * 10) if pasos else 0

        # ρ de cada medida (mismas fórmulas de la hoja Resistividad)
        rhos = []
        for r in datos:
            r1, r2, r3 = r.get('r1'), r.get('r2'), r.get('r3')
            if r1 is not None:
                rhos.append(2 * math.pi * r1 * 100)
            if r1 is not None and r2 is not None and abs(r2 - r1) > 0:
                rhos.append(((r1 * r2) / abs(r2 - r1)) * 2 * math.pi * 100)
            if r2 is not None and r3 is not None and abs(r3 - r2) > 0:
                rhos.append(((r2 * r3) / abs(r3 - r2)) * 2 * math.pi * 100)
        if not rhos:
            return ""
        clases = [self._clase_corrosividad(v) for v in rhos]
        zona = {n: z for _t, n, z in self._CORROSIVIDAD}
        presentes = [n for _t, n, _z in self._CORROSIVIDAD if n in clases]
        p1 = (f"Se realizaron mediciones de resistividades"
              + (f" cada {paso} m de distancia" if paso else "")
              + " a 1, 2 y 3 metros de profundidad con el objetivo de estimar el "
                "grado de corrosividad del suelo en las ubicaciones de los "
                "defectos encontrados en la tubería.")
        if len(presentes) > 1:
            p1 += (f" Para las ubicaciones medidas, se encontró que la tubería "
                   f"se encuentra en zonas desde {presentes[-1]} hasta "
                   f"{presentes[0]} (Ver pestaña Datos).")
        else:
            p1 += (f" Para las ubicaciones medidas, se encontró que la tubería "
                   f"se encuentra en zonas {presentes[0]} (Ver pestaña Datos).")
        pcts = sorted(((clases.count(c) * 100.0 / len(clases), c) for c in presentes),
                      reverse=True)
        det = [f"{self._num_es(p, 1)}% a zonas {zona[c]}" for p, c in pcts[1:]]
        p2 = ("Basado en la totalidad de las mediciones para 1, 2 y 3 m de "
              f"profundidad se puede considerar que el {self._num_es(pcts[0][0], 1)}% "
              f"aproximadamente de los registros corresponden a zonas "
              f"{zona[pcts[0][1]]}"
              + (" y " + self._unir(det) if det else "") + ".")
        return p1 + "\n\n" + p2

    def fill_observaciones_dcvg(self, info, postes, defectos, resistividades=None):
        """Escribe las OBSERVACIONES de las dos gráficas del informe DCVG
        (celda F32 de 'GRAFICA DCVG' y de 'Gráfica Resistividad').

        La plantilla trae el texto de un informe de ejemplo; si no se
        reemplaza, el informe sale con las observaciones de otra inspección.
        Siempre se sobrescribe, aunque no haya datos."""
        for hoja, texto in (('GRAFICA DCVG',
                             self._texto_obs_dcvg(info or {}, postes, defectos)),
                            ('Gráfica Resistividad',
                             self._texto_obs_resistividad(resistividades))):
            if hoja in self.wb.sheetnames:
                self._safe_write(self.wb[hoja], 32, 6, texto)

    def fill_graficas_dcvg(self, n_inspeccion, n_resist):
        """Ajusta el rango de las series de datos de las 2 gráficas DCVG a las
        filas realmente escritas (Inspección DCVG y Resistividad). En la
        gráfica DCVG la severidad se grafica como PORCENTAJE: las líneas de
        criterio se ponen en fracción (0.15/0.35/0.60, no 15/35/60) y el eje Y
        se formatea como '0%', para que cuadren con el %IR (=M/Q) que es
        fracción."""
        import re
        from openpyxl.chart.data_source import NumFmt
        planes = [
            ('GRAFICA DCVG', 'Inspección DCVG', 8, n_inspeccion),
            ('Gráfica Resistividad', 'Resistividad', 9, n_resist),
        ]
        for hoja_g, hoja_dato, inicio, n in planes:
            if n <= 0 or hoja_g not in self.wb.sheetnames:
                continue
            ws = self.wb[hoja_g]
            if not getattr(ws, '_charts', None):
                continue
            last = inicio + n - 1
            for s in ws._charts[0].series:
                for ref in (s.xVal, s.yVal):
                    f = ref.numRef.f if (ref and ref.numRef) else None
                    if not f or hoja_dato not in f:
                        continue
                    ref.numRef.f = re.sub(r'(\$?[A-Z]+\$?)\d+:(\$?[A-Z]+\$?)\d+',
                                          lambda m: f"{m.group(1)}{inicio}:{m.group(2)}{last}", f)
                    if ref.numRef.numCache:
                        ref.numRef.numCache = None

            if hoja_g == 'GRAFICA DCVG':
                insp = self.wb['Inspección DCVG']
                d_ini = insp.cell(row=inicio, column=4).value
                d_fin = insp.cell(row=last, column=4).value
                # extremos de abscisa de las líneas de criterio (D39:D40)
                self._safe_write(ws, 39, 4, d_ini if d_ini is not None else 0)
                self._safe_write(ws, 40, 4, d_fin if d_fin is not None else 0)
                # criterios como fracción (%): 15%→0.15, 35%→0.35, 60%→0.60
                for rc in (39, 40):
                    self._safe_write(ws, rc, 5, 0.15)   # E
                    self._safe_write(ws, rc, 6, 0.35)   # F
                    self._safe_write(ws, rc, 7, 0.60)   # G
                try:
                    ws._charts[0].y_axis.numFmt = NumFmt(formatCode='0%',
                                                         sourceLinked=False)
                except Exception:
                    pass

    # Criterios de la plantilla que no coinciden con lo que escribe el informe.
    _CRITERIOS_RESUMEN = (('"A-A"', '"AA"'),        # el carácter se escribe 'AA'
                          ('"Mediana"', '"Mediano"'))  # la clasificación, 'Mediano'

    def ajustar_resumenes_dcvg(self, n_inspeccion, n_resist):
        """Ajusta las fórmulas de RESUMEN DE INDICACIONES y CLASIFICACIÓN
        RESISTIVIDAD (hoja Informe) a las filas realmente escritas.

        La plantilla trae los rangos a mano y descuadrados entre sí (V8:V37 en
        una celda y V8:V237 en la de al lado; U9:U25 junto a U9:U244/246/247),
        así que con inspecciones largas esas tablas contaban solo una parte de
        la data. De paso corrige dos criterios que nunca casaban: 'A-A' (el
        informe escribe 'AA') y 'Mediana' (escribe 'Mediano').
        """
        ws = self.ws_informe
        if ws is None:
            return
        fin = {"'Inspección DCVG'": 7 + max(int(n_inspeccion or 0), 1),
               "Resistividad": 8 + max(int(n_resist or 0), 1)}
        ini = {"'Inspección DCVG'": 8, "Resistividad": 9}
        patron = re.compile(
            r"('Inspección DCVG'|Resistividad)!"
            r"(\$?[A-Z]{1,3}\$?)(\d+):(\$?[A-Z]{1,3}\$?)(\d+)")

        def _reescribir(m):
            hoja = m.group(1)
            return (f"{hoja}!{m.group(2)}{ini[hoja]}:{m.group(4)}{fin[hoja]}")

        for fila in ws.iter_rows():
            for celda in fila:
                f = celda.value
                if not (isinstance(f, str) and f.startswith("=")):
                    continue
                if "Inspección DCVG" not in f and "Resistividad" not in f:
                    continue
                nueva = patron.sub(_reescribir, f)
                for viejo, bueno in self._CRITERIOS_RESUMEN:
                    nueva = nueva.replace(viejo, bueno)
                if nueva != f:
                    self._safe_write(ws, celda.row, celda.column, nueva)

    def fill_rangos_dcvg(self, postes, defectos, seg_m=5000, max_hojas=60):
        """Crea las hojas por rango (~5 km) del informe DCVG: cada una es una
        copia de 'GRAFICA DCVG' (mismo chart de %IR vs abscisa, leyendo
        'Inspección DCVG') con el eje X limitado a su segmento. Los segmentos
        cubren la extensión de abscisas del survey. Devuelve el nº de hojas."""
        import copy as _copy
        if 'GRAFICA DCVG' not in self.wb.sheetnames:
            return 0
        absc = [x['pk_m'] for x in (list(postes or []) + list(defectos or []))
                if x.get('pk_m') is not None]
        if not absc:
            return 0
        lo, hi = min(absc), max(absc)
        b0 = ((lo // seg_m) + 1) * seg_m if lo % seg_m else lo + seg_m
        # límites de cada segmento: primer parcial + bloques de seg_m
        limites = []
        if lo < b0:
            limites.append((lo, min(b0 - 1, hi)))
        b = b0
        while b <= hi and len(limites) < max_hojas:
            limites.append((b, min(b + seg_m, hi)))
            b += seg_m
        if not limites:
            limites = [(lo, hi)]

        src = self.wb['GRAFICA DCVG']
        src_merges = [str(m) for m in src.merged_cells.ranges]

        def _k(m):
            return f"K {int(m) // 1000:03d}+{int(m) % 1000:03d}"

        creadas = 0
        for ini, fin in limites:
            nombre = f"{_k(ini)} - {_k(fin)}"[:31]
            if nombre in self.wb.sheetnames:
                continue
            ws = self.wb.create_sheet(nombre)
            # copiar celdas/estilos (encabezado, tabla de criterios) del modelo
            for col, dim in src.column_dimensions.items():
                ws.column_dimensions[col].width = dim.width
            for row in src.iter_rows():
                for c in row:
                    if c.value is None and not c.has_style:
                        continue
                    tc = ws.cell(row=c.row, column=c.column, value=c.value)
                    if c.has_style:
                        tc.font = copy(c.font)
                        tc.border = copy(c.border)
                        tc.fill = copy(c.fill)
                        tc.number_format = c.number_format
                        tc.alignment = copy(c.alignment)
            for m in src_merges:
                try:
                    ws.merge_cells(m)
                except Exception:
                    pass
            # chart: copia con el eje X del segmento
            if getattr(src, '_charts', None):
                ch = _copy.deepcopy(src._charts[0])
                try:
                    ch.x_axis.scaling.min = ini
                    ch.x_axis.scaling.max = fin
                except Exception:
                    pass
                ws.add_chart(ch, "A9")
            creadas += 1
        return creadas

    def save(self, output_path: str):
        """Save the completed report"""
        self.wb.save(output_path)
        return output_path
