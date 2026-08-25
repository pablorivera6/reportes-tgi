"""Lectura de los archivos FastField de DCVG y Resistividades.

El export de FastField trae una hoja `Root` con la metadata y subformularios
(`subform_N`) con las filas repetibles. Para DCVG:
  - subform_5 = postes (potenciales de estructura)
  - subform_9 = defectos de recubrimiento
Para Resistividades:
  - subform_7 = medidas Wenner (1/2/3 m)

Las funciones son tolerantes: nombres de columna con espacios/tildes variables,
celdas vacías, y coordenadas "lat,lon" en un solo campo.
"""
import re

import openpyxl

# Etiqueta de abscisa que escribe el técnico: "5+760", "136+300", "PK 2+000".
_RE_PK = re.compile(r'(\d+)\s*\+\s*(\d+)')


def parse_pk(texto):
    """'5+760' -> 5760 (metros). None si no hay patрón K+M."""
    if texto is None:
        return None
    m = _RE_PK.search(str(texto))
    if not m:
        return None
    return int(m.group(1)) * 1000 + int(m.group(2))


def parse_coords(texto):
    """'4.5290577,-75.7461' -> (4.5290577, -75.7461). (None, None) si no aplica."""
    if texto is None:
        return (None, None)
    partes = str(texto).replace(";", ",").split(",")
    if len(partes) < 2:
        return (None, None)
    try:
        return (float(partes[0].strip()), float(partes[1].strip()))
    except (ValueError, TypeError):
        return (None, None)


def caracter_corto(txt):
    """'Catódico-Catódico' -> 'CC'; 'Catódico-Anódico' -> 'CA'; etc."""
    if not txt:
        return ""
    t = str(txt).lower()
    partes = re.split(r'[-/]', t)
    if len(partes) >= 2:
        a = "C" if "cat" in partes[0] else ("A" if "an" in partes[0] else "")
        b = "C" if "cat" in partes[1] else ("A" if "an" in partes[1] else "")
        if a and b:
            return a + b
    return ""


def clasificar_severidad(pct_ir):
    """Clasificación de severidad DCVG por %IR (umbrales PCC/TGI)."""
    if pct_ir is None:
        return ""
    v = abs(float(pct_ir))
    if v <= 15:
        return "Muy Pequeño"
    if v <= 35:
        return "Pequeño"
    if v <= 60:
        return "Mediano"
    return "Grande"


# ── Helpers de hoja/columna ───────────────────────────────────────────────────

def _norm(s):
    return re.sub(r'\s+', ' ', str(s or "").strip().lower())


def _sin_tildes(s):
    import unicodedata
    return ''.join(c for c in unicodedata.normalize('NFD', str(s or ""))
                   if unicodedata.category(c) != 'Mn')


def _es_defecto_dcvg(texto):
    """True si la fila de la data cruda es un defecto DCVG ('DCVG Anomaly',
    'DCVG Anómalo'…). Esos defectos llegan por FastField, que trae la medición
    completa; tomarlos también del logger los duplicaría en el informe."""
    return 'anomal' in _sin_tildes(texto).lower()


def _indice_columnas(fila_encabezados):
    """Mapa nombre_normalizado -> índice de columna (0-based)."""
    return {_norm(v): i for i, v in enumerate(fila_encabezados) if v is not None}


def _col(idx_map, *nombres):
    """Devuelve el índice de la primera columna cuyo nombre normalizado empiece
    por alguno de los dados; None si no está."""
    for n in nombres:
        nn = _norm(n)
        for nombre, i in idx_map.items():
            if nombre == nn or nombre.startswith(nn):
                return i
    return None


def _num(v):
    try:
        return float(v) if v not in (None, "") else None
    except (ValueError, TypeError):
        return None


def _filas(ws):
    return [r for r in ws.iter_rows(values_only=True) if r and any(c is not None for c in r)]


# ── Lectores ──────────────────────────────────────────────────────────────────

def leer_dcvg_fastfield(ruta):
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    meta = {"contratista": "", "fecha": "", "cliente": "", "tramo": "", "tecnico": ""}
    if "Root" in wb.sheetnames:
        rf = _filas(wb["Root"])
        if len(rf) >= 2:
            im = _indice_columnas(rf[0])
            fila = rf[1]

            def g(*n):
                i = _col(im, *n)
                return str(fila[i]).strip() if i is not None and fila[i] is not None else ""
            meta.update(contratista=g("Contratista"), fecha=g("Fecha"),
                        cliente=g("Cliente"), tramo=g("Troncal o ramal"),
                        tecnico=g("Técnico a cargo", "Tecnico a cargo"))

    postes = []
    if "subform_5" in wb.sheetnames:
        rf = _filas(wb["subform_5"])
        if len(rf) >= 2:
            im = _indice_columnas(rf[0])
            ci = {k: _col(im, *v) for k, v in {
                "tipo": ("Tipo de poste",), "pk": ("PK",), "on": ("ON",),
                "off": ("OFF",), "vac": ("Voltaje AC",), "res": ("Resistencia",),
                "coord": ("Coordenadas",)}.items()}
            for fila in rf[1:]:
                def v(k):
                    i = ci[k]
                    return fila[i] if i is not None and i < len(fila) else None
                lat, lon = parse_coords(v("coord"))
                postes.append({
                    "tipo": str(v("tipo") or "").strip(), "pk_m": parse_pk(v("pk")),
                    "on": _num(v("on")), "off": _num(v("off")), "vac": _num(v("vac")),
                    "resistencia": _num(v("res")), "lat": lat, "lon": lon})

    defectos = []
    if "subform_9" in wb.sheetnames:
        rf = _filas(wb["subform_9"])
        if len(rf) >= 2:
            im = _indice_columnas(rf[0])
            ci = {k: _col(im, *v) for k, v in {
                "sector": ("Sector",), "ubic": ("Ubicación", "Ubicacion"),
                "pk": ("PK del defecto",), "fn": ("Forma N",), "fs": ("Forma S",),
                "fe": ("Forma E",), "fo": ("Forma O",), "olre": ("OL/RE",),
                "prof": ("Profundidad",), "clas": ("Clasificación", "Clasificacion"),
                "reloj": ("Forma del defecto",), "com": ("Comentarios",),
                "car": ("Caracter de la indicación", "Caracter")}.items()}
            for fila in rf[1:]:
                def v(k):
                    i = ci[k]
                    return fila[i] if i is not None and i < len(fila) else None
                lat, lon = parse_coords(v("ubic"))
                defectos.append({
                    "sector": str(v("sector") or "").strip(), "lat": lat, "lon": lon,
                    "pk_m": parse_pk(v("pk")),
                    "forma_n": _num(v("fn")), "forma_s": _num(v("fs")),
                    "forma_e": _num(v("fe")), "forma_o": _num(v("fo")),
                    "ol_re": _num(v("olre")), "profundidad": _num(v("prof")),
                    "clasificacion_campo": str(v("clas") or "").strip(),
                    "posicion_reloj": str(v("reloj") or "").strip(),
                    "comentarios": str(v("com") or "").strip(),
                    "caracter": caracter_corto(v("car"))})
    wb.close()
    return {"meta": meta, "postes": postes, "defectos": defectos}


_RE_SOLO_RESIST = re.compile(
    r'^\s*(pk\s*\d+\s*\+?\s*\d*\s*)?toma\s+resistivida[d]?\s*$', re.IGNORECASE)


def tecnico_del_logger(rutas):
    """Técnico que aparece en la DATA CRUDA del equipo (hoja 'Survey Info',
    campo 'Technician Name'). Es el mismo dato que ya usa CIPS y el nombre que
    coincide con `Listado equipos TGI.xlsx`, a diferencia del que se escribe a
    mano en FastField."""
    if not rutas:
        return ""
    try:
        from cips_lrs import tecnico_de_archivos
        return tecnico_de_archivos(rutas) or ""
    except Exception:
        return ""


def info_desde_meta(meta, rutas_logger=None):
    """Campos de Datos Generales que trae la cabecera del FastField DCVG.

    El técnico ya diligenció en campo el tramo ('Troncal o ramal'), la fecha y
    el contratista: se llevan al informe para no tener que reescribirlos (y
    para que no queden vacíos el encabezado, la columna TRAMO de Hallazgos y la
    sigla del nombre del archivo).

    El **inspector** se toma de la data cruda del logger cuando está disponible
    (`rutas_logger`): ese es el nombre con el que se autollenan el serial y los
    equipos. El 'Técnico a cargo' del FastField queda como respaldo."""
    meta = meta or {}
    pares = (('tramo', 'tramo'), ('fecha', 'fecha'),
             ('contratista', 'contratista'), ('tecnico', 'inspector'))
    out = {destino: str(meta[origen]).strip()
           for origen, destino in pares
           if str(meta.get(origen) or '').strip()}
    del_logger = tecnico_del_logger(rutas_logger)
    if del_logger:
        out['inspector'] = del_logger
    return out


def leer_dcvg_fastfield_varios(rutas):
    """Combina varios FastField DCVG en un solo dict (postes/defectos unidos;
    meta del primero que traiga técnico)."""
    postes, defectos, meta = [], [], {}
    for r in rutas:
        d = leer_dcvg_fastfield(r)
        postes += d["postes"]
        defectos += d["defectos"]
        if not meta.get("tecnico") and d["meta"].get("tecnico"):
            meta = d["meta"]
        elif not meta:
            meta = d["meta"]
    return {"meta": meta, "postes": postes, "defectos": defectos}


def leer_resistividades_fastfield_varios(rutas):
    out = []
    for r in rutas:
        out += leer_resistividades_fastfield(r)
    return out


def leer_hallazgos_logger_varios(rutas):
    out = []
    for r in rutas:
        out += leer_hallazgos_logger(r)
    return out


def leer_hallazgos_logger(ruta):
    """Hallazgos del informe DCVG desde la data cruda del logger (hoja DCP
    Data): filas con comentario de campo (cruces, tramos enmontados, saltos,
    mallas, válvulas…). abscisa = Station No; GPS tomado de la hoja Survey Data
    por Station. Excluye los comentarios de carácter (Cathodic/Anodic) y los de
    solo 'toma resistividad'. Excluye TAMBIÉN las filas de defecto DCVG
    ('DCVG Anomaly'/'DCVG Anómalo'): esos defectos entran por FastField y
    tomarlos aquí los duplicaría. Devuelve dicts listos para cips_a_hallazgos."""
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    gps = {}
    if "Survey Data" in wb.sheetnames:
        sf = _filas(wb["Survey Data"])
        if sf:
            im = _indice_columnas(sf[0])
            si = _col(im, "Station No"); la = _col(im, "Latitude"); lo = _col(im, "Longitude")
            for r in sf[1:]:
                if si is not None and si < len(r) and r[si] is not None:
                    try:
                        st = int(float(r[si]))
                    except (ValueError, TypeError):
                        continue
                    lat = r[la] if la is not None and la < len(r) else None
                    lon = r[lo] if lo is not None and lo < len(r) else None
                    if st not in gps and lat is not None:
                        gps[st] = (_num(lat), _num(lon))
    salida = []
    if "DCP Data" in wb.sheetnames:
        df = _filas(wb["DCP Data"])
        if df:
            im = _indice_columnas(df[0])
            si = _col(im, "Station No"); ci = _col(im, "Comments")
            # columna del tipo de fila: 'Highway', 'Flag', 'DCVG Anomaly'…
            fi = _col(im, "DCP/Feature", "Feature", "Tipo")
            for r in df[1:]:
                tipo_fila = r[fi] if fi is not None and fi < len(r) else None
                if _es_defecto_dcvg(tipo_fila):
                    continue          # el defecto ya viene del FastField
                com = r[ci] if ci is not None and ci < len(r) else None
                com = str(com or "").strip()
                if not com:
                    continue
                if _es_defecto_dcvg(com):
                    continue          # por si el tipo va dentro del comentario
                low = com.lower()
                if low in ("cathodic/cathodic", "cathodic/anodic", "anodic/anodic",
                           "anodic/cathodic"):
                    continue
                if _RE_SOLO_RESIST.match(com):
                    continue
                com = _corregir_hallazgo(com)
                if not com:
                    continue
                try:
                    st = int(float(r[si])) if (si is not None and r[si] is not None) else None
                except (ValueError, TypeError):
                    st = None
                lat, lon = gps.get(st, (None, None))
                salida.append({"abscisa_val": st, "observaciones": com,
                               "referencia": com, "lat": lat, "lon": lon})
    wb.close()
    return salida


def _corregir_hallazgo(texto):
    """Repara mojibake, quita el ruido de 'toma resistividad' de comentarios
    combinados y normaliza. Devuelve '' si no queda contenido útil."""
    from ortografia import reparar_texto
    t = reparar_texto(texto)
    t = re.sub(r'\btoma\s+resistivida[d]?\b', '', t, flags=re.IGNORECASE)
    t = re.sub(r'\s{2,}', ' ', t).strip(' ,;-')
    return t


def leer_resistividades_fastfield(ruta):
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    out = []
    if "subform_7" in wb.sheetnames:
        rf = _filas(wb["subform_7"])
        if len(rf) >= 2:
            im = _indice_columnas(rf[0])
            ci = {k: _col(im, *v) for k, v in {
                "pk": ("PK",), "sector": ("Sector",), "prof": ("Profundidad",),
                "ubic": ("Ubicación", "Ubicacion"),
                "r1": ("Resistencia 1 metro",), "r2": ("Resistencia 2 metros",),
                "r3": ("Resistencia 3 metros",)}.items()}
            for fila in rf[1:]:
                def v(k):
                    i = ci[k]
                    return fila[i] if i is not None and i < len(fila) else None
                lat, lon = parse_coords(v("ubic"))
                out.append({
                    "pk_m": parse_pk(v("pk")), "sector": str(v("sector") or "").strip(),
                    "profundidad": _num(v("prof")), "lat": lat, "lon": lon,
                    "r1": _num(v("r1")), "r2": _num(v("r2")), "r3": _num(v("r3"))})
    wb.close()
    return out
